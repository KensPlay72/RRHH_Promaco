from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db import models
from django.contrib.auth.hashers import make_password
from django.contrib.auth.hashers import check_password
from django.utils import timezone
from django.shortcuts import redirect
from django.contrib.auth import logout
import json
import os
import random
import string
from django.conf import settings
from django.http import HttpResponse
from pathlib import Path
import mimetypes
from django.db.models import Q
import pandas as pd
from dateutil import parser
import uuid
from datetime import datetime, timedelta
from openpyxl import load_workbook
from django.views import View
from num2words import num2words
from django.db.models.functions import TruncDate
from django.db.models import Avg
from django.db.models import Count
from datetime import date
import calendar
from .models import *
from django.contrib.auth.models import Permission
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from apscheduler.schedulers.background import BackgroundScheduler


#--------SIN ACCESO VIEW--------#
def sin_acceso_view(request):
    return render(request, 'errors/sinacceso.html')

#--------USUARIOS VIEW--------#
@csrf_exempt
@login_required(login_url='/login/') 
def users_view(request):
    if request.method == 'POST':
        if not request.user.has_perm('auth.add_user'):
            return JsonResponse({'error': 'No tienes permisos para crear usuarios'}, status=403)
        try:
            data = json.loads(request.body)

            username = data.get('username')
            password = data.get('password')
            first_name = data.get('nombreusuario')
            last_name = data.get('apellidousuario')
            email = data.get('email')
            is_active = data.get('estado')
            permisos_ids = data.get('permisos', [])

            if User.objects.filter(username=username).exists():
                return JsonResponse({'error': 'El usuario ya existe'}, status=400)
            
            user = User.objects.create(
                username=username,
                password=make_password(password),
                first_name=first_name,
                last_name=last_name,
                email=email,
                is_active=is_active
            )

            permisos = Permission.objects.filter(id__in=permisos_ids)
            user.user_permissions.set(permisos)

            return JsonResponse({'message': 'Usuario creado correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'GET' and 'user_id' not in request.GET:

        if not request.user.has_perm('auth.view_user'):
            return redirect('sinacceso')

        search = request.GET.get('search', '')
        estado = request.GET.get('estado', '')

        permisos = Permission.objects.filter(
            Q(content_type_id=4) | Q(content_type_id__gt=6)
        )

        permisos_agrupados = {}
        for permiso in permisos:
            modelo = permiso.content_type.model
            if modelo not in permisos_agrupados:
                permisos_agrupados[modelo] = []
            permisos_agrupados[modelo].append(permiso.id)

        usuarios = User.objects.filter(is_superuser=False)

        if search:
            usuarios = usuarios.filter(
                Q(id__icontains=search) | 
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )

        if estado:
            usuarios = usuarios.filter(activo__icontains=estado)

        paginator = Paginator(usuarios, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        total_pages = paginator.num_pages
        current_page = page_obj.number

        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'usuarios': page_obj,
            'permisos_agrupados': permisos_agrupados,
            'search': search,
            'estado': estado,
            'page_range': page_range,
        }

        return render(request, 'users.html', context)

    elif request.method == 'GET' and 'user_id' in request.GET:
        user_id = request.GET.get('user_id')
        try:
            user = User.objects.get(id=user_id)
            permisos = user.user_permissions.all()

            return JsonResponse({
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                    'is_active': user.is_active,
                    'permisos': [
                        {'id': permiso.id, 'nombre': permiso.name, 'codename': permiso.codename}
                        for permiso in permisos
                    ] 
                },
            })

        except User.DoesNotExist:
            return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
        
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            user_id = data.get('id')
            username = data.get('username')
            password = data.get('password')
            first_name = data.get('nombreusuario')
            last_name = data.get('apellidousuario')
            email = data.get('email')
            is_active = data.get('estado')
            permisos_ids = data.get('permisos', [])

            user = User.objects.get(id=user_id)
            user.username = username
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.is_active = is_active

            if User.objects.filter(username=username).exclude(id=user_id).exists():
                return JsonResponse({'error': 'El usuario ya existe'}, status=400)
            user.save()

            if isinstance(permisos_ids, list):
                permisos = Permission.objects.filter(id__in=permisos_ids)
                user.user_permissions.set(permisos)

            return JsonResponse({'message': 'Usuario actualizado correctamente'}, status=200)
        
        except User.DoesNotExist:
            return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
@login_required(login_url='/login/') 
def resetpassword(request, id):
    if request.method == 'PUT':
        try:
            if request.user.has_perm('Reclutamiento.res_password'):
                return JsonResponse({'error': 'No tienes permisos para restablecer la contraseña'}, status=403)
            
            data = json.loads(request.body)

            password = data.get('password')

            user = User.objects.get(id=id)
            user.password = make_password(password)
            user.save()

            return JsonResponse({'message': 'Contraseña restablecida correctamente'}, status=200)
        
        except User.DoesNotExist:
            return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

#--------LOGIN VIEW--------#
def login_view(request):
    if request.method == 'GET':
            return render(request, 'login.html')
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            username = data.get('username')
            password = data.get('password')

            if not username or not password:
                return JsonResponse({'error': 'Usuario y contraseña son requeridos'}, status=400)
            
            user = User.objects.filter(username=username).first()

            if user is None:
                return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
            
            if user.is_active == False:
                return JsonResponse({'error': 'Usuario inactivo'}, status=400)
            
            if not check_password(password, user.password):
                return JsonResponse({'error': 'Contraseña incorrecta'}, status=400)
            
            if password == '12345678':
                return JsonResponse({'showModal': True}, status=200)
            
            login(request, user)
                        
            return JsonResponse({'redirect': '/Dashboard/'}, status=200)

        except Exception as e:
            return JsonResponse({'error': f'Error inesperado: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)
           
@csrf_exempt
def change_password_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            username = data.get('username')
            password = data.get('password')
            password_confirmation = data.get('password_confirmation')

            user = User.objects.get(username=username)

            if user is None:
                return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
            
            if password != password_confirmation:
                return JsonResponse({'error': 'La contraseña no coinciden'}, status=400)

            user.password = make_password(password)
            user.save()

            return JsonResponse({'message': 'Contraseña actualizada correctamente'}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

#--------DASHBOARD VIEW--------#
@login_required(login_url='/login/') 
def dashboard_view(request):
    # Filtros
    empresa = request.GET.get('empresa')
    año = request.GET.get('año')
    mes_corte = request.GET.get('mes_corte')
    sucursal = request.GET.get('sucursal')

    # Base QuerySet
    plazas = ControlDePlazas.objects.all()

    # Aplicar filtros si existen
    if empresa:
        plazas = plazas.filter(empresa__nombre_empresa=empresa)
    if año:
        plazas = plazas.filter(año=año)
    if mes_corte:
        plazas = plazas.filter(mes_corte=mes_corte)
    if sucursal:
        plazas = plazas.filter(sucursal__nombre_sucursal=sucursal)

    # Conteos después de aplicar los filtros
    canceladas_count = plazas.filter(estatus='CANCELADA').count()
    en_proceso_count = plazas.filter(estatus='EN PROCESO').count()
    cerradas_count = plazas.filter(estatus='CERRADA').count()

    # Promedio de tiempo_efectivo_cobertura
    promedio_tiempo_efectivo_cobertura = plazas.filter(estatus='CERRADA').aggregate(Avg('tiempo_efectivo_cobertura'))['tiempo_efectivo_cobertura__avg']

    # Conteo de plazas cerradas por género
    genero_count = plazas.filter(estatus='CERRADA').aggregate(
        total_masculino=Count('id', filter=Q(genero='M')),
        total_femenino=Count('id', filter=Q(genero='F'))
    )

    # Promedio de edades
    promedio_edad = plazas.aggregate(Avg('edad'))['edad__avg']

    # Obtener empresas, años, meses de corte, y sucursales únicos para los selectores
    empresas = ControlDePlazas.objects.values('empresa__nombre_empresa').distinct()
    años = ControlDePlazas.objects.values('año').distinct()
    meses_corte = ControlDePlazas.objects.values('mes_corte').distinct()
    sucursales = ControlDePlazas.objects.values('sucursal__nombre_sucursal').distinct()

    # Consultar el estado de pago en cesantías
    cesantias_count = Cesantias.objects.values('estado_pago').annotate(total=Count('id')).filter(estado_pago__in=['AUTORIZADO', 'CALCULADO', 'PAGADO'])

    # Crear un diccionario para los resultados de cesantías
    estado_pago_counts = {estado['estado_pago']: estado['total'] for estado in cesantias_count}

    # Nuevo conteo para BolsaEmpleos basado en el estado
    bolsa_estado_count = BolsaEmpleos.objects.values('estado').annotate(total=Count('id')).filter(estado__in=['CONTACTADO', 'DESCARTADO', 'ENTREVISTADO', 'REGISTRADO'])
    
    # Crear un diccionario para los resultados de BolsaEmpleos
    bolsa_estado_counts = {estado['estado']: estado['total'] for estado in bolsa_estado_count}

    # Obtener la fecha actual
    fecha_actual = date.today()

    # Pasar los datos al contexto
    context = {
        'canceladas_count': canceladas_count,
        'en_proceso_count': en_proceso_count,
        'cerradas_count': cerradas_count,
        'empresas': empresas,
        'años': años,
        'meses_corte': meses_corte,
        'sucursales': sucursales,
        'promedio_tiempo_efectivo_cobertura': promedio_tiempo_efectivo_cobertura,
        'total_masculino': genero_count['total_masculino'],
        'total_femenino': genero_count['total_femenino'],
        'promedio_edad': promedio_edad,
        'fecha_actual': fecha_actual,
        'estado_pago_labels': list(estado_pago_counts.keys()),  # Etiquetas para el gráfico de cesantías
        'estado_pago_values': list(estado_pago_counts.values()),  # Valores para el gráfico de cesantías
        'bolsa_estado_labels': list(bolsa_estado_counts.keys()),  # Etiquetas para el nuevo gráfico de BolsaEmpleos
        'bolsa_estado_values': list(bolsa_estado_counts.values()),  # Valores para el nuevo gráfico de BolsaEmpleos
    }
    
    return render(request, 'dashboard.html', context)

def logout_view(request):
    logout(request)
    return redirect('login')

#--------EMPRESAS VIEW--------#
@csrf_exempt
@login_required(login_url='/login/')
def empresas_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_empresas'):
                return JsonResponse({'error': 'No tienes permisos para crear sucursales'}, status=403)
    
            data = json.loads(request.body)
            nombre_empresa = data.get('nombre_empresa')
            estado = data.get('estado') 

            if not nombre_empresa:
                return JsonResponse({'success': False, 'message': 'El campo nombre_empresa es obligatorio.'}, status=400)

            if Empresas.objects.filter(nombre_empresa=nombre_empresa).exists():
                return JsonResponse({'success': False, 'message': 'La empresa ya está registrada.'}, status=409)

            empresa = Empresas.objects.create(
                nombre_empresa=nombre_empresa,
                activo=estado,
                usuario_creo = request.user
            )
            
            return JsonResponse({'message': 'Sucursal registrada correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        if not request.user.has_perm('Reclutamiento.change_empresas'):
            return JsonResponse({'error': 'No tienes permisos para actualizas'}, status=403)
        
        try:
            data = json.loads(request.body)

            id = data.get('id')
            nombre_empresa = data.get('nombre_empresa')
            activo = data.get('estado')

            if Empresas.objects.filter(nombre_empresa=nombre_empresa).exclude(id=id).exists():
                return JsonResponse({'error': 'La empresa ya esta registrada'})
            
            empresa = Empresas.objects.get(id=id)
            empresa.nombre_empresa = nombre_empresa
            empresa.activo = activo
            empresa.usuario_modifico = request.user
            empresa.save()

            return JsonResponse({'message': 'Empresa actualizada correctamente'}, status=200)
        
        except Empresas.DoesNotExist:
            return JsonResponse({'error':'Empresa no encontrada'}, status=403)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_empresas'):
            return redirect('sinacceso')
        empresas_id = request.GET.get('empresas_id')

        if empresas_id:
            try:
                empresa = Empresas.objects.get(id=empresas_id)

                return JsonResponse({
                    'empresas':{
                        'id': empresa.id,
                        'nombre_empresa': empresa.nombre_empresa,
                        'activo': empresa.activo
                    }
                })
            except Empresas.DoesNotExist:
                return JsonResponse({'error': 'Empresa no encontrada'}, status=404)
            
        else:
            search = request.GET.get('search', '')  
            estado = request.GET.get('estado', '') 
            empresas = Empresas.objects.all()

            if search:
                empresas = empresas.filter(
                    models.Q(id__icontains=search) |  
                    models.Q(nombre_empresa__icontains=search)
                )

            if estado:
                empresas = empresas.filter(activo__icontains=estado) 

            paginator = Paginator(empresas, 10) 
            page_number = request.GET.get('page') 
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1) 

            context = {
                'empresas': page_obj,  
                'search': search, 
                'estado': estado,  
                'page_range': page_range, 
            }
            return render(request, 'modulos/empresas.html', context)
    
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------SURCURSALES VIEW--------#
@csrf_exempt
@login_required(login_url='/login/')
def sucursales_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_sucursal'):
                return JsonResponse({'error': 'No tienes permisos para crear sucursales'}, status=403)
            
            data = json.loads(request.body)
            nombre_sucursal = data.get('nombre_sucursal')
            estado = data.get('estado') 

            if not nombre_sucursal:
                return JsonResponse({'success': False, 'message': 'El campo nombre sucursal es obligatorio.'}, status=400)

            if Sucursal.objects.filter(nombre_sucursal=nombre_sucursal).exists():
                return JsonResponse({'error': 'La sucursal ya está registrada'}, status=409)

            sucursal = Sucursal.objects.create(
                nombre_sucursal=nombre_sucursal,
                activo=estado,
                usuario_creo=request.user
            )
            
            return JsonResponse({'message': 'Sucursal registrada correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        if not request.user.has_perm('Reclutamiento.change_sucursal'):
            return JsonResponse({'error': 'No tienes permisos para actualizar sucursales'}, status=403)
        
        try:
            data = json.loads(request.body)

            sucursal_id = data.get('id')
            nombre_sucursal = data.get('nombre_sucursal')
            estado = data.get('estado')

            if Sucursal.objects.filter(nombre_sucursal=nombre_sucursal).exclude(id=sucursal_id).exists():
                return JsonResponse({'error': 'La sucursal ya está registrada'}, status=409)
            
            sucursal = Sucursal.objects.get(id=sucursal_id)
            sucursal.nombre_sucursal = nombre_sucursal
            sucursal.activo = estado
            sucursal.usuario_modifico = request.user
            sucursal.save()

            return JsonResponse({'message': 'Sucursal actualizada correctamente'}, status=200)
        
        except Sucursal.DoesNotExist:
            return JsonResponse({'error': 'Sucursal no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_sucursal'):
            return redirect('sinacceso')
        sucursal_id = request.GET.get('sucursal_id') 

        if sucursal_id:
            try:
                sucursal = Sucursal.objects.get(id=sucursal_id)

                return JsonResponse({
                    'sucursales': {
                        'id': sucursal.id,
                        'nombre_sucursal': sucursal.nombre_sucursal,
                        'activo': sucursal.activo
                    }
                })

            except Sucursal.DoesNotExist:
                return JsonResponse({'error': 'Sucursal no encontrada'}, status=404)
        
        else: 
            search = request.GET.get('search', '')  
            estado = request.GET.get('estado', '')
            sucursales = Sucursal.objects.all()

            if search:
                sucursales = sucursales.filter(
                    models.Q(id__icontains=search) |  
                    models.Q(nombre_sucursal__icontains=search)
                )

            if estado:
                sucursales = sucursales.filter(activo__icontains=estado) 

            paginator = Paginator(sucursales, 10) 
            page_number = request.GET.get('page') 
            page_obj = paginator.get_page(page_number)  

            total_pages = paginator.num_pages
            current_page = page_obj.number

            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'sucursales': page_obj,  
                'search': search, 
                'estado': estado, 
                'page_range': page_range,
            }

            return render(request, 'modulos/sucursales.html', context)

    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------UNIDAD DE NEGOCIO VIEW--------#
@csrf_exempt
@login_required(login_url='/login/')
def unidad_negocio_view(request):
    if request.method == 'POST':
        if not request.user.has_perm('Reclutamiento.add_unidad_negocio'):
            return JsonResponse({'error': 'No tienes permisos para registrar sucursales'}, status=403)
        
        try:
            data = json.loads(request.body)
            nombre_unidad_de_negocio = data.get('nombre_unidad_negocio')
            activo = data.get('estado')  

            if not nombre_unidad_de_negocio:
                return JsonResponse({'success': False, 'message': 'El campo nombre unidad de negocio es obligatorio.'}, status=400)

            if Unidad_Negocio.objects.filter(nombre_unidad_de_negocio=nombre_unidad_de_negocio).exists():
                return JsonResponse({'error': 'La sucursal ya está registrada'}, status=409)

            unidad_negocio = Unidad_Negocio.objects.create(
                nombre_unidad_de_negocio=nombre_unidad_de_negocio,
                activo=activo,
                usuario_creo = request.user
            )
            
            return JsonResponse({'success': True, 'message': 'Unidad de negocio creada exitosamente'}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        

    elif request.method == 'PUT':
        if not request.user.has_perm('Reclutamiento.change_unidad_negocio'):
            return JsonResponse({'error': 'No tienes permisos para actualizar sucursales'}, status=403)
        
        try:
            data = json.loads(request.body)

            unidad_id = data.get('id')
            nombre_unidad_de_negocio = data.get('nombre_unidad_negocio')
            activo = data.get('estado')

            if Unidad_Negocio.objects.filter(nombre_unidad_de_negocio = nombre_unidad_de_negocio).exclude(id=unidad_id).exists():
                return JsonResponse({'error': 'La unidad de negocio ya esta registrada'}, status=409)
            
            unidad = Unidad_Negocio.objects.get(id=unidad_id)
            unidad.nombre_unidad_de_negocio = nombre_unidad_de_negocio
            unidad.activo = activo
            unidad.usuario_modifico = request.user
            unidad.save()

            return JsonResponse({'message': 'Unidad de negocio actualizada correctamente'}, status=200)
        
        except Unidad_Negocio.DoesNotExist:
            return JsonResponse({'error': 'Unidad de negocio no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_unidad_negocio'):
            return redirect('sinacceso')
        unidad_id = request.GET.get('unidad_id')

        if unidad_id:
            try:
                unidad_negocio = Unidad_Negocio.objects.get(id=unidad_id)

                return JsonResponse({
                    'unidad': {
                        'id': unidad_negocio.id,
                        'nombre_unidad_de_negocio': unidad_negocio.nombre_unidad_de_negocio,
                        'activo': unidad_negocio.activo
                    }
                })
            except Unidad_Negocio.DoesNotExist:
                return JsonResponse({'error': 'Unidad de negocio no encontrada'}, status=404)
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=400)
        
        else:
            if not request.user.has_perm('Reclutamiento.view_unidad_negocio'):
                return redirect('sinacceso')
            
            search = request.GET.get('search', '')  
            estado = request.GET.get('estado', '')
            unidades_negocio = Unidad_Negocio.objects.all()

            if search:
                unidades_negocio = unidades_negocio.filter(
                    models.Q(id__icontains=search) |  
                    models.Q(nombre_unidad_de_negocio__icontains=search)
                )

            if estado:
                unidades_negocio = unidades_negocio.filter(activo__icontains=estado)

            paginator = Paginator(unidades_negocio, 10)
            page_number = request.GET.get('page') 
            page_obj = paginator.get_page(page_number) 

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'unidades_negocio': page_obj,  
                'search': search,  
                'estado': estado,
                'page_range': page_range, 
            }
            return render(request, 'modulos/unidad_negocio.html', context)
        
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------DEPARTAMENTOS VIEW--------#
@csrf_exempt
@login_required(login_url='/login/')
def departamentos_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_departamento'):
                return JsonResponse({'error': 'No tienes permisos para crear departamentos'}, status=403)
            
            data = json.loads(request.body)
            nombre_departamento = data.get('nombre_departamento')
            estado = data.get('estado')

            if not nombre_departamento:
                return JsonResponse({'success': False, 'message': 'El campo nombre departamento es obligatorio.'}, status=400)
            
            if Departamento.objects.filter(nombre_departamento=nombre_departamento).exists():
                return JsonResponse({'error': 'El departamento ya está registrado'}, status=409)
            
            departamento = Departamento.objects.create(
                nombre_departamento = nombre_departamento,
                activo = estado,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Departamento registrado correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        if not request.user.has_perm('Reclutamiento.change_departamento'):
            return JsonResponse({'error': 'No tienes permisos para actualizar Departamentos'})
        
        try:
            
            data = json.loads(request.body)
            id = data.get('id')
            nombre_departamento = data.get('nombre_departamento')
            activo = data.get('estado')

            if Departamento.objects.filter(nombre_departamento=nombre_departamento).exclude(id=id).exists():
                return JsonResponse({'error': 'El departamento ya está registrado'}, status=409)
            
            departamento = Departamento.objects.get(id=id)
            departamento.nombre_departamento = nombre_departamento
            departamento.activo = activo
            departamento.usuario_modifico = request.user
            departamento.save()

            return JsonResponse({'message': 'Departamento actualizado correctamente'}, status=200)
        
        except Departamento.DoesNotExist:
            return JsonResponse({'error': 'Departamento no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_departamento'):
            return redirect('sinacceso')
        departamento_id = request.GET.get('departamento_id')

        if departamento_id:
            try:
                departamento = Departamento.objects.get(id=departamento_id)

                return JsonResponse({
                    'departamentos':{
                        'id': departamento.id,
                        'nombre_departamento': departamento.nombre_departamento,
                        'activo': departamento.activo
                    }
                })
            except Departamento.DoesNotExist:
                return JsonResponse({'error':'Departamento no encontrado'}, status=400)
            
        else:
            if not request.user.has_perm('Reclutamiento.view_departamento'):
                return redirect('sinacceso')
            
            search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  
        departamentos = Departamento.objects.all()

        if search:
            departamentos = departamentos.filter(
                models.Q(id__icontains=search) |  
                models.Q(nombre_departamento__icontains=search)
            )

        if estado:
            departamentos = departamentos.filter(activo__icontains=estado)  

        paginator = Paginator(departamentos, 10)
        page_number = request.GET.get('page')  
        page_obj = paginator.get_page(page_number)  

        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'departamentos': page_obj,  
            'search': search,  
            'estado': estado,
            'page_range': page_range,
        }
        return render(request, 'modulos/departamentos.html', context)
    
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------CIUDADES VIEW--------#
@csrf_exempt
@login_required(login_url='/login/')
def ciudades_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_ciudades'):
                return JsonResponse({'error': 'No tienes permisos para registrar Ciudades'})
            
            data = json.loads(request.body)
            nombre_ciudades = data.get('nombre_ciudades')
            activo = data.get('estado')

            if not nombre_ciudades:
                return JsonResponse({'success': False, 'message': 'El campo nombre Ciudades es obligatorio.'}, status=400)
            
            if Ciudades.objects.filter(nombre_ciudades=nombre_ciudades).exists():
                return JsonResponse({'error': 'La Ciudad ya está registrada'}, status=409)
            
            ciudades = Ciudades.objects.create(
                nombre_ciudades = nombre_ciudades,
                activo = activo,
                usuario_creo = request.user
            )
            return JsonResponse({'message': 'Ciudad registrada correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif request.method == 'PUT':
        if not request.user.has_perm('Reclutamiento.change_ciudades'):
            return JsonResponse({'error': 'No tienes permisos para actualizar Ciudades'})
        try: 
            data = json.loads(request.body)

            ciudades_id = data.get('id')
            nombre_ciudades = data.get('nombre_ciudades')
            activo = data.get('estado')

            if Ciudades.objects.filter(nombre_ciudades=nombre_ciudades).exclude(id=ciudades_id).exists():
                return JsonResponse({'error': 'La ciudad ya está registrada'}, status=409)
            
            ciudades = Ciudades.objects.get(id=ciudades_id)
            ciudades.nombre_ciudades = nombre_ciudades
            ciudades.activo = activo
            ciudades.usuario_modifico = request.user
            ciudades.save()

            return JsonResponse({'message': 'Ciudad actualizada correctamente'}, status=200)
        
        except Ciudades.DoesNotExist:
            return JsonResponse({'error': 'Ciudad no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
  
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_ciudades'):
            return redirect('sinacceso')
        
        ciudades_id = request.GET.get('ciudades_id')

        if ciudades_id:
            try:
                ciudades = Ciudades.objects.get(id=ciudades_id)

                return JsonResponse({
                    'ciudades':{
                        'id': ciudades.id,
                        'nombre_ciudades': ciudades.nombre_ciudades,
                        'activo': ciudades.activo
                    }
                })
            
            except Ciudades.DoesNotExist:
                return JsonResponse({'error': 'Ciudad no encontrada'}, status=404)
            
        else: 
            search = request.GET.get('search', '') 
            estado = request.GET.get('estado', '') 
            ciudades = Ciudades.objects.all()

            if search:
                ciudades = ciudades.filter(
                    models.Q(id__icontains=search) |
                    models.Q(nombre_ciudades__icontains=search)
                )

            if estado:
                ciudades = ciudades.filter(activo__icontains=estado)

            paginator = Paginator(ciudades, 10) 
            page_number = request.GET.get('page') 
            page_obj = paginator.get_page(page_number) 

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'ciudades': page_obj,
                'search': search,
                'estado': estado,
                'page_range': page_range,
            }
        return render(request, 'modulos/ciudades.html', context)
    
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------PUESTOS VIEW--------#
@csrf_exempt
@login_required(login_url='/login/')
def puestos_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_puestos'):
                return JsonResponse({'error': 'No tienes permisos para registrar puestos'})
            
            data = json.loads(request.body)

            nombre_puestos = data.get('nombre_puestos')
            activo = data.get('estado')

            if not nombre_puestos:
                return JsonResponse({'error': 'El nombre del puesto es obligatorio'})
            
            if Puestos.objects.filter(nombre_puestos=nombre_puestos).exists():
                return JsonResponse({'error': 'El puesto ya está registrado'}, status=409)
            
            puesto = Puestos.objects.create(
                nombre_puestos = nombre_puestos,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Puesto registrada correctamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_puestos'):
                return JsonResponse({'error': 'No tienes permisos para actualizar puestos'})
            
            data = json.loads(request.body)

            puestos_id = data.get('id')
            nombre_puestos = data.get('nombre_puestos')
            activo = data.get('estado')

            if Puestos.objects.filter(nombre_puestos=nombre_puestos).exclude(id=puestos_id).exists():
                return JsonResponse({'error': 'El puesto ya está registrada'}, status=409)

            puestos = Puestos.objects.get(id=puestos_id)
            puestos.nombre_puestos = nombre_puestos
            puestos.activo = activo
            puestos.usuario_modifico = request.user
            puestos.save()

            return JsonResponse({'message': 'Puesto actualizado correctamente'}, status=200)
        
        except Puestos.DoesNotExist:
            return JsonResponse({'error': 'Puesto no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_puestos'):
            return redirect('sinacceso')
        
        puestos_id = request.GET.get('puestos_id')

        if puestos_id:
            try:
                puesto = Puestos.objects.get(id=puestos_id)

                return JsonResponse({
                    'puestos':{
                        'id': puesto.id,
                        'nombre_puestos': puesto.nombre_puestos,
                        'activo': puesto.activo
                    }
                })
            except Puestos.DoesNotExist:
                return JsonResponse({'error': 'Puesto no encontrada'}, status=404)  
        else:
            search = request.GET.get('search', '')  
            estado = request.GET.get('estado', '')  
            puestos = Puestos.objects.all()

            if search:
                puestos = puestos.filter(
                    models.Q(id__icontains=search) |  
                    models.Q(nombre_puestos__icontains=search) 
                )

            if estado:
                puestos = puestos.filter(activo__icontains=estado)  

            paginator = Paginator(puestos, 10)
            page_number = request.GET.get('page')  
            page_obj = paginator.get_page(page_number)  

            total_pages = paginator.num_pages
            current_page = page_obj.number

            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = list(range(start_page, end_page + 1))

            context = {
                'puestos': page_obj,  
                'search': search,  
                'estado': estado,
                'page_range': page_range,
            }
            return render(request, 'modulos/puestos.html', context)          
    
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------PRIORIDADES VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def prioridades_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_prioridad'):
                return JsonResponse({'error': 'No tienes permisos para registrar prioridades'})
    
            data = json.loads(request.body)

            nombre_prioridad = data.get('nombre_prioridad')
            activo = data.get('estado')

            if not nombre_prioridad:
                return JsonResponse({'error': 'El nombre de la prioridad es obligatorio'})
            
            if Prioridad.objects.filter(nombre_prioridad=nombre_prioridad).exists():
                return JsonResponse({'error': 'La Prioridad ya está registrada'}, status=409)

            prioridades = Prioridad.objects.create(
                nombre_prioridad = nombre_prioridad,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Prioridad registrada correctamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_prioridad'):
                return JsonResponse({'error': 'No tienes permisos para actualizar prioridades'})
            
            data = json.loads(request.body)

            id_prioridad = data.get('id')
            nombre_prioridad = data.get('nombre_prioridad')
            activo = data.get('estado')

            if Prioridad.objects.filter(nombre_prioridad=nombre_prioridad).exclude(id=id_prioridad).exists():
                return JsonResponse({'error': 'Prioridad ya está registrada'})
            
            prioridades = Prioridad.objects.get(id=id_prioridad)
            prioridades.nombre_prioridad = nombre_prioridad
            prioridades.activo = activo
            prioridades.usuario_modifico = request.user
            prioridades.save()

            return JsonResponse({'message': 'Prioridad actualizado correctamente'}, status=200)
        
        except Prioridad.DoesNotExist:
            return JsonResponse({'error': 'Prioridad no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_prioridad'):
            return redirect('sinacceso')
        
        prioridades_id = request.GET.get('prioridades_id')

        if prioridades_id:
            try:
                prioridades = Prioridad.objects.get(id=prioridades_id)

                return JsonResponse({
                    'prioridades':{
                        'id': prioridades.id,
                        'nombre_prioridad': prioridades.nombre_prioridad,
                        'activo': prioridades.activo
                    }
                })
            except Puestos.DoesNotExist:
                return JsonResponse({'error': 'Prioridades no encontrada'}, status=404)   

        else:
            search = request.GET.get('search', '')
            estado = request.GET.get('estado', '')
            prioridades = Prioridad.objects.all()

            if search:
                prioridades = prioridades.filter(
                    models.Q(id__icontains=search) |
                    models.Q(nombre_prioridad__icontains=search)
                )

            if estado:
                prioridades = prioridades.filter(activo__icontains=estado)

            paginator = Paginator(prioridades, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'prioridades': page_obj,
                'search': search,
                'estado': estado,
                'page_range': page_range,
            }
            return render(request, 'modulos/prioridades.html', context)      

    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------MODO VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def modos_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_modo'):
                return JsonResponse({'error': 'No tienes permisos para registrar modos'})
            
            data = json.loads(request.body)

            nombre_modo = data.get('nombre_modo')
            activo = data.get('estado')

            if not nombre_modo:
                return JsonResponse({'error': 'El nombre del modo es obligatorio'})
            
            if Modo.objects.filter(nombre_modo=nombre_modo).exists():
                return JsonResponse({'error': 'El Modo ya está registrado'}, status=409)
            
            modos = Modo.objects.create(
                nombre_modo = nombre_modo,
                activo = activo,
                usuario_creo = request.user
            )
            return JsonResponse({'message': 'Modo registrado correctamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_modo'):
                return JsonResponse({'error': 'No tienes permisos para actualizar modos'})
            
            data = json.loads(request.body)

            modo_id = data.get('id')
            nombre_modo = data.get('nombre_modo')
            activo = data.get('estado')

            if Modo.objects.filter(nombre_modo=nombre_modo).exclude(id=modo_id).exists():
                return JsonResponse({'error': 'El Modo ya está registrado'}, status=409)
            
            modos = Modo.objects.get(id=modo_id)
            modos.nombre_modo = nombre_modo
            modos.activo = activo
            modos.usuario_modifico = request.user
            modos.save()

            return JsonResponse({'message': 'Modo actualizado correctamente'}, status=200)
        
        except Modo.DoesNotExist:
            return JsonResponse({'error': 'Modo no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_modo'):
            return redirect('sinacceso')
        
        modos_id = request.GET.get('modo_id')
        
        if modos_id:
            try:
                modos = Modo.objects.get(id=modos_id)

                return JsonResponse({
                    'modo' :{
                        'id': modos.id,
                        'nombre_modo': modos.nombre_modo,
                        'activo': modos.activo
                    }
                })
            except Modo.DoesNotExist:
                return JsonResponse({'error': 'Modo no encontrado'}, status=404)

        else:
            search = request.GET.get('search', '')
            estado = request.GET.get('estado', '')
            modos = Modo.objects.all()

            if search:
                modos = modos.filter(
                    models.Q(id__icontains=search) |
                    models.Q(nombre_modo__icontains=search)
                )

            if estado:
                modos = modos.filter(activo__icontains=estado)

            paginator = Paginator(modos, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'modos': page_obj,
                'search': search,
                'estado': estado,
                'page_range': page_range,
            }
            return render(request, 'modulos/modos.html', context)
  

#--------MOTIVO VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def motivos_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_motivo'):
                return JsonResponse({'error': 'No tienes permisos para registrar motivos'})
            
            data = json.loads(request.body)

            nombre_motivo = data.get('nombre_motivo')
            activo = data.get('estado')

            if not nombre_motivo:
                return JsonResponse({'error': 'El nombre del motivo es obligatorio'})
            
            if Motivo.objects.filter(nombre_motivo=nombre_motivo).exists():
                return JsonResponse({'error': 'El Motivo ya está registrado'}, status=409)
            
            motivos = Motivo.objects.create(
                nombre_motivo = nombre_motivo,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Motivo registrada correctamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_motivo'):
                return JsonResponse({'error': 'No tienes permisos para actualizar motivos'})
            
            data = json.loads(request.body)

            motivos_id = data.get('id')
            nombre_motivo = data.get('nombre_motivo')
            activo = data.get('estado')

            if Motivo.objects.filter(nombre_motivo=nombre_motivo).exclude(id=motivos_id).exists():
                return JsonResponse({'error': 'El motivo ya está registrada'}, status=409)
            
            motivos = Motivo.objects.get(id=motivos_id)
            motivos.nombre_motivo = nombre_motivo
            motivos.activo = activo
            motivos.usuario_modifico = request.user
            motivos.save()

            return JsonResponse({'message': 'Motivo actualizado correctamente'}, status=200)
        
        except Motivo.DoesNotExist:
            return JsonResponse({'error': 'Motivo no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_motivo'):
            return redirect('sinacceso')
        
        motivos_id = request.GET.get('motivo_id')

        if motivos_id:
            try:
                motivos = Motivo.objects.get(id=motivos_id)

                return JsonResponse({
                    'motivos':{
                        'id': motivos.id,
                        'nombre_motivo': motivos.nombre_motivo,
                        'activo': motivos.activo
                    }
                })
            except Motivo.DoesNotExist:
                return JsonResponse({'error': 'Puesto no encontrada'}, status=404)  
        else:
            search = request.GET.get('search', '')  
            estado = request.GET.get('estado', '')  
        
            motivos = Motivo.objects.all()

            if search:
                motivos = motivos.filter(
                    models.Q(id__icontains=search) |
                    models.Q(nombre_motivo__icontains=search)
                )

            if estado:
                motivos = motivos.filter(activo__icontains=estado)

            paginator = Paginator(motivos, 10)  
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'motivos': page_obj,
                'search': search,
                'estado': estado,
                'page_range': page_range,
            }

            return render(request, 'modulos/motivos.html', context)

    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------MEDIO DE RECLUTAMIENTO VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def medios_reclutamiento_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_medioreclutamiento'):
                return JsonResponse({'error': 'No tienes permisos para registrar medios de reclutamiento'})
    
            data = json.loads(request.body)

            nombre_medio_de_reclutamiento= data.get('nombre_medio_de_reclutamiento')
            activo = data.get('estado')

            if not nombre_medio_de_reclutamiento:
                return JsonResponse({'error': 'El nombre del medio de reclutamiento es obligatorio'})
            
            if MedioReclutamiento.objects.filter(nombre_medio_de_reclutamiento=nombre_medio_de_reclutamiento).exists():
                return JsonResponse({'error': 'El medio de reclutamiento ya está registrado'}, status=409)
                
            medios = MedioReclutamiento.objects.create(
                nombre_medio_de_reclutamiento = nombre_medio_de_reclutamiento,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Medio de reclutamiento registrado correctamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_medioreclutamiento'):
                return JsonResponse({'error': 'No tienes permisos para actualizar medios de reclutamiento'})
            
            data = json.loads(request.body)

            medio_reclutamiento_id = data.get('id')
            nombre_medio_de_reclutamiento = data.get('nombre_medio_de_reclutamiento')
            activo = data.get('estado')

            if MedioReclutamiento.objects.filter(nombre_medio_de_reclutamiento=nombre_medio_de_reclutamiento).exclude(id=medio_reclutamiento_id).exists():
                return JsonResponse({'error': 'El medio de reclutamiento ya está registrado'}, status=409)

            medios = MedioReclutamiento.objects.get(id=medio_reclutamiento_id)
            medios.nombre_medio_de_reclutamiento = nombre_medio_de_reclutamiento
            medios.activo = activo
            medios.usuario_modifico = request.user
            medios.save()

            return JsonResponse({'message': 'Medio de reclutamiento actualizado correctamente'}, status=200)
        
        except MedioReclutamiento.DoesNotExist:
            return JsonResponse({'error': 'Medio de reclutamiento no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_medioreclutamiento'):
            return redirect('sinacceso')
        
        medio_reclutamiento_id = request.GET.get('medio_reclutamiento_id')

        if medio_reclutamiento_id:
            try:
                medios = MedioReclutamiento.objects.get(id=medio_reclutamiento_id)

                return JsonResponse({
                    'medio_reclutamiento':{
                        'id': medios.id,
                        'nombre_medio_de_reclutamiento': medios.nombre_medio_de_reclutamiento,
                        'activo': medios.activo
                    }
                })
            except MedioReclutamiento.DoesNotExist:
                return JsonResponse({'error': 'Medio de reclutamiento no encontrada'}, status=404)  
        else:
            search = request.GET.get('search', '')  
            estado = request.GET.get('estado', '')  
        
            medios_reclutamiento = MedioReclutamiento.objects.all()

            if search:
                medios_reclutamiento = medios_reclutamiento.filter(
                    models.Q(id__icontains=search) |  
                    models.Q(nombre_medio_de_reclutamiento__icontains=search)
                )

            if estado:
                medios_reclutamiento = medios_reclutamiento.filter(activo__icontains=estado)

            paginator = Paginator(medios_reclutamiento, 10)  
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'medios_reclutamiento': page_obj,
                'search': search,
                'estado': estado,
                'page_range': page_range,
            }

            return render(request, 'modulos/medios_reclutamiento.html', context)

    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)

#--------TIPOS DE CONTRATO VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def tipo_contrato_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_tipocontrato'):
                return JsonResponse({'error': 'No tienes permisos para registrar tipos de contrato'})
    
            data = json.loads(request.body)

            nombre_tipo_de_contrato= data.get('nombre_tipo_de_contrato')
            activo = data.get('estado')

            if not nombre_tipo_de_contrato:
                return JsonResponse({'error': 'El nombre del tipo de contrato es obligatorio'})
            
            if TipoContrato.objects.filter(nombre_tipo_de_contrato=nombre_tipo_de_contrato).exists():
                return JsonResponse({'error': 'El tipo de contrato ya está registrado'}, status=409)
                
            tipocontrato = TipoContrato.objects.create(
                nombre_tipo_de_contrato = nombre_tipo_de_contrato,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Tipo de contrato registrado correctamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_tipocontrato'):
                return JsonResponse({'error': 'No tienes permisos para actualizar tipos de contratos'})
            
            data = json.loads(request.body)

            tipo_de_contrato_id = data.get('id')
            nombre_tipo_de_contrato = data.get('nombre_tipo_de_contrato')
            activo = data.get('estado')

            if TipoContrato.objects.filter(nombre_tipo_de_contrato=nombre_tipo_de_contrato).exclude(id=tipo_de_contrato_id).exists():
                return JsonResponse({'error': 'El tipo de contrato ya está registrado'}, status=409)

            tipocontrato = TipoContrato.objects.get(id=tipo_de_contrato_id)
            tipocontrato.nombre_tipo_de_contrato = nombre_tipo_de_contrato
            tipocontrato.activo = activo
            tipocontrato.usuario_modifico = request.user
            tipocontrato.save()

            return JsonResponse({'message': 'Tipo de contrato actualizado correctamente'}, status=200)
        
        except TipoContrato.DoesNotExist:
            return JsonResponse({'error': 'Tipo de contrato no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_tipocontrato'):
            return redirect('sinacceso')
        
        tipo_de_contrato_id = request.GET.get('tipo_de_contrato_id')

        if tipo_de_contrato_id:
            try:
                tipocontrato = TipoContrato.objects.get(id=tipo_de_contrato_id)

                return JsonResponse({
                    'tipocontrato':{
                        'id': tipocontrato.id,
                        'nombre_tipo_de_contrato': tipocontrato.nombre_tipo_de_contrato,
                        'activo': tipocontrato.activo
                    }
                })
            except TipoContrato.DoesNotExist:
                return JsonResponse({'error': 'Medio de reclutamiento no encontrada'}, status=404)  
        else:
            search = request.GET.get('search', '')  
            estado = request.GET.get('estado', '')  
        
            tipos_contrato = TipoContrato.objects.all()

            if search:
                tipos_contrato = tipos_contrato.filter(
                    models.Q(id__icontains=search) |  
                    models.Q(nombre_tipo_de_contrato__icontains=search)
                )

            if estado:
                tipos_contrato = tipos_contrato.filter(activo__icontains=estado)

            paginator = Paginator(tipos_contrato, 10)  
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'tipos_contrato': page_obj,
                'search': search,
                'estado': estado,
                'page_range': page_range, 
            }

            return render(request, 'modulos/tipo_contrato.html', context)

    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------PSICOSMART VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def pruebapsicosmart_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_spicosmart'):
                return JsonResponse({'error': 'No tienes permisos para registrar pruebas psicosmart'})
    
            data = json.loads(request.body)

            nivel = data.get('nivel')
            nombre_prueba = data.get('nombre_prueba')
            prueba_mide= data.get('prueba_mide')
            activo = data.get('estado')

            if not nombre_prueba:
                return JsonResponse({'error': 'El nombre de la prueba psicosmart es obligatorio'})
            
            if Spicosmart.objects.filter(nombre_prueba=nombre_prueba, nivel=nivel).exists():
                return JsonResponse({'error': 'La prueba psicosmart ya está registrada'}, status=409)
                
            psicosmart = Spicosmart.objects.create(
                nivel = nivel,
                nombre_prueba = nombre_prueba,
                prueba_mide = prueba_mide,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Prueba psicosmart registrada correctamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_spicosmart'):
                return JsonResponse({'error': 'No tienes permisos para actualizar pruebas psicosmart'})
            
            data = json.loads(request.body)

            prueba_id = data.get('id')
            nivel = data.get('nivel')
            nombre_prueba = data.get('nombre_prueba')
            prueba_mide= data.get('prueba_mide')
            activo = data.get('estado')

            if Spicosmart.objects.filter(nombre_prueba=nombre_prueba).exclude(id=prueba_id).exists():
                return JsonResponse({'error': 'La prueba psicosmart ya está registrada'}, status=409)
            
            psicosmart = Spicosmart.objects.get(id=prueba_id)
            psicosmart.nivel = nivel
            psicosmart.nombre_prueba = nombre_prueba
            psicosmart.prueba_mide = prueba_mide
            psicosmart.activo = activo
            psicosmart.usuario_modifico = request.user
            psicosmart.save()

            return JsonResponse({'message': 'Prueba psicosmart actualizada correctamente'}, status=200)
        
        except Spicosmart.DoesNotExist:
            return JsonResponse({'error': 'Prueba psicosmart no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
             
    elif request.method == 'GET':
            if not request.user.has_perm('Reclutamiento.view_spicosmart'):
                return redirect('sinacceso')
            
            prueba_id = request.GET.get('prueba_id')

            if prueba_id:
                try:
                    psicosmart = Spicosmart.objects.get(id=prueba_id)

                    return JsonResponse({
                        'psicosmart':{
                            'id': psicosmart.id,
                            'nivel': psicosmart.nivel,
                            'nombre_prueba': psicosmart.nombre_prueba,
                            'prueba_mide': psicosmart.prueba_mide,
                            'activo': psicosmart.activo
                        }
                    })
                except Spicosmart.DoesNotExist:
                    return JsonResponse({'error': 'Prueba psicosmart no encontrada'}, status=404)
            else:
                search = request.GET.get('search', '')
                estado = request.GET.get('estado', '')

                pruebas = Spicosmart.objects.all()

                if search:
                    pruebas = pruebas.filter(
                        models.Q(nombre_prueba__icontains=search) |
                        models.Q(nivel__icontains=search) |
                        models.Q(prueba_mide__icontains=search) |
                        models.Q(activo__icontains=search) 
                    )

                if estado:
                    pruebas = pruebas.filter(activo__icontains=estado)

                pruebas = pruebas.order_by('id')

                paginator = Paginator(pruebas, 10)
                page_number = request.GET.get('page')
                page_obj = paginator.get_page(page_number)

                total_pages = paginator.num_pages
                current_page = page_obj.number
                start_page = max(1, current_page - 2)
                end_page = min(total_pages, current_page + 2)

                if current_page <= 3:
                    end_page = min(5, total_pages)
                elif current_page >= total_pages - 2:
                    start_page = max(1, total_pages - 4)

                page_range = range(start_page, end_page + 1)

                context = {
                    'pruebas': page_obj,
                    'search': search,
                    'estado': estado,
                    'page_range': page_range,
                }

                return render(request, 'modulos/psicosmart.html', context)
    
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------PAISES VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def paises_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_pais'):
                return JsonResponse({'error': 'No tienes permisos para registrar paises'})
            
            data = json.loads(request.body)

            nombre_pais = data.get('nombre_pais')
            activo = data.get('estado')

            if not nombre_pais:
                return JsonResponse({'error': 'El nombre del pais es obligatorio'})
            
            if Pais.objects.filter(nombre_pais=nombre_pais).exists():
                return JsonResponse({'error': 'El pais ya está registrado'}, status=409)
            
            pais = Pais.objects.create(
                nombre_pais = nombre_pais,
                activo = activo,
                usuario_creo = request.user
            )
            return JsonResponse({'message': 'Pais registrado correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_pais'):
                return JsonResponse({'error': 'No tienes permisos para actualizar paises'})
            
            data = json.loads(request.body)

            id = data.get('id')
            nombre_pais = data.get('nombre_pais')
            activo = data.get('estado')

            if Pais.objects.filter(nombre_pais=nombre_pais).exclude(id=id).exists():
                return JsonResponse({'error': 'El pais ya está registrado'}, status=409)
            
            paises = Pais.objects.get(id=id)
            paises.nombre_pais = nombre_pais
            paises.activo = activo
            paises.usuario_modifico = request.user
            paises.save()

            return JsonResponse({'message': 'Pais actualizado correctamente'}, status=200)
        
        except Pais.DoesNotExist:
            return JsonResponse({'error':'Pais no encontrada'}, status=403)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_pais'):
            return redirect('sinacceso')
        
        pais_id = request.GET.get('pais_id')
        
        if pais_id:
            try:
                pais = Pais.objects.get(id=pais_id)

                return JsonResponse({
                    'paises':{
                        'id': pais.id,
                        'nombre_pais': pais.nombre_pais,
                        'activo': pais.activo
                    }
                })
            except Pais.DoesNotExist:
                return JsonResponse({'error': 'Pais no encontrado'}, status=404)
        else:
            search = request.GET.get('search', '')
            estado = request.GET.get('estado', '')

            paises = Pais.objects.all()

            if search:
                paises = paises.filter(
                    models.Q(id__icontains=search) |
                    models.Q(nombre_pais__icontains=search)
                )

            if estado:
                paises = paises.filter(activo__icontains=estado)

            paginator = Paginator(paises, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'paises': page_obj,
                'search': search,
                'estado': estado,
                'page_range': page_range,
            }

            return render(request, 'modulos/pais.html', context)
    
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------DEPARTAMENTO PAISES VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def departamento_paises_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_departamentopais'):
                return JsonResponse({'error': 'No tienes permisos para registrar departamentos'})
            
            data = json.loads(request.body)

            nombre_departamento = data.get('nombre_departamento')
            pais = data.get('pais')
            activo = data.get('estado')

            if not nombre_departamento:
                return JsonResponse({'error': 'El nombre del departamento es obligatorio'})
            
            if DepartamentoPais.objects.filter(nombre_departamento=nombre_departamento).exists():
                return JsonResponse({'error': 'El departamento ya está registrado'}, status=409)
            
            departamento_pais = DepartamentoPais.objects.create(
                nombre_departamento = nombre_departamento,
                pais_id = pais,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Departamento registrado correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_departamentopais'):
                return JsonResponse({'error': 'No tienes permisos para actualizar departamentos'})
            
            data = json.loads(request.body)

            departamentopais_id = data.get('id')
            nombre_departamento = data.get('nombre_departamento')
            pais = data.get('pais')
            activo = data.get('estado')

            if DepartamentoPais.objects.filter(nombre_departamento=nombre_departamento).exclude(id=departamentopais_id).exists():
                return JsonResponse({'error': 'El departamento ya está registrado'}, status=409)
            
            departamento_pais = DepartamentoPais.objects.get(id=departamentopais_id)
            departamento_pais.nombre_departamento = nombre_departamento
            departamento_pais.pais_id = pais
            departamento_pais.activo = activo
            departamento_pais.usuario_modifico = request.user
            departamento_pais.save()

            return JsonResponse({'message': 'Departamento actualizado correctamente'}, status=200)
        
        except DepartamentoPais.DoesNotExist:
            return JsonResponse({'error':'Departamento no encontrada'}, status=403)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_departamentopais'):
            return redirect('sinacceso')
        
        departamentopais_id = request.GET.get('departamentopais_id')

        if departamentopais_id:
            try:
                departamento_pais = DepartamentoPais.objects.get(id=departamentopais_id)

                return JsonResponse({
                    'departamento_pais':{
                        'id': departamento_pais.id,
                        'nombre_departamento': departamento_pais.nombre_departamento,
                        'pais_id': departamento_pais.pais_id,
                        'activo': departamento_pais.activo
                    }
                })
            except DepartamentoPais.DoesNotExist:
                return JsonResponse({'error': 'Departamento no encontrado'}, status=404)
        else:
            search = request.GET.get('search', '')
            estado = request.GET.get('estado', '')

            departamentos = DepartamentoPais.objects.select_related('pais').all()
            paises = Pais.objects.all()

            if search:
                departamentos = departamentos.filter(
                    models.Q(id__icontains=search) |
                    models.Q(nombre_departamento__icontains=search) |
                    models.Q(pais__nombre_pais__icontains=search)
                )

            if estado:
                departamentos = departamentos.filter(activo__icontains=estado)

            paginator = Paginator(departamentos, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'departamentos': page_obj,
                'paises': paises,
                'search': search,
                'estado': estado,
                'page_range': page_range,
            }

            return render(request, 'modulos/departamentospais.html', context)
    
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------MUNICIPIOS VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def municipios_paises_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_municipiopais'):
                return JsonResponse({'error': 'No tienes permisos para registrar municipios'})
            
            data = json.loads(request.body)

            nombre_municipio = data.get('nombre_municipio')
            departamento = data.get('departamento')
            activo = data.get('estado')

            if not nombre_municipio:
                return JsonResponse({'error': 'El nombre del municipio es obligatorio'})
            
            if MunicipioPais.objects.filter(nombre_municipio=nombre_municipio).exists():
                return JsonResponse({'error': 'El municipio ya está registrado'}, status=409)
            
            municipio = MunicipioPais.objects.create(
                nombre_municipio = nombre_municipio,
                departamento_id = departamento,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Municipio registrado correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_municipiopais'):
                return JsonResponse({'error': 'No tienes permisos para actualizar municipios'})
            
            data = json.loads(request.body)

            municipio_id = data.get('id')
            nombre_municipio = data.get('nombre_municipio')
            departamento = data.get('departamento')
            activo = data.get('estado')

            if MunicipioPais.objects.filter(nombre_municipio=nombre_municipio).exclude(id=municipio_id).exists():
                return JsonResponse({'error': 'El municipio ya está registrado'}, status=409)
            
            municipio = MunicipioPais.objects.get(id=municipio_id)
            municipio.nombre_municipio = nombre_municipio
            municipio.departamento_id = departamento
            municipio.activo = activo
            municipio.usuario_modifico = request.user
            municipio.save()

            return JsonResponse({'message': 'Municipio actualizado correctamente'}, status=200)
        
        except MunicipioPais.DoesNotExist:
            return JsonResponse({'error':'Municipio no encontrada'}, status=403)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_municipiopais'):
            return redirect('sinacceso')
        
        municipiopais_id = request.GET.get('municipiopais_id')

        if municipiopais_id:
            try:
                municipio = MunicipioPais.objects.get(id=municipiopais_id)

                return JsonResponse({
                    'municipio':{
                        'id': municipio.id,
                        'nombre_municipio': municipio.nombre_municipio,
                        'departamento_id': municipio.departamento_id,
                        'activo': municipio.activo
                    }
                })
            except MunicipioPais.DoesNotExist:
                return JsonResponse({'error': 'Municipio no encontrado'}, status=404)
        else:
            search = request.GET.get('search', '')

            municipios = MunicipioPais.objects.select_related('departamento').all()
            departamentos = DepartamentoPais.objects.all()

            if search:
                municipios = municipios.filter(
                    models.Q(id__icontains=search) |
                    models.Q(nombre_municipio__icontains=search) |
                    models.Q(departamento__nombre_departamento__icontains=search)
                )

            paginator = Paginator(municipios, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'municipios': page_obj,
                'departamentos': departamentos,
                'search': search,
                'page_range': page_range,
            }

            return render(request, 'modulos/municipiospais.html', context)

    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------JEFES VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def jefes_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_jefes'):
                return JsonResponse({'error': 'No tienes permisos para registrar jefes'})
            
            data = json.loads(request.body)

            nombrejefe = data.get('nombrejefe')
            codigo = data.get('codigo')
            identidadjefe = data.get('identidadjefe')
            activo = data.get('estado')

            if not nombrejefe:
                return JsonResponse({'error': 'El nombre del jefe es obligatorio'})
            
            if Jefes.objects.filter(nombrejefe=nombrejefe).exists():
                return JsonResponse({'error': 'El jefe ya está registrado'}, status=409)
            
            jefe = Jefes.objects.create(
                nombrejefe = nombrejefe,
                codigo = codigo,
                identidadjefe = identidadjefe,
                activo = activo,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Jefe registrado correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_jefes'):
                return JsonResponse({'error': 'No tienes permisos para actualizar jefes'})
            
            data = json.loads(request.body)

            jefe_id = data.get('id')
            nombrejefe = data.get('nombrejefe')
            identidadjefe = data.get('identidadjefe')
            codigo = data.get('codigo')
            activo = data.get('estado')

            if Jefes.objects.filter(nombrejefe=nombrejefe).exclude(id=jefe_id).exists():
                return JsonResponse({'error': 'El jefe ya está registrado'}, status=409)
            
            jefe = Jefes.objects.get(id=jefe_id)
            jefe.nombrejefe = nombrejefe
            jefe.codigo = codigo
            jefe.identidadjefe = identidadjefe
            jefe.activo = activo
            jefe.usuario_modifico = request.user
            jefe.save()

            return JsonResponse({'message': 'Jefe actualizado correctamente'}, status=200)
        
        except Jefes.DoesNotExist:
            return JsonResponse({'error':'Jefe no encontrado'}, status=403)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_jefes'):
            return redirect('sinacceso')
        
        jefe_id = request.GET.get('jefe_id')

        if jefe_id:
            try:
                jefe = Jefes.objects.get(id=jefe_id)

                return JsonResponse({
                    'jefe':{
                        'id': jefe.id,
                        'nombrejefe': jefe.nombrejefe,
                        'codigo': jefe.codigo,
                        'identidadjefe': jefe.identidadjefe,
                        'activo': jefe.activo
                    }
                })
            except Jefes.DoesNotExist:
                return JsonResponse({'error': 'Jefe no encontrado'}, status=404)
            
        else: 
            search = request.GET.get('search', '')

            jefes = Jefes.objects.all()

            if search:
                jefes = jefes.filter(
                    models.Q(id__icontains=search) |
                    models.Q(nombrejefe__icontains=search) |
                    models.Q(codigo__icontains=search) |
                    models.Q(identidadjefe__icontains=search)
                )

            paginator = Paginator(jefes, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'jefes': page_obj,
                'search': search,
                'page_range': page_range,
            }

            return render(request, 'modulos/jefes.html', context)
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------COLABORADORES VIEW--------#   
@csrf_exempt
@login_required(login_url='/login/')
def colaboradores_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_colaboradores'):
                return JsonResponse({'error': 'No tienes permisos para registrar colaboradores'})
            
            data = json.loads(request.body)

            nombrecolaborador = data.get('nombrecolaborador')
            sucursal_id = data.get('sucursal')
            empresa_id = data.get('empresa')
            unidad_de_negocio_id = data.get('unidadnegocio')
            departamento_id = data.get('departamento')
            jefe_id = data.get('jefes')
            activo = data.get('estado')
            codigocolaborador = data.get('codigocolaborador')

            if not nombrecolaborador:
                return JsonResponse({'error': 'El nombre del colaborador es obligatorio'})
            
            if Colaboradores.objects.filter(nombrecolaborador=nombrecolaborador,sucursal_id=sucursal_id, empresa_id=empresa_id,unidad_de_negocio_id=unidad_de_negocio_id,departamento_id=departamento_id,jefe_id=jefe_id,codigocolaborador=codigocolaborador).exists():
                return JsonResponse({'error': 'El colaborador ya está registrado'}, status=409)
            
            colaboradores = Colaboradores.objects.create(
                nombrecolaborador = nombrecolaborador,
                sucursal_id = sucursal_id,
                empresa_id = empresa_id,
                unidad_de_negocio_id = unidad_de_negocio_id,
                departamento_id = departamento_id,
                jefe_id = jefe_id,
                activo = activo,
                codigocolaborador = codigocolaborador,
                usuario_creo = request.user
            )

            return JsonResponse({'message': 'Colaborador registrado correctamente'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)    


    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_colaboradores'):
                return JsonResponse({'error': 'No tienes permisos para actualizar colaboradores'})

            data = json.loads(request.body)

            id = data.get('id')
            nombrecolaborador = data.get('nombrecolaborador')
            sucursal = data.get('sucursal')or None
            empresa = data.get('empresa')or None
            unidad_de_negocio = data.get('unidad_de_negocio')or None
            departamento = data.get('departamento')or None
            jefes = data.get('jefes')or None
            activo = data.get('estado')
            codigocolaborador = data.get('codigocolaborador')


            if Colaboradores.objects.filter(nombrecolaborador=nombrecolaborador,sucursal_id=sucursal, empresa_id=empresa,unidad_de_negocio_id=unidad_de_negocio,departamento_id=departamento,jefe_id=jefes,codigocolaborador=codigocolaborador).exclude(id=id).exists():
                return JsonResponse({'error': 'El colaborador ya está registrado'}, status=409)
            
            colaborador = Colaboradores.objects.get(id=id)
            colaborador.nombrecolaborador = nombrecolaborador
            colaborador.sucursal_id = sucursal
            colaborador.empresa_id = empresa
            colaborador.unidad_de_negocio_id = unidad_de_negocio 
            colaborador.departamento_id = departamento
            colaborador.jefe_id = jefes
            colaborador.activo = activo
            colaborador.codigocolaborador = codigocolaborador
            colaborador.usuario_modifico = request.user
            colaborador.save()

            return JsonResponse({'message': 'Colaborador actualizado correctamente'}, status=200)
        
        except Colaboradores.DoesNotExist:
            return JsonResponse({'error':'Colaborador no encontrado'}, status=403)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_colaboradores'):
            return redirect('sinacceso')
        
        colaborador_id = request.GET.get('colaborador_id')

        if colaborador_id:
            try:
                colaborador = Colaboradores.objects.get(id=colaborador_id)

                return JsonResponse({
                    'colaborador':{
                        'id': colaborador.id,
                        'nombrecolaborador': colaborador.nombrecolaborador,
                        'sucursal_id': colaborador.sucursal_id,
                        'empresa_id': colaborador.empresa_id,
                        'unidad_de_negocio_id': colaborador.unidad_de_negocio_id,
                        'departamento_id': colaborador.departamento_id,
                        'jefe_id': colaborador.jefe_id,
                        'activo': colaborador.activo,
                        'codigocolaborador': colaborador.codigocolaborador
                    }
                })
            except Colaboradores.DoesNotExist:
                return JsonResponse({'error': 'Colaborador no encontrado'}, status=404)
            
        else:
            search = request.GET.get('search', '') 

            # Obtener todos los colaboradores
            colaboradores = Colaboradores.objects.all()

            # Filtrar por búsqueda en múltiples campos
            if search:
                colaboradores = colaboradores.filter(
                    models.Q(nombrecolaborador__icontains=search) |
                    models.Q(sucursal__nombre_sucursal__icontains=search) |
                    models.Q(empresa__nombre_empresa__icontains=search) |
                    models.Q(unidad_de_negocio__nombre_unidad_de_negocio__icontains=search) |
                    models.Q(departamento__nombre_departamento__icontains=search) |
                    models.Q(jefe__nombrejefe__icontains=search) |
                    models.Q(codigocolaborador__icontains=search)
                )

            # Paginación
            paginator = Paginator(colaboradores, 10)  # 10 colaboradores por página
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            # Calcular el rango de páginas a mostrar
            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            # Obtener listas para los selects del formulario
            all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()
            all_empresas = Empresas.objects.filter(activo=True).order_by('nombre_empresa').distinct()
            all_unidades_negocio = Unidad_Negocio.objects.filter(activo=True).order_by('nombre_unidad_de_negocio').distinct()
            all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
            all_jefes = Jefes.objects.filter(activo=True).order_by('nombrejefe').distinct()

            context = {
                'colaboradores': page_obj,
                'search': search,
                'page_range': page_range,
                'all_sucursales': all_sucursales,
                'all_empresas': all_empresas,
                'all_unidades_negocio': all_unidades_negocio,
                'all_departamentos': all_departamentos,
                'all_jefes': all_jefes
            }

            return render(request, 'modulos/colaboradores.html', context)
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------CONTROL TIEMPO VIEW--------#   
@csrf_exempt 
@login_required(login_url='/login/')
def control_tiempo_view(request):
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_controldetiempo'):
                return JsonResponse({'error': 'No tienes permisos para registrar control de tiempo'})
            
            data = json.loads(request.body)

            unidad_de_negocio_id = data.get('unidadNegocio')
            puestos_id = data.get('puesto')
            departamento_id = data.get('departamento')
            tiempo = data.get('tiempo')
            estado = data.get('estado') 

            unidad_negocio = Unidad_Negocio.objects.get(id=unidad_de_negocio_id)
            puesto = Puestos.objects.get(id=puestos_id)

            unid_puesto = f"{unidad_negocio.nombre_unidad_de_negocio}-{puesto.nombre_puestos}"

            ControlDeTiempo.objects.create(
                unidad_de_negocio_id=unidad_de_negocio_id,
                puestos_id=puestos_id,
                unid_puesto=unid_puesto,
                departamento_id=departamento_id,
                tiempo=tiempo,
                activo=estado,
                usuario_creo=request.user
            )
            return JsonResponse({'success': True, 'message': 'Control de Tiempo registrado correctamente.'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_controldetiempo'):
                return JsonResponse({'error': 'No tienes permisos para actualizar control de tiempo'})
            
            data = json.loads(request.body)

            controltiempo_id = data.get('id')
            unidadnegocio = data.get('unidadnegocio')
            departamento = data.get('departamento')
            puestos = data.get('puestos')
            tiempo = data.get('tiempo')
            estado = data.get('estado')

            unidad_negocio = Unidad_Negocio.objects.get(id=unidadnegocio)
            puest = Puestos.objects.get(id=puestos)

            if ControlDeTiempo.objects.filter(unidad_de_negocio_id=unidadnegocio, departamento_id=departamento, puestos_id=puestos, tiempo=tiempo).exclude(id=controltiempo_id).exists():
                return JsonResponse({'error': 'Control de tiempo ya está registrada'}, status=409)
            
            controltiempo = ControlDeTiempo.objects.get(id=controltiempo_id)
            controltiempo.unidad_de_negocio_id = unidadnegocio
            controltiempo.departamento_id = departamento
            controltiempo.puestos_id = puestos
            controltiempo.unid_puesto = f"{unidad_negocio}-{puest}"
            controltiempo.tiempo = tiempo
            controltiempo.activo = estado
            controltiempo.usuario_modifico = request.user
            controltiempo.save()

            return JsonResponse({'message': 'Control de tiempo actualizada correctamente'}, status=200)
        
        except ControlDeTiempo.DoesNotExist:
            return JsonResponse({'error': 'Control de tiempo no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
 
    elif request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_controldetiempo'):
            return redirect('sinacceso')
        
        control_tiempo_id = request.GET.get('control_tiempo_id')

        if control_tiempo_id:
            try:
                control_tiempo = ControlDeTiempo.objects.get(id=control_tiempo_id)

                return JsonResponse({
                    'control_tiempo':{
                        'id': control_tiempo.id,
                        'unidad_de_negocio': control_tiempo.unidad_de_negocio_id,
                        'departamento': control_tiempo.departamento_id,
                        'puestos': control_tiempo.puestos_id,
                        'tiempo': control_tiempo.tiempo,
                        'activo': control_tiempo.activo
                    }
                })
            
            except ControlDeTiempo.DoesNotExist:
                return JsonResponse({'error': 'Control de Tiempo no encontrado'}, status=404)
            
        else:
            unidades_negocio = Unidad_Negocio.objects.filter(activo=True).order_by('nombre_unidad_de_negocio')
            departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento')
            puestos = Puestos.objects.filter(activo=True).order_by('nombre_puestos')

            # Obtener todos los registros de ControlDeTiempo
            controles_tiempo = ControlDeTiempo.objects.all().prefetch_related('unidad_de_negocio', 'departamento', 'puestos')

            # Parámetros de búsqueda (opcional)
            search = request.GET.get('search', '')
            if search:
                controles_tiempo = controles_tiempo.filter(
                    Q(unidad_de_negocio__nombre_unidad_de_negocio__icontains=search) |
                    Q(departamento__nombre_departamento__icontains=search) |
                    Q(puestos__nombre_puestos__icontains=search) |
                    Q(tiempo__icontains=search)
                )

            # Paginación
            paginator = Paginator(controles_tiempo, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            # Calcular el rango de páginas a mostrar
            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            context = {
                'unidadesNegocio': unidades_negocio,
                'departamentos': departamentos,
                'puestos': puestos,
                'controlesTiempo': page_obj,
                'page_range': page_range,
                'search': search,
            }
            return render(request, 'controltiempo.html', context)
    
    return JsonResponse({'success': False, 'message': 'Metodo no permitido.'}, status=405)


#--------CONTROL PLAZAS VIEW--------#  
@csrf_exempt  
@login_required(login_url='/login/')
def control_plazas_view(request, id=None):
    if request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_controldeplazas'):
            return redirect('sinacceso')
        
        control_plaza_id = request.GET.get('control_plaza_id')
        
        if control_plaza_id:
            try:
                control_plaza = ControlDePlazas.objects.get(id=control_plaza_id)

                return JsonResponse({
                    'control_plaza':{
                        'id': control_plaza.id,
                        'analista': control_plaza.analista,
                        'sucursal': control_plaza.sucursal_id,
                        'empresa': control_plaza.empresa_id,
                        'unidad_negocio': control_plaza.unidad_de_negocio_id,
                        'ano': control_plaza.año,
                        'mes_corte': control_plaza.mes_corte,
                        'mes_solicitud': control_plaza.mes_solicitud,
                        'modo': control_plaza.modo_id,
                        'motivo': control_plaza.motivo_ingreso_id,
                        'nombre_reemplazo': control_plaza.nombre_reemplazo,
                        'puesto': control_plaza.puestos_id,
                        'departamento': control_plaza.departamento_id,
                        'prioridad': control_plaza.prioridad_id,
                        'fecha_solicitud': control_plaza.fecha_solicitud,
                        'fecha_cobertura': control_plaza.fecha_cobertura,
                        'fecha_inicio_busqueda': control_plaza.fecha_inicio_busqueda,
                        'fecha_ingreso': control_plaza.fecha_ingreso,
                        'fuente_reclutamiento': control_plaza.fuente_reclutamiento,
                        'nombre_contratado': control_plaza.nombre_contratado,
                        'dni': control_plaza.dni,
                        'genero': control_plaza.genero,
                        'edad': control_plaza.edad,
                        'fecha_nacimiento': control_plaza.fecha_nacimiento,
                        'medio_reclutamiento': control_plaza.medio_reclutamiento_id,
                        'salario': control_plaza.salario,
                        'combustible': control_plaza.combustible,
                        'depreciacion': control_plaza.depreciacion,
                        'comision': control_plaza.comision,
                        'bono': control_plaza.bono,
                        'hospedaje': control_plaza.hospedaje,
                        'tipo_contrato': control_plaza.tipo_contrato_id,
                        'nombreimagen': control_plaza.nombreimagen,
                        'nombreimagen1': control_plaza.nombreimagen1,
                    }
                })
            
            except ControlDePlazas.DoesNotExist:
                return JsonResponse({'error': 'Control de Plazas no encontrado'}, status=404)
        else:
                
            control_de_plazas = ControlDePlazas.objects.all().order_by('-id')

            # Obtener los registros de ControlDePlazas y valores únicos para los filtros
            estatus_options = set(control_de_plazas.values_list('estatus', flat=True))
            año_options = set(control_de_plazas.values_list('año', flat=True))
            mes_corte_options = set(control_de_plazas.values_list('mes_corte', flat=True))
            mes_solicitud_options = set(control_de_plazas.values_list('mes_solicitud', flat=True))

            # Obtener todas las sucursales, empresas, etc. para el formulario
            all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()
            all_empresas = Empresas.objects.filter(activo=True).order_by('nombre_empresa').distinct()
            all_unidades_negocio = Unidad_Negocio.objects.filter(activo=True).order_by('nombre_unidad_de_negocio').distinct()
            all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
            all_puestos = Puestos.objects.filter(activo=True).order_by('nombre_puestos').distinct()

            # Obtener opciones para Modo, Motivo, Prioridad, Tipo de Contrato y Medio de Reclutamiento
            modos = Modo.objects.filter(activo=True).order_by('nombre_modo').distinct()
            motivos = Motivo.objects.filter(activo=True).order_by('nombre_motivo').distinct()
            prioridades = Prioridad.objects.filter(activo=True).order_by('nombre_prioridad').distinct()
            tipos_contrato = TipoContrato.objects.filter(activo=True).order_by('nombre_tipo_de_contrato').distinct()
            mediosReclutamiento = MedioReclutamiento.objects.filter(activo=True).order_by('nombre_medio_de_reclutamiento').distinct()

            # Filtrar solo las sucursales, empresas, unidades de negocio, departamentos y puestos que estén en ControlDePlazas
            sucursales_options = Sucursal.objects.filter(
                id__in=control_de_plazas.values_list('sucursal__id', flat=True)
            ).distinct().order_by('nombre_sucursal')

            empresas_options = Empresas.objects.filter(
                id__in=control_de_plazas.values_list('empresa__id', flat=True)
            ).distinct().order_by('nombre_empresa')

            unidades_negocio_options = Unidad_Negocio.objects.filter(
                id__in=control_de_plazas.values_list('unidad_de_negocio__id', flat=True)
            ).distinct().order_by('nombre_unidad_de_negocio')

            departamentos_options = Departamento.objects.filter(
                id__in=control_de_plazas.values_list('departamento__id', flat=True)
            ).distinct().order_by('nombre_departamento')

            puestos_options = Puestos.objects.filter(
                id__in=control_de_plazas.values_list('puestos__id', flat=True)
            ).distinct().order_by('nombre_puestos')

            # Parámetros de búsqueda (opcional)
            search = request.GET.get('search', '')
            sucursal_filter = request.GET.getlist('sucursal')
            empresa_filter = request.GET.getlist('empresa')
            año_filter = request.GET.getlist('año')
            mes_corte_filter = request.GET.getlist('mes_corte')
            mes_solicitud_filter = request.GET.getlist('mes_solicitud')
            puesto_filter = request.GET.getlist('puesto')
            departamento_filter = request.GET.getlist('departamento')
            estatus_filter = request.GET.getlist('estatus')

            # Aplicar filtros en ControlDePlazas
            if search:
                control_de_plazas = control_de_plazas.filter(
                    Q(sucursal__nombre_sucursal__icontains=search) |
                    Q(empresa__nombre_empresa__icontains=search) |
                    Q(año__icontains=search) |
                    Q(mes_corte__icontains=search) |
                    Q(mes_solicitud__icontains=search) |
                    Q(puestos__nombre_puestos__icontains=search) |
                    Q(departamento__nombre_departamento__icontains=search) |
                    Q(estatus__icontains=search)
                )

            # Filtrar por Sucursal
            if sucursal_filter:
                control_de_plazas = control_de_plazas.filter(sucursal__id__in=sucursal_filter)

            # Filtrar por Empresa
            if empresa_filter:
                control_de_plazas = control_de_plazas.filter(empresa__id__in=empresa_filter)

            # Filtrar por Año
            if año_filter:
                control_de_plazas = control_de_plazas.filter(año__in=año_filter)

            # Filtrar por Mes de Corte
            if mes_corte_filter:
                control_de_plazas = control_de_plazas.filter(mes_corte__in=mes_corte_filter)

            # Filtrar por Mes de Solicitud
            if mes_solicitud_filter:
                control_de_plazas = control_de_plazas.filter(mes_solicitud__in=mes_solicitud_filter)

            # Filtrar por Puesto
            if puesto_filter:
                control_de_plazas = control_de_plazas.filter(puestos__id__in=puesto_filter)

            # Filtrar por Departamento
            if departamento_filter:
                control_de_plazas = control_de_plazas.filter(departamento__id__in=departamento_filter)

            # Filtrar por Estatus
            if estatus_filter:
                control_de_plazas = control_de_plazas.filter(estatus__in=estatus_filter)

            # Paginación
            paginator = Paginator(control_de_plazas, 10) 
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            # Calcular el rango de páginas a mostrar
            total_pages = paginator.num_pages
            current_page = page_obj.number
            start_page = max(1, current_page - 2)
            end_page = min(total_pages, current_page + 2)

            if current_page <= 3:
                end_page = min(5, total_pages)
            elif current_page >= total_pages - 2:
                start_page = max(1, total_pages - 4)

            page_range = range(start_page, end_page + 1)

            # Contexto para renderizar la plantilla
            context = {
                'sucursales': sucursales_options,
                'empresas': empresas_options,
                'unidadesNegocio': unidades_negocio_options,
                'departamentos': departamentos_options,
                'puestos': puestos_options,
                'estatus_options': list(estatus_options),  # Convertir set a list para el contexto
                'año_options': list(año_options),  # Convertir set a list para el contexto
                'mes_corte_options': list(mes_corte_options),  # Convertir set a list para el contexto
                'mes_solicitud_options': list(mes_solicitud_options),  # Convertir set a list para el contexto
                'controlDePlazas': page_obj,
                'page_range': page_range,
                'search': search,
                'sucursal_filter': sucursal_filter,
                'empresa_filter': empresa_filter,
                'año_filter': año_filter,
                'mes_corte_filter': mes_corte_filter,
                'mes_solicitud_filter': mes_solicitud_filter,
                'puesto_filter': puesto_filter,
                'departamento_filter': departamento_filter,
                'estatus_filter': estatus_filter,
                'all_sucursales': all_sucursales,
                'all_empresas': all_empresas,
                'all_unidades_negocio': all_unidades_negocio,
                'all_departamentos': all_departamentos,
                'all_puestos': all_puestos,
                'modos': modos,
                'motivos': motivos,
                'prioridades': prioridades,
                'tipos_contrato': tipos_contrato,
                'mediosReclutamiento': mediosReclutamiento,
            }

            return render(request, 'controlplazas.html', context)
    
        
    elif request.method == 'POST':
        if not request.user.has_perm('Reclutamiento.add_controldeplazas'):
            return JsonResponse({'error': 'No tienes permisos para crear plazas'}, status=403)
        try:
            # Validación de los datos de entrada
            data = json.loads(request.POST.get('data'))

            # Consultar la instancia de empresa (puede ser None)
            empresa = Empresas.objects.filter(id=data.get('empresa')).first()

            # Consultar la instancia de sucursal (puede ser None)
            sucursal = Sucursal.objects.filter(id=data.get('sucursal')).first()

            # Consultar Unidad de Negocio (puede ser None)
            unidad_negocio = Unidad_Negocio.objects.filter(id=data.get('unidad_de_negocio')).first()

            # Consultar Puesto (puede ser None)
            puesto = Puestos.objects.filter(id=data.get('puestos')).first()

            # Consultar otros modelos relacionados (pueden ser None)
            modo = Modo.objects.filter(id=data.get('modo')).first()
            motivo_ingreso = Motivo.objects.filter(id=data.get('motivo_ingreso')).first()
            departamento = Departamento.objects.filter(id=data.get('departamento')).first() if data.get('departamento') else None
            prioridad = Prioridad.objects.filter(id=data.get('prioridad')).first() if data.get('prioridad') else None
            medio_reclutamiento = MedioReclutamiento.objects.filter(id=data.get('medio_reclutamiento')).first() if data.get('medio_reclutamiento') else None
            tipo_contrato = TipoContrato.objects.filter(id=data.get('tipo_contrato')).first() if data.get('tipo_contrato') else None

            # Construir unidad_puesto (puede ser None)
            unidad_puesto = f"{unidad_negocio.nombre_unidad_de_negocio} - {puesto.nombre_puestos}" if unidad_negocio and puesto else None

            # Consultar ControlDeTiempo para obtener el tiempo de cobertura (puede ser None)
            control_de_tiempo = ControlDeTiempo.objects.filter(
                unidad_de_negocio=unidad_negocio,
                puestos=puesto,
                unid_puesto=unidad_puesto
            ).first()
            tiempo_cobertura = control_de_tiempo.tiempo if control_de_tiempo else None

            # Manejo de las fechas
            fecha_solicitud = parser.parse(data.get('fecha_solicitud')) if data.get('fecha_solicitud') else datetime.now()
            fecha_cobertura = parser.parse(data.get('fecha_cobertura')) if data.get('fecha_cobertura') else datetime.now()
            fecha_ingreso = parser.parse(data.get('fecha_ingreso')) if data.get('fecha_ingreso') else datetime.now()

            # Calcular tiempos efectivos
            tiempo_efectivo_cobertura = (fecha_cobertura - fecha_solicitud).days if fecha_solicitud and fecha_cobertura else None
            tiempo_efectivo_cobertura = tiempo_efectivo_cobertura if tiempo_efectivo_cobertura > 0 else None

            tiempo_efectivo_fecha_ingreso = (fecha_ingreso - fecha_solicitud).days if fecha_solicitud and fecha_ingreso else None
            tiempo_efectivo_fecha_ingreso = tiempo_efectivo_fecha_ingreso if tiempo_efectivo_fecha_ingreso > 0 else None

            # Calcular fecha límite de cobertura
            fecha_limite_cobertura_str = (fecha_solicitud + timedelta(days=tiempo_cobertura)).date() if fecha_solicitud and tiempo_cobertura else None

            # Capturar y convertir las fechas
            fecha_limite_cobertura_cal = datetime.combine(fecha_limite_cobertura_str, datetime.min.time()) if fecha_limite_cobertura_str else None
            fecha_cobertura_cal = parser.parse(data.get('fecha_cobertura')) if data.get('fecha_cobertura') else None

            # Calcular el tiempo disponible
            if fecha_limite_cobertura_cal and fecha_cobertura_cal:
                tiempo_disponible = (fecha_limite_cobertura_cal - fecha_cobertura_cal).days
                tiempo_disponible = tiempo_disponible if tiempo_disponible >= 0 else None
            else:
                tiempo_disponible = None

            # Determinar el estatus
            estatus = 'EN PROCESO' if not data.get('fecha_cobertura') else 'CERRADA' if 'fecha_ingreso' in data else 'ESPERA DE INGRESO'

            # Asignar cantidad solicitada y cantidad cubierta
            cantidad_solicitada = 1
            cantidad_cubierta = 1 if estatus == 'CERRADA' else None

            # Manejo de la imagen
            imagen_dni = request.FILES.get('imagen_dni')
            nombre_imagen = None
            ruta_imagen = None

            if imagen_dni:
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen_dni.name}"
                ruta_imagen = f"reclutamiento/static/img/controlplazas/{nombre_imagen}"

                with open(ruta_imagen, 'wb+') as destino:
                    for chunk in imagen_dni.chunks():
                        destino.write(chunk)

            # Manejo de la imagen del reverso
            imagen_dni_reverso = request.FILES.get('imagen_dni_reverso')
            nombre_imagen_reverso = None
            ruta_imagen_reverso = None

            if imagen_dni_reverso:
                nombre_imagen_reverso = f"{uuid.uuid4().hex[:4]}_{imagen_dni_reverso.name}"  # Generar un nombre único
                ruta_imagen_reverso = f"reclutamiento/static/img/controlplazas/{nombre_imagen_reverso}"

                with open(ruta_imagen_reverso, 'wb+') as destino:
                    for chunk in imagen_dni_reverso.chunks():
                        destino.write(chunk)

            control_plaza = ControlDePlazas.objects.create(
                analista=data.get('analistas'),
                sucursal=sucursal,  # Puede ser None
                empresa=empresa,  # Puede ser None
                unidad_de_negocio=unidad_negocio,  # Puede ser None
                año=data.get('año'),
                mes_corte=data.get('mes_corte'),
                mes_solicitud=data.get('mes_solicitud'),
                modo=modo,
                motivo_ingreso=motivo_ingreso,
                nombre_reemplazo=data.get('nombre_reemplazo'),
                puestos=puesto,  # Puede ser None
                departamento=departamento,  # Puede ser None
                prioridad=prioridad,  # Puede ser None
                unidad_puesto=unidad_puesto,
                tiempo_cobertura=tiempo_cobertura,
                fecha_solicitud=data.get('fecha_solicitud'),
                fecha_inicio_busqueda=parser.parse(data.get('fecha_inicio_busqueda')) if data.get('fecha_inicio_busqueda') else None,
                fecha_cobertura=data.get('fecha_cobertura'),
                fecha_ingreso=data.get('fecha_ingreso'),
                tiempo_efectivo_cobertura=tiempo_efectivo_cobertura,
                tiempo_efectivo_fecha_ingreso=tiempo_efectivo_fecha_ingreso,
                fecha_limite_cobertura=fecha_limite_cobertura_str,
                tiempo_disponible=tiempo_disponible,
                estatus=estatus,
                cantidad_solicitada=cantidad_solicitada,
                cantidad_cubierta=cantidad_cubierta,
                fuente_reclutamiento=data.get('fuente_reclutamiento'),
                nombre_contratado=data.get('nombre_contratado'),
                dni=data.get('dni'),
                genero=data.get('genero'),
                edad=data.get('edad'),
                fecha_nacimiento=data.get('fecha_nacimiento'),
                medio_reclutamiento=medio_reclutamiento,  # Puede ser None
                salario=data.get('salario'),
                combustible=data.get('combustible'),
                depreciacion=data.get('depreciacion'),
                comision=data.get('comision'),
                bono=data.get('bono'),
                hospedaje=data.get('hospedaje'),
                tipo_contrato=tipo_contrato,  # Puede ser None
                ruta=ruta_imagen,
                nombreimagen=nombre_imagen,
                ruta1=ruta_imagen_reverso,
                nombreimagen1=nombre_imagen_reverso,
                usuario_creo=request.user,
            )

            return JsonResponse({'success': True, 'message': 'Plaza registrada exitosamente', 'controlPlaza': control_plaza.id}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    if request.method == 'DELETE':
        try:
            if not request.user.has_perm('Reclutamiento.delete_controldeplazas'):
                return JsonResponse({'error': 'No tienes permisos para eliminar plazas'}, status=403)
            control_plaza = ControlDePlazas.objects.get(id=id)
            data = json.loads(request.body)
            estatus = data.get('estatus', None)
            
            if estatus:
                control_plaza.estatus = estatus 
                control_plaza.save()

            return JsonResponse({'success': True, 'message': 'Plaza eliminada exitosamente'}, status=200)

        except ControlDePlazas.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Control de plazas no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    if request.method == 'PATCH':
        try:
            control_plaza = ControlDePlazas.objects.get(id=id)
            data = json.loads(request.body) 
            estatus = data.get('estatus', None)

            if estatus:
                control_plaza.estatus = estatus 
                control_plaza.save() 

            return JsonResponse({'success': True, 'message': 'Estatus actualizado exitosamente'}, status=200)

        except ControlDePlazas.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Control de plazas no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt  
def control_plazas_update(request, id):
    if request.method == 'POST':
        try:
            control_plaza = ControlDePlazas.objects.get(id=id)

            # Validación de los datos de entrada
            data = json.loads(request.POST.get('data'))

            # Consultar la instancia de empresa (puede ser None)
            empresa = Empresas.objects.filter(id=data.get('empresa')).first()

            # Consultar la instancia de sucursal (puede ser None)
            sucursal = Sucursal.objects.filter(id=data.get('sucursal')).first()

            # Consultar Unidad de Negocio (puede ser None)
            unidad_negocio = Unidad_Negocio.objects.filter(id=data.get('unidad_de_negocio')).first()

            # Consultar Puesto (puede ser None)
            puesto = Puestos.objects.filter(id=data.get('puestos')).first()

            # Consultar otros modelos relacionados (pueden ser None)
            modo = Modo.objects.filter(id=data.get('modo')).first()
            motivo_ingreso = Motivo.objects.filter(id=data.get('motivo_ingreso')).first()
            departamento = Departamento.objects.filter(id=data.get('departamento')).first() if data.get('departamento') else None
            prioridad = Prioridad.objects.filter(id=data.get('prioridad')).first() if data.get('prioridad') else None
            medio_reclutamiento = MedioReclutamiento.objects.filter(id=data.get('medio_reclutamiento')).first() if data.get('medio_reclutamiento') else None
            tipo_contrato = TipoContrato.objects.filter(id=data.get('tipo_contrato')).first() if data.get('tipo_contrato') else None

            # Construir unidad_puesto (puede ser None)
            unidad_puesto = f"{unidad_negocio.nombre_unidad_de_negocio} - {puesto.nombre_puestos}" if unidad_negocio and puesto else None

            # Consultar ControlDeTiempo para obtener el tiempo de cobertura (puede ser None)
            control_de_tiempo = ControlDeTiempo.objects.filter(
                unidad_de_negocio=unidad_negocio,
                puestos=puesto,
                unid_puesto=unidad_puesto
            ).first()
            tiempo_cobertura = control_de_tiempo.tiempo if control_de_tiempo else None

            # Manejo de las fechas
            fecha_solicitud = parser.parse(data.get('fecha_solicitud')) if data.get('fecha_solicitud') else datetime.now()
            fecha_cobertura = parser.parse(data.get('fecha_cobertura')) if data.get('fecha_cobertura') else datetime.now()
            fecha_ingreso = parser.parse(data.get('fecha_ingreso')) if data.get('fecha_ingreso') else datetime.now()

            # Calcular tiempos efectivos
            tiempo_efectivo_cobertura = (fecha_cobertura - fecha_solicitud).days if fecha_solicitud and fecha_cobertura else None
            tiempo_efectivo_cobertura = tiempo_efectivo_cobertura if tiempo_efectivo_cobertura > 0 else None

            tiempo_efectivo_fecha_ingreso = (fecha_ingreso - fecha_solicitud).days if fecha_solicitud and fecha_ingreso else None
            tiempo_efectivo_fecha_ingreso = tiempo_efectivo_fecha_ingreso if tiempo_efectivo_fecha_ingreso > 0 else None

            # Calcular fecha límite de cobertura
            fecha_limite_cobertura_str = (fecha_solicitud + timedelta(days=tiempo_cobertura)).date() if fecha_solicitud and tiempo_cobertura else None

            # Capturar y convertir las fechas
            fecha_limite_cobertura_cal = datetime.combine(fecha_limite_cobertura_str, datetime.min.time()) if fecha_limite_cobertura_str else None
            fecha_cobertura_cal = parser.parse(data.get('fecha_cobertura')) if data.get('fecha_cobertura') else None

            # Calcular el tiempo disponible
            if fecha_limite_cobertura_cal and fecha_cobertura_cal:
                tiempo_disponible = (fecha_limite_cobertura_cal - fecha_cobertura_cal).days
                tiempo_disponible = tiempo_disponible if tiempo_disponible >= 0 else None
            else:
                tiempo_disponible = None

            # Determinar el estatus
            estatus = 'EN PROCESO' if not data.get('fecha_cobertura') else 'CERRADA' if 'fecha_ingreso' in data else 'ESPERA DE INGRESO'

            # Asignar cantidad solicitada y cantidad cubierta
            cantidad_solicitada = 1
            cantidad_cubierta = 1 if estatus == 'CERRADA' else None

            # Manejo de la imagen
            imagen_dni = request.FILES.get('imagen_dni')
            if imagen_dni:
                # Eliminar la imagen anterior si existe
                if control_plaza.nombreimagen:
                    old_image_path = os.path.join('reclutamiento/static/img/controlplazas', control_plaza.nombreimagen)
                    if os.path.isfile(old_image_path):
                        os.remove(old_image_path)

                # Guardar la nueva imagen
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen_dni.name}"  # Toma solo los primeros 4 caracteres
                ruta_imagen = f"reclutamiento/static/img/controlplazas/{nombre_imagen}"

                with open(ruta_imagen, 'wb+') as destino:
                    for chunk in imagen_dni.chunks():
                        destino.write(chunk)

                # Actualizar el nombre y la ruta de la imagen
                control_plaza.ruta = ruta_imagen
                control_plaza.nombreimagen = nombre_imagen

            imagen_dni_reverso = request.FILES.get('imagen_dni_reverso')
            if imagen_dni_reverso:
                # Eliminar la imagen anterior si existe
                if control_plaza.nombreimagen1:  # Asegúrate de usar el campo correcto
                    old_image_path_reverso = os.path.join('reclutamiento/static/img/controlplazas', control_plaza.nombreimagen1)
                    if os.path.isfile(old_image_path_reverso):
                        os.remove(old_image_path_reverso)

                # Guardar la nueva imagen del reverso
                nombre_imagen_reverso = f"{uuid.uuid4().hex[:4]}_{imagen_dni_reverso.name}"  # Toma solo los primeros 4 caracteres
                ruta_imagen_reverso = f"reclutamiento/static/img/controlplazas/{nombre_imagen_reverso}"

                with open(ruta_imagen_reverso, 'wb+') as destino:
                    for chunk in imagen_dni_reverso.chunks():
                        destino.write(chunk)

                # Actualizar el nombre y la ruta de la imagen del reverso
                control_plaza.ruta1 = ruta_imagen_reverso
                control_plaza.nombreimagen1 = nombre_imagen_reverso

            # Actualizar el registro en la base de datos
            control_plaza.empresa = empresa
            control_plaza.sucursal = sucursal
            control_plaza.unidad_de_negocio = unidad_negocio
            control_plaza.año = data.get('año')
            control_plaza.mes_corte = data.get('mes_corte')
            control_plaza.mes_solicitud = data.get('mes_solicitud')
            control_plaza.modo = modo
            control_plaza.motivo_ingreso = motivo_ingreso
            control_plaza.nombre_reemplazo = data.get('nombre_reemplazo')
            control_plaza.puestos = puesto
            control_plaza.departamento = departamento
            control_plaza.prioridad = prioridad
            control_plaza.unidad_puesto = unidad_puesto
            control_plaza.tiempo_cobertura = tiempo_cobertura
            control_plaza.fecha_solicitud = fecha_solicitud
            control_plaza.fecha_inicio_busqueda = parser.parse(data.get('fecha_inicio_busqueda')) if data.get('fecha_inicio_busqueda') else None
            control_plaza.fecha_cobertura = fecha_cobertura
            control_plaza.fecha_ingreso = fecha_ingreso
            control_plaza.tiempo_efectivo_cobertura = tiempo_efectivo_cobertura
            control_plaza.tiempo_efectivo_fecha_ingreso = tiempo_efectivo_fecha_ingreso
            control_plaza.fecha_limite_cobertura = fecha_limite_cobertura_str
            control_plaza.tiempo_disponible = tiempo_disponible
            control_plaza.estatus = estatus
            control_plaza.cantidad_solicitada = cantidad_solicitada
            control_plaza.cantidad_cubierta = cantidad_cubierta
            control_plaza.fuente_reclutamiento = data.get('fuente_reclutamiento')
            control_plaza.nombre_contratado = data.get('nombre_contratado')
            control_plaza.dni = data.get('dni')
            control_plaza.genero = data.get('genero')
            control_plaza.edad = data.get('edad')
            control_plaza.fecha_nacimiento = data.get('fecha_nacimiento')
            control_plaza.medio_reclutamiento = medio_reclutamiento
            control_plaza.salario = data.get('salario')
            control_plaza.combustible = data.get('combustible')
            control_plaza.depreciacion = data.get('depreciacion')
            control_plaza.comision = data.get('comision')
            control_plaza.bono = data.get('bono')
            control_plaza.hospedaje = data.get('hospedaje')
            control_plaza.tipo_contrato = tipo_contrato
            control_plaza.usuario_modifico = request.user

            control_plaza.save()  # Guardar los cambios

            return JsonResponse({'success': True, 'message': 'Plaza actualizada exitosamente'}, status=200)

        except ControlDePlazas.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Control de plazas no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def descargar_imagen(request, imagen_id):
    try:
        registro = ControlDePlazas.objects.get(id=imagen_id)
    except ControlDePlazas.DoesNotExist:
        return HttpResponse(status=404, content="El registro no existe.")

    imagenes = {}

    if registro.nombreimagen:
        imagen_path = Path(settings.BASE_DIR) / registro.ruta
        if imagen_path.exists():
            # Cambiar a una URL accesible a través del servidor
            imagenes['frontal'] = f"/static/img/controlplazas/{registro.nombreimagen}"

    if registro.nombreimagen1:
        imagen_path_reverso = Path(settings.BASE_DIR) / registro.ruta1
        if imagen_path_reverso.exists():
            # Cambiar a una URL accesible a través del servidor
            imagenes['reverso'] = f"/static/img/controlplazas/{registro.nombreimagen1}"

    if not imagenes:
        return HttpResponse(status=404, content="No hay imágenes disponibles para descargar.")

    return JsonResponse({'success': True, 'imagenes': imagenes})

class ExportExcelPlazas(View):
    def get(self, request, *args, **kwargs):
        template_path = os.path.join('reclutamiento/static/templates/formatoplazaoriginal.xlsx')
        
        # Cargar la plantilla de Excel
        try:
            wb = load_workbook(template_path)
            sheet = wb['CUADRO GENERAL DE PLAZAS']  # Ajusta esto según el nombre de tu hoja
        except Exception as e:
            return HttpResponse(f"Error al cargar la plantilla de Excel: {e}", status=500)

        # Obtener los datos de ControlDePlazas
        control_de_plazas = ControlDePlazas.objects.all().order_by('-id')
        
        # Rango de la fila donde se comenzarán a escribir los datos
        start_row = 9
        
        # Escribir los datos en la hoja de Excel
        for idx, plaza in enumerate(control_de_plazas, start=start_row):
            try:
                sheet.cell(row=idx, column=2, value=plaza.analista)
                sheet.cell(row=idx, column=3, value=plaza.sucursal.nombre_sucursal if plaza.sucursal else '')
                sheet.cell(row=idx, column=4, value=plaza.empresa.nombre_empresa if plaza.empresa else '')
                sheet.cell(row=idx, column=5, value=plaza.unidad_de_negocio.nombre_unidad_de_negocio if plaza.unidad_de_negocio else '')
                sheet.cell(row=idx, column=6, value=plaza.año)
                sheet.cell(row=idx, column=7, value=plaza.mes_corte)
                sheet.cell(row=idx, column=8, value=plaza.mes_solicitud)
                sheet.cell(row=idx, column=9, value=plaza.modo.nombre_modo if plaza.modo else '')
                sheet.cell(row=idx, column=10, value=plaza.motivo_ingreso.nombre_motivo if plaza.motivo_ingreso else '')
                sheet.cell(row=idx, column=11, value=plaza.nombre_reemplazo)
                sheet.cell(row=idx, column=12, value=plaza.puestos.nombre_puestos if plaza.puestos else '')
                sheet.cell(row=idx, column=13, value=plaza.departamento.nombre_departamento if plaza.departamento else '')
                sheet.cell(row=idx, column=14, value=plaza.prioridad.nombre_prioridad if plaza.prioridad else '')
                sheet.cell(row=idx, column=15, value=plaza.unidad_puesto)
                sheet.cell(row=idx, column=16, value=plaza.tiempo_cobertura)
                sheet.cell(row=idx, column=17, value=plaza.fecha_solicitud)
                sheet.cell(row=idx, column=18, value=plaza.fecha_inicio_busqueda)
                sheet.cell(row=idx, column=19, value=plaza.fecha_cobertura)
                sheet.cell(row=idx, column=20, value=plaza.fecha_ingreso)
                sheet.cell(row=idx, column=21, value=plaza.tiempo_efectivo_cobertura)
                sheet.cell(row=idx, column=22, value=plaza.tiempo_efectivo_fecha_ingreso)
                sheet.cell(row=idx, column=23, value=plaza.fecha_limite_cobertura)
                sheet.cell(row=idx, column=24, value=plaza.tiempo_disponible)
                sheet.cell(row=idx, column=25, value=plaza.estatus)
                sheet.cell(row=idx, column=26, value=plaza.cantidad_solicitada)
                sheet.cell(row=idx, column=27, value=plaza.cantidad_cubierta)
                sheet.cell(row=idx, column=28, value=plaza.fuente_reclutamiento)
                sheet.cell(row=idx, column=29, value=plaza.nombre_contratado)
                sheet.cell(row=idx, column=30, value=plaza.dni)
                sheet.cell(row=idx, column=31, value=plaza.genero)
                sheet.cell(row=idx, column=32, value=plaza.edad)
                sheet.cell(row=idx, column=33, value=plaza.medio_reclutamiento.nombre_medio_de_reclutamiento if plaza.medio_reclutamiento else '')
                sheet.cell(row=idx, column=34, value=plaza.salario)
                sheet.cell(row=idx, column=35, value=plaza.combustible)
                sheet.cell(row=idx, column=36, value=plaza.depreciacion)
                sheet.cell(row=idx, column=37, value=plaza.comision)
                sheet.cell(row=idx, column=38, value=plaza.bono)
                sheet.cell(row=idx, column=39, value=plaza.hospedaje)
                sheet.cell(row=idx, column=40, value=plaza.tipo_contrato.nombre_tipo_de_contrato if plaza.tipo_contrato else '')
            except Exception as e:
                return HttpResponse(f"Error escribiendo datos en Excel en la fila {idx}: {e}", status=500)

        # Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Control_de_Plazas.xlsx"'

        # Guardar el libro de trabajo en la respuesta
        wb.save(response)

        return response

#--------BOLSA DE EMPLEO VIEW--------#  
@csrf_exempt
@login_required(login_url='/login/')
def bolsaempleo_view(request, id=None):
    if request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_bolsaempleos'):
            return redirect('sinacceso')
        
        # Obtener ciudades y medios de reclutamiento activos
        ciudades = Ciudades.objects.filter(activo=True).order_by('nombre_ciudades')
        medios_reclutamiento = MedioReclutamiento.objects.filter(activo=True).order_by('nombre_medio_de_reclutamiento')

        # Obtener todos los puestos
        puestos = Puestos.objects.all()

        # Obtener solo los puestos que están en BolsaEmpleos
        puestos_ids = set()
        for empleo in BolsaEmpleos.objects.all():
            puestos_ids.update(empleo.puestosaspira or [])
            puestos_ids.update(empleo.puestosaplica or [])
        
        puestos_en_bolsa = Puestos.objects.filter(id__in=puestos_ids)

        # Obtener los distintos estados únicos presentes en la base de datos
        estados = BolsaEmpleos.objects.values_list('estado', flat=True).distinct()

        # Parámetros de búsqueda
        search = request.GET.get('search', '').upper()  # Convertir a mayúsculas para evitar problemas de case-sensitive
        estado = request.GET.getlist('estado', [])  # Obtener la lista de estados seleccionados
        puestos_aspira = request.GET.getlist('puestosaspira', [])
        puestos_aplica = request.GET.getlist('puestosaplica', [])
        ciudad_id = request.GET.getlist('ciudad', [])  # Filtro de ciudad

        # Obtener la queryset inicial
        bolsa_empleos = BolsaEmpleos.objects.all().prefetch_related('ciudad', 'medio_reclutamiento').order_by('-id')

        # Filtrar por ciudad si se proporciona y no está vacío
        if ciudad_id and all(ciudad_id):  # Asegúrate de que ciudad_id no tenga valores vacíos
            bolsa_empleos = bolsa_empleos.filter(ciudad__id__in=ciudad_id)

        # Filtrar por búsqueda
        if search:
            bolsa_empleos = bolsa_empleos.filter(
                Q(id__icontains=search) |
                Q(dni__icontains=search) |
                Q(nombre_candidato__icontains=search) |
                Q(telefono__icontains=search) |
                Q(telefono2__icontains=search) |
                Q(estado__icontains=search) |
                Q(ciudad__nombre_ciudades__icontains=search) |
                Q(medio_reclutamiento__nombre_medio_de_reclutamiento__icontains=search) |
                Q(edad__icontains=search) |
                Q(fecha_nacimiento__icontains=search) |
                Q(estadocivil__icontains=search) |
                Q(nhijos__icontains=search) |
                Q(direccion__icontains=search) |
                Q(mediomovilizacion__icontains=search) |
                Q(experiencia__icontains=search) |
                Q(observacion__icontains=search) |
                Q(ruta__icontains=search) |
                Q(nombredocumento__icontains=search) |
                Q(puestosaspira__icontains=search) |  # Busca en puestosaspira (parcial)
                Q(puestosaplica__icontains=search)    # Busca en puestosaplica (parcial)
            )

        # Filtrar por estado (múltiples estados seleccionados)
        if estado:
            bolsa_empleos = bolsa_empleos.filter(estado__in=estado)

        # Filtrado manual por puestos que aspira
        if puestos_aspira:
            bolsa_empleos = [
                empleo for empleo in bolsa_empleos
                if any(int(puesto) in (int(p) for p in empleo.puestosaspira or []) for puesto in puestos_aspira)
            ]

        # Filtrado manual por puestos que aplica
        if puestos_aplica:
            bolsa_empleos = [
                empleo for empleo in bolsa_empleos
                if any(int(puesto) in (int(p) for p in (empleo.puestosaplica or [])) for puesto in puestos_aplica)
            ]

        # Paginación: mostrar 10 registros por página
        paginator = Paginator(bolsa_empleos, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calcular el rango de páginas a mostrar en la paginación
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        # Convertir IDs de puestosaspira y puestosaplica a nombres de puestos
        for empleo in page_obj:
            empleo.puestosaspira_nombres = [
                Puestos.objects.get(id=puesto_id).nombre_puestos 
                for puesto_id in (empleo.puestosaspira or [])
            ]
            
            empleo.puestosaplica_nombres = [
                Puestos.objects.get(id=puesto_id).nombre_puestos 
                for puesto_id in (empleo.puestosaplica or [])
            ]

        # Contexto para renderizar la plantilla
        context = {
            'puestos': puestos,
            'puestos_en_bolsa': puestos_en_bolsa,
            'ciudades': ciudades,
            'medios_reclutamiento': medios_reclutamiento,
            'bolsa_empleos': page_obj,
            'page_range': page_range,
            'search': search,
            'estado': estado,  # Incluir el filtro de estado en el contexto
            'puestos_aspira': puestos_aspira,
            'puestos_aplica': puestos_aplica,
            'ciudad_id': ciudad_id,
            'estados': estados,  # Incluir los estados disponibles para el filtro
        }
        return render(request, 'bolsaempleo.html', context)
    elif request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_bolsaempleos'):
                return JsonResponse({'error': 'No tienes permisos para agregar candidatos'}, status=403)
            
            data = json.loads(request.POST.get('data'))  # Datos JSON del formulario
            dni = data.get('dni')
            nombre_candidato = data.get('nombre_candidato')
            puestosaspira = data.get('puestoaspira', [])
            telefono = data.get('telefono')
            telefono2 = data.get('telefono2')
            estado = data.get('estado')
            ciudad_id = data.get('ciudad')
            medio_reclutamiento_id = data.get('mediosReclutamiento')
            edad = data.get('edad') or None  # Asignar None si está vacío
            fecha_nacimiento = data.get('fechanacimiento') or None  # Asignar None si está vacío
            estadocivil = data.get('estadocivil')
            nhijos = data.get('nhijos') or None  # Asignar None si está vacío
            direccion = data.get('direccion')
            experiencia = data.get('experiencia')
            observacion = data.get('observacion')
            mediomovilizacion = data.get('mediomovilizacion')
            archivo = request.FILES.get('cv')  # Archivo PDF subido

            # Validación de campos obligatorios
            if not dni or not nombre_candidato or not telefono or not puestosaspira or not estado:
                return JsonResponse({'success': False, 'message': 'Los campos DNI, Nombre del Candidato, Teléfono, Puesto al que Aspira y Estado son obligatorios.'}, status=400)
            
            # Validar que telefono y telefono2 no sean iguales entre sí
            if telefono == telefono2:
                return JsonResponse({'success': False, 'message': 'El teléfono principal y el secundario no pueden ser iguales.'}, status=400)

            # Validar que telefono y telefono2 no existan ya en la base de datos
            if BolsaEmpleos.objects.filter(telefono=telefono).exists() or BolsaEmpleos.objects.filter(telefono2=telefono).exists():
                return JsonResponse({'success': False, 'message': 'El número de teléfono principal ya está registrado.'}, status=400)

            if telefono2 and (BolsaEmpleos.objects.filter(telefono=telefono2).exists() or BolsaEmpleos.objects.filter(telefono2=telefono2).exists()):
                return JsonResponse({'success': False, 'message': 'El número de teléfono secundario ya está registrado.'}, status=400)
            
            # Validar formato de fecha
            if fecha_nacimiento:
                try:
                    # Intentar convertir a objeto de fecha
                    fecha_nacimiento = timezone.datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'El formato de la fecha de nacimiento es inválido. Debe ser YYYY-MM-DD.'}, status=400)

            # Guardar archivo en ruta especificada con nombre único
            ruta = None
            nombredocumento = None
            if archivo:
                if archivo.content_type != 'application/pdf':
                    return JsonResponse({'success': False, 'message': 'Solo se permiten documentos en formato PDF.'}, status=400)
                
                # Generar nombre único para el archivo con números aleatorios
                random_str = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                nuevo_nombre_archivo = f"{random_str}_{archivo.name}"
                
                # Guardar archivo en reclutamiento/static/document
                ruta_relativa = os.path.join('reclutamiento', 'static', 'document', nuevo_nombre_archivo)
                ruta_archivo = os.path.join(settings.BASE_DIR, ruta_relativa)
                
                # Crear directorio si no existe
                os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)
                
                # Guardar el archivo en la ruta especificada
                with open(ruta_archivo, 'wb+') as destination:
                    for chunk in archivo.chunks():
                        destination.write(chunk)
                
                ruta = ruta_relativa
                nombredocumento = nuevo_nombre_archivo

            # Crear nueva entrada en BolsaEmpleos
            bolsa_empleo = BolsaEmpleos.objects.create(
                dni=dni,
                nombre_candidato=nombre_candidato,
                puestosaspira=puestosaspira,
                telefono=telefono,
                telefono2=telefono2,
                estado=estado,
                ciudad_id=ciudad_id,
                medio_reclutamiento_id=medio_reclutamiento_id,
                edad=edad,
                fecha_nacimiento=fecha_nacimiento,
                estadocivil=estadocivil,
                nhijos=nhijos,
                direccion=direccion,
                experiencia=experiencia,
                observacion=observacion,
                mediomovilizacion=mediomovilizacion,
                ruta=ruta,
                nombredocumento=nombredocumento,
                usuario_creo=request.user
            )
            return JsonResponse({'success': True, 'message': 'Candidato registrado exitosamente'}, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
          
def bolsaempleo_update_post_view(request, id=None):
    if request.method == 'POST' and id is not None:
        try:
            if not request.user.has_perm('Reclutamiento.change_bolsaempleos'):
                return JsonResponse({'error': 'No tienes permisos para editar candidatos'}, status=403)
            data_json = request.POST.get('data')

            if not data_json:
                return JsonResponse({'success': False, 'message': 'No se recibieron datos.'}, status=400)

            data = json.loads(data_json)

            dni = data.get('dni')
            nombre_candidato = data.get('nombre_candidato')
            puestosaspira = data.get('puestoaspira', [])
            puestosaplica = data.get('puestoaplica', [])
            telefono = data.get('telefono')
            estado = data.get('estado')

            # Validación de campos obligatorios
            if not dni or not nombre_candidato or not telefono or not puestosaspira or not estado:
                return JsonResponse({'success': False, 'message': 'Los campos DNI, Nombre del Candidato, Teléfono, Puesto al que Aspira y Estado son obligatorios.'}, status=400)

            # Buscar el candidato en la base de datos
            try:
                bolsa_empleo = BolsaEmpleos.objects.get(id=id)
            except BolsaEmpleos.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'El candidato no existe.'}, status=404)

            # Validar formato de fecha de nacimiento si se proporciona
            fecha_nacimiento = data.get('fechanacimiento')
            if fecha_nacimiento:
                try:
                    fecha_nacimiento = timezone.datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'El formato de la fecha de nacimiento es inválido. Debe ser YYYY-MM-DD.'}, status=400)

            # Obtener el archivo subido
            archivo = request.FILES.get('cveditar')  # Archivo PDF subido, si existe

            # Si se sube un nuevo archivo, eliminar el anterior y guardar el nuevo
            if archivo:
                if archivo.content_type != 'application/pdf':
                    return JsonResponse({'success': False, 'message': 'Solo se permiten documentos en formato PDF.'}, status=400)

                # Eliminar archivo anterior si existe
                if bolsa_empleo.ruta and os.path.exists(os.path.join(settings.BASE_DIR, bolsa_empleo.ruta)):
                    os.remove(os.path.join(settings.BASE_DIR, bolsa_empleo.ruta))

                # Generar nombre único para el nuevo archivo
                random_str = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                nuevo_nombre_archivo = f"{random_str}_{archivo.name}"

                # Guardar nuevo archivo
                ruta_relativa = os.path.join('reclutamiento', 'static', 'document', nuevo_nombre_archivo)
                ruta_archivo = os.path.join(settings.BASE_DIR, ruta_relativa)

                os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)

                with open(ruta_archivo, 'wb+') as destination:
                    for chunk in archivo.chunks():
                        destination.write(chunk)

                bolsa_empleo.ruta = ruta_relativa
                bolsa_empleo.nombredocumento = nuevo_nombre_archivo

            # Actualizar los campos del modelo
            bolsa_empleo.dni = dni
            bolsa_empleo.nombre_candidato = nombre_candidato
            bolsa_empleo.puestosaspira = puestosaspira
            bolsa_empleo.puestosaplica = puestosaplica
            bolsa_empleo.telefono = telefono
            bolsa_empleo.estado = estado
            bolsa_empleo.telefono2 = data.get('telefono2') or bolsa_empleo.telefono2
            bolsa_empleo.ciudad_id = data.get('ciudad') or bolsa_empleo.ciudad_id
            bolsa_empleo.medio_reclutamiento_id = data.get('mediosReclutamiento') or bolsa_empleo.medio_reclutamiento_id
            bolsa_empleo.edad = data.get('edad') or bolsa_empleo.edad
            bolsa_empleo.fecha_nacimiento = fecha_nacimiento or bolsa_empleo.fecha_nacimiento
            bolsa_empleo.estadocivil = data.get('estadocivil') or bolsa_empleo.estadocivil
            nhijos_value = data.get('nhijos')
            if nhijos_value is None:
                bolsa_empleo.nhijos = None  # Asignar None si el valor es None
            else:
                bolsa_empleo.nhijos = nhijos_value            
            bolsa_empleo.direccion = data.get('direccion') or bolsa_empleo.direccion
            bolsa_empleo.experiencia = data.get('experiencia') or bolsa_empleo.experiencia
            bolsa_empleo.observacion = data.get('observacion') or bolsa_empleo.observacion
            bolsa_empleo.mediomovilizacion = data.get('mediomovilizacion') or bolsa_empleo.mediomovilizacion
            bolsa_empleo.usuario_modifico = request.user
            # Guardar los cambios en la base de datos
            bolsa_empleo.save()

            return JsonResponse({'success': True, 'message': 'Candidato actualizado exitosamente'}, status=200)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def descargar_pdf(request, archivo_id):
    try:
        registro = BolsaEmpleos.objects.get(id=archivo_id)
    except BolsaEmpleos.DoesNotExist:
        return HttpResponse(status=404, content="El archivo no existe.")

    # Verificar si el registro tiene una ruta válida
    if not registro.ruta:
        return HttpResponse(status=404, content="El archivo no tiene una ruta válida.")

    # Obtener la ruta completa del archivo
    archivo_path = Path(settings.BASE_DIR) / registro.ruta

    # Verificar si el archivo existe en el sistema de archivos
    if not archivo_path.exists():
        return HttpResponse(status=404, content="El archivo no se encuentra en el sistema de archivos.")

    # Leer el archivo y prepararlo para la descarga
    with open(archivo_path, 'rb') as archivo:
        response = HttpResponse(archivo.read(), content_type=mimetypes.guess_type(str(archivo_path))[0])
        response['Content-Disposition'] = f'attachment; filename="{archivo_path.name}"'
        return response

def exportar_bolsa_empleos(request):
    # Obtener todos los registros de BolsaEmpleos
    datos = (
        BolsaEmpleos.objects
        .prefetch_related('ciudad', 'medio_reclutamiento')
        .values(
            'id', 
            'dni', 
            'nombre_candidato',
            'puestosaspira',  
            'puestosaplica',
            'telefono', 
            'telefono2', 
            'estado', 
            'ciudad__nombre_ciudades', 
            'medio_reclutamiento__nombre_medio_de_reclutamiento', 
            'edad', 
            'fecha_nacimiento', 
            'estadocivil', 
            'nhijos', 
            'direccion', 
            'mediomovilizacion', 
            'experiencia', 
            'observacion',
            'creado'  
        )
    )

    # Convertir los datos a un DataFrame de pandas
    df = pd.DataFrame(list(datos))
    
    # Convertir IDs a nombres de puestos
    if 'puestosaspira' in df.columns:
        df['puestosaspira'] = df['puestosaspira'].apply(lambda x: 
            ', '.join(Puestos.objects.get(id=puesto_id).nombre_puestos for puesto_id in x) 
            if isinstance(x, list) else ''
        )
    else:
        df['puestosaspira'] = ''

    if 'puestosaplica' in df.columns:
        df['puestosaplica'] = df['puestosaplica'].apply(lambda x: 
            ', '.join(Puestos.objects.get(id=puesto_id).nombre_puestos for puesto_id in x) 
            if isinstance(x, list) else ''
        )
    else:
        df['puestosaplica'] = ''

    # Crear un archivo Excel
    excel_file = "bolsa_empleos.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={excel_file}'

    # Escribir el DataFrame en el archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='BolsaEmpleos')

    return response

#--------CONTRATACIONES VIEW--------#  
def contratacionesform_view(request):
    if request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_contratacionempleados'):
            return redirect('sinacceso')
        
        all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()
        all_empresas = Empresas.objects.filter(activo=True).order_by('nombre_empresa').distinct()
        all_unidades_negocio = Unidad_Negocio.objects.filter(activo=True).order_by('nombre_unidad_de_negocio').distinct()
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
        all_puestos = Puestos.objects.filter(activo=True).order_by('nombre_puestos').distinct()

        # Obtener opciones para Modo, Motivo, Prioridad, Tipo de Contrato y Medio de Reclutamiento
        all_modos = Modo.objects.filter(activo=True).order_by('nombre_modo').distinct()
        all_motivos = Motivo.objects.filter(activo=True).order_by('nombre_motivo').distinct()
        all_prioridades = Prioridad.objects.filter(activo=True).order_by('nombre_prioridad').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(activo=True).order_by('nombre_tipo_de_contrato').distinct()
        all_mediosReclutamiento = MedioReclutamiento.objects.filter(activo=True).order_by('nombre_medio_de_reclutamiento').distinct()

        all_municipios = MunicipioPais.objects.filter(activo=True).order_by('nombre_municipio').distinct()
        all_departamentos_hn = DepartamentoPais.objects.filter(activo=True).order_by('nombre_departamento').distinct()
        all_paises = Pais.objects.filter(activo=True).order_by('nombre_pais').distinct()

        context = {
            'all_sucursales': all_sucursales,
            'all_empresas': all_empresas,
            'all_unidades_negocio': all_unidades_negocio,
            'all_departamentos': all_departamentos,
            'all_puestos': all_puestos,
            'all_modos': all_modos,
            'all_motivos': all_motivos,
            'all_prioridades': all_prioridades,
            'all_tipos_contrato': all_tipos_contrato,
            'all_mediosReclutamiento': all_mediosReclutamiento,
            'all_municipios': all_municipios, 
            'all_departamentos_hn': all_departamentos_hn,
            'all_paises': all_paises,
        }
        return render(request, 'contrataciones/registrarcontrataciones.html', context)
    
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_contratacionempleados'):
                return JsonResponse({'error': 'No tienes permisos para agregar contrataciones'}, status=403)
            
            primer_nombre = request.POST.get('primerNombre') or None
            segundo_nombre = request.POST.get('segundoNombre') or None
            primer_apellido = request.POST.get('primerApellido') or None
            segundo_apellido = request.POST.get('segundoApellido') or None
            fecha_nacimiento = request.POST.get('fechaNacimiento') or None
            municipio = request.POST.get('municipio') or None
            pais = request.POST.get('pais') or None
            genero = request.POST.get('genero') or None
            direccion = request.POST.get('direccion') or None
            dni = request.POST.get('cedula') or None
            estado_civil = request.POST.get('estadoCivil') or None
            hijos = 1 if request.POST.get('hijos') == '1' else 0
            profesion = request.POST.get('profesion') or None
            correo = request.POST.get('correo') or None
            departamento = request.POST.get('departamento') or None
            telefono = request.POST.get('telefono') or None

            # Datos de emergencia
            emergencia1 = request.POST.get('emergencia1') or None
            parentesco1 = request.POST.get('parentesco1') or None
            telefono_emergencia1 = request.POST.get('telefonoEmergencia1') or None
            emergencia2 = request.POST.get('emergencia2') or None
            parentesco2 = request.POST.get('parentesco2') or None
            telefono_emergencia2 = request.POST.get('telefonoEmergencia2') or None

            # Nivel educativo
            nivel_educativo = request.POST.get('alfabeta') or None
            ultimo_grado_estudio = request.POST.get('ultimoGrado') or None
            ultimogradodetalle = request.POST.get('ultimogradodetalle') or None
            padecimiento = 1 if request.POST.get('padecimiento') == '1' else 0
            detalle_enfermedad = request.POST.get('detalleEnfermedad') or None

            # Datos laborales
            puesto = request.POST.get('puesto') or None
            unidad_negocio = request.POST.get('unidadnegocio') or None
            salario = request.POST.get('salario') or None
            comision = request.POST.get('comision') or None
            bofa = request.POST.get('bofa') or None
            sucursal = request.POST.get('sucursal') or None
            tipo_contrato = request.POST.get('tipocontrato') or None
            fecha_ingreso = request.POST.get('fechaIngreso') or None
            fecha_vencimiento = request.POST.get('fechavencimiento') or None
            departamento_empresa = request.POST.get('departamentoEmpresa') or None
            direccion_empresa = request.POST.get('direccionempresa') or None
            nombre_empresa = request.POST.get('nombreempresa') or None
            telefono_empresa = request.POST.get('telefonoempresa') or None

            # Beneficiario
            nombre_beneficiario = request.POST.get('nombreBeneficiario') or None
            identidad_beneficiario = request.POST.get('identidadBeneficiario') or None
            parentesco_beneficiario = request.POST.get('parentescoBeneficiario') or None
            porcentaje_beneficiario = request.POST.get('porcentajeBeneficiario') or None

            imagen = request.FILES.get('imagen')
            nombre_imagen = None
            ruta_imagen = None

            if imagen:
                # Si hay imagen, generar un nombre único para el archivo
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"  # Usando UUID para nombre único
                ruta_imagen = f"reclutamiento/static/img/contrataciones/{nombre_imagen}"

                # Crear directorio si no existe
                try:
                    os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                    # Guardar la imagen en la ruta especificada
                    with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                        for chunk in imagen.chunks():
                            destino.write(chunk)
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Error al guardar la imagen: {str(e)}'}, status=500)
            else:
                # Si no hay imagen, mantener nombre_imagen y ruta_imagen como None
                nombre_imagen = None
                ruta_imagen = None

            # Crear nueva entrada en ContratacionEmpleados
            contratacion = ContratacionEmpleados.objects.create(
                tipo_contratacion=request.POST.get('tipoIngreso') or None,
                nombre1=primer_nombre,
                nombre2=segundo_nombre,
                apellido1=primer_apellido,
                apellido2=segundo_apellido,
                fecha_nacimiento=fecha_nacimiento,
                municipio_id=municipio,
                pais_id=pais,
                genero=genero,
                direccionexacta=direccion,
                dni=dni,
                estado_civil=estado_civil,
                hijos=hijos,
                profecion_oficio=profesion,
                correo=correo,
                departamento_id=departamento,
                telefono=telefono,
                nombre1_emergencia=emergencia1,
                parentesco1=parentesco1,
                telefonoemergencia1=telefono_emergencia1,
                nombre2_emergencia=emergencia2,
                parentesco2=parentesco2,
                telefonoemergencia2=telefono_emergencia2,
                nivel_educativo=nivel_educativo,
                ultimo_grado_estudio=ultimo_grado_estudio,
                ultimogradodetalle=ultimogradodetalle,
                padecimiento=padecimiento,
                detalle_enfermedad=detalle_enfermedad,
                puestos_id=puesto,
                unidad_de_negocio_id=unidad_negocio,
                sucursal_id=sucursal,
                departamento_empresa_id=departamento_empresa,
                tipo_contrato_id=tipo_contrato,
                salario=salario,
                comision=comision,
                bofa=bofa,
                fecha_ingreso=fecha_ingreso,
                fecha_vencimiento=fecha_vencimiento,
                direccionempresa=direccion_empresa,
                nombre_empresa_id=nombre_empresa,
                telefono_empresa=telefono_empresa,
                nombre_beneficiario=nombre_beneficiario,
                dni_beneficiario=identidad_beneficiario,
                parentesco_beneficiario=parentesco_beneficiario,
                porcentaje=porcentaje_beneficiario,
                ruta=ruta_imagen,
                nombreimagen=nombre_imagen,
                usuario_creo = request.user,
            )

            return JsonResponse({'success': True, 'message': 'Candidato registrado exitosamente'}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        
def contratacionestable_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')

        contrataciones = ContratacionEmpleados.objects.select_related(
            'puestos', 
            'unidad_de_negocio', 
            'sucursal', 
            'municipio',
        )

        # Obtener valores únicos para los filtros
        municipios_options = set(contrataciones.values_list('municipio__nombre_municipio', flat=True))
        puestos_options = set(contrataciones.values_list('puestos__nombre_puestos', flat=True))
        unidades_negocio_options = set(contrataciones.values_list('unidad_de_negocio__nombre_unidad_de_negocio', flat=True))
        sucursales_options = set(contrataciones.values_list('sucursal__nombre_sucursal', flat=True))
        fechas_ingreso_options = set(contrataciones.values_list('fecha_ingreso', flat=True))

        # Filtrar por los filtros seleccionados (si se aplican)
        municipio_filter = request.GET.getlist('municipio')
        puesto_filter = request.GET.getlist('puesto')
        unidad_negocio_filter = request.GET.getlist('unidad_negocio')
        sucursal_filter = request.GET.getlist('sucursal')
        fecha_ingreso_filter = request.GET.getlist('fecha_ingreso')

        # Aplicar filtros a la consulta
        if municipio_filter:
            contrataciones = contrataciones.filter(municipio__nombre_municipio__in=municipio_filter)
        
        if puesto_filter:
            contrataciones = contrataciones.filter(puestos__nombre_puestos__in=puesto_filter)
        
        if unidad_negocio_filter:
            contrataciones = contrataciones.filter(unidad_de_negocio__nombre_unidad_de_negocio__in=unidad_negocio_filter)
        
        if sucursal_filter:
            contrataciones = contrataciones.filter(sucursal__nombre_sucursal__in=sucursal_filter)

        if fecha_ingreso_filter:
            contrataciones = contrataciones.filter(fecha_ingreso__in=fecha_ingreso_filter)

        if search:
            contrataciones = contrataciones.filter(
                models.Q(nombre1__icontains=search) |
                models.Q(nombre2__icontains=search) |
                models.Q(apellido1__icontains=search) |
                models.Q(apellido2__icontains=search) |
                models.Q(telefono__icontains=search) |
                models.Q(dni__icontains=search) |
                models.Q(correo__icontains=search)
            )

        # Paginación
        paginator = Paginator(contrataciones, 10) 
        page_number = request.GET.get('page') 
        page_obj = paginator.get_page(page_number) 

        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1) 

        context = {
            'contrataciones': page_obj, 
            'search': search,  
            'page_range': page_range,
            'municipios_options': list(municipios_options),
            'puestos_options': list(puestos_options),
            'unidades_negocio_options': list(unidades_negocio_options),
            'sucursales_options': list(sucursales_options),
            'fechas_ingreso_options': list(fechas_ingreso_options),
            'municipio_filter': municipio_filter,
            'puesto_filter': puesto_filter,
            'unidad_negocio_filter': unidad_negocio_filter,
            'sucursal_filter': sucursal_filter,
            'fecha_ingreso_filter': fecha_ingreso_filter,
        }

        return render(request, 'contrataciones/contrataciones.html', context)
    
def imprimir_contratacion_view(request, contratacion_id):
    contratacion = get_object_or_404(ContratacionEmpleados, id=contratacion_id)
    
    context = {
        'contratacion': contratacion,
    }
    return render(request, 'contrataciones/imprimir_contratacion.html', context)

def updatecontrataciones_view(request, contratacion_id):
    contratacion = get_object_or_404(ContratacionEmpleados, id=contratacion_id)

    # Si es una solicitud GET, se muestra el formulario con los datos actuales
    if request.method == 'GET':
        context = {
            'contratacion': contratacion,
            'all_municipios': MunicipioPais.objects.all(),
            'all_paises': Pais.objects.all(),
            'all_departamentos': DepartamentoPais.objects.all(),
            'all_puestos': Puestos.objects.all(),
            'all_unidades_negocio': Unidad_Negocio.objects.all(),
            'all_sucursales': Sucursal.objects.all(),
            'all_tipos_contrato': TipoContrato.objects.all(),
            'all_empresas': Empresas.objects.all(),
            'all_departamentoempresa': Departamento.objects.all(),
        }
        return render(request, 'contrataciones/updatecontrataciones.html', context)

    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.change_contratacionempleados'):
                return JsonResponse({'error': 'No tienes permisos para editar contrataciones'}, status=403)
            # Obtener campos individuales y asignar None si están vacíos
            primer_nombre = request.POST.get('primerNombre') or None
            segundo_nombre = request.POST.get('segundoNombre') or None
            primer_apellido = request.POST.get('primerApellido') or None
            segundo_apellido = request.POST.get('segundoApellido') or None
            fecha_nacimiento = request.POST.get('fechaNacimiento') or None
            pais = request.POST.get('pais') or None

            # Claves foráneas deben ser números o None
            municipio = request.POST.get('municipio')
            municipio_id = int(municipio) if municipio and municipio.isdigit() else None

            genero = request.POST.get('genero') or None
            direccion = request.POST.get('direccion') or None
            dni = request.POST.get('cedula') or None
            estado_civil = request.POST.get('estadoCivil') or None
            hijos = 1 if request.POST.get('hijos') == '1' else 0
            profesion = request.POST.get('profesion') or None
            correo = request.POST.get('correo') or None

            departamento = request.POST.get('departamento')
            departamento_id = int(departamento) if departamento and departamento.isdigit() else None

            telefono = request.POST.get('telefono') or None

            # Datos laborales
            puesto = request.POST.get('puesto')
            puesto_id = int(puesto) if puesto and puesto.isdigit() else None

            unidad_negocio = request.POST.get('unidadnegocio')
            unidad_negocio_id = int(unidad_negocio) if unidad_negocio and unidad_negocio.isdigit() else None

            sucursal = request.POST.get('sucursal')
            sucursal_id = int(sucursal) if sucursal and sucursal.isdigit() else None

            tipo_contrato = request.POST.get('tipocontrato')
            tipo_contrato_id = int(tipo_contrato) if tipo_contrato and tipo_contrato.isdigit() else None

            fecha_ingreso = request.POST.get('fechaIngreso') or None
            fecha_vencimiento = request.POST.get('fechavencimiento') or None

            departamento_empresa = request.POST.get('departamentoEmpresa')
            departamento_empresa_id = int(departamento_empresa) if departamento_empresa and departamento_empresa.isdigit() else None

            direccion_empresa = request.POST.get('direccionempresa') or None
            nombre_empresa = request.POST.get('nombreempresa')
            nombre_empresa_id = int(nombre_empresa) if nombre_empresa and nombre_empresa.isdigit() else None

            telefono_empresa = request.POST.get('telefonoempresa') or None

            # Beneficiario
            nombre_beneficiario = request.POST.get('nombreBeneficiario') or None
            identidad_beneficiario = request.POST.get('identidadBeneficiario') or None
            parentesco_beneficiario = request.POST.get('parentescoBeneficiario') or None
            porcentaje_beneficiario = request.POST.get('porcentajeBeneficiario') or None

            # Variables para imagen
            imagen = request.FILES.get('imagen')
            nombre_imagen = None
            ruta_imagen = None

            if imagen:
                # Si ya existe una imagen, eliminarla
                if contratacion.ruta:  # Verifica si hay una imagen ya existente en la ruta
                    ruta_anterior = os.path.join(settings.BASE_DIR, contratacion.ruta)
                    if os.path.exists(ruta_anterior):
                        try:
                            os.remove(ruta_anterior)  # Elimina la imagen anterior
                        except Exception as e:
                            return JsonResponse({'success': False, 'message': f'Error al eliminar la imagen anterior: {str(e)}'}, status=500)

                # Generar un nombre único para la nueva imagen
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/contrataciones/{nombre_imagen}"

                # Crear el directorio si no existe
                try:
                    os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                    # Guardar la nueva imagen en la ruta especificada
                    with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                        for chunk in imagen.chunks():
                            destino.write(chunk)

                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Error al guardar la nueva imagen: {str(e)}'}, status=500)

                # Actualizar la ruta y nombre de imagen en el modelo de contratación
                contratacion.ruta = ruta_imagen
                contratacion.nombreimagen = nombre_imagen
            else:
                # Si no se proporciona una nueva imagen, mantener los valores anteriores
                nombre_imagen = contratacion.nombreimagen
                ruta_imagen = contratacion.ruta


            # Valida "padecimiento"
            padecimiento = 1 if request.POST.get('padecimiento') == 'si' else 0

            # Actualizar la entrada existente en ContratacionEmpleados
            contratacion.tipo_contratacion = request.POST.get('tipoIngreso') or None
            contratacion.nombre1 = primer_nombre
            contratacion.nombre2 = segundo_nombre
            contratacion.apellido1 = primer_apellido
            contratacion.apellido2 = segundo_apellido
            contratacion.fecha_nacimiento = fecha_nacimiento
            contratacion.municipio_id = municipio_id
            contratacion.genero = genero
            contratacion.direccionexacta = direccion
            contratacion.dni = dni
            contratacion.estado_civil = estado_civil
            contratacion.hijos = hijos
            contratacion.profecion_oficio = profesion
            contratacion.correo = correo
            contratacion.departamento_id = departamento_id
            contratacion.telefono = telefono
            contratacion.nombre1_emergencia = request.POST.get('emergencia1') or None
            contratacion.parentesco1 = request.POST.get('parentesco1') or None
            contratacion.telefonoemergencia1 = request.POST.get('telefonoEmergencia1') or None
            contratacion.nombre2_emergencia = request.POST.get('emergencia2') or None
            contratacion.parentesco2 = request.POST.get('parentesco2') or None
            contratacion.telefonoemergencia2 = request.POST.get('telefonoEmergencia2') or None
            contratacion.nivel_educativo = request.POST.get('alfabeta') or None
            contratacion.ultimo_grado_estudio = request.POST.get('ultimoGrado') or None
            contratacion.ultimogradodetalle = request.POST.get('ultimogradodetalle') or None
            contratacion.padecimiento = padecimiento
            contratacion.detalle_enfermedad = request.POST.get('detalleEnfermedad') or None
            contratacion.puestos_id = puesto_id
            contratacion.pais_id = pais
            contratacion.usuario_modifico = request.user
            contratacion.unidad_de_negocio_id = unidad_negocio_id
            contratacion.sucursal_id = sucursal_id
            contratacion.departamento_empresa_id = departamento_empresa_id
            contratacion.tipo_contrato_id = tipo_contrato_id
            contratacion.salario = request.POST.get('salario') or None
            contratacion.comision = request.POST.get('comision') or None
            contratacion.bofa = request.POST.get('bofa') or None
            contratacion.fecha_ingreso = fecha_ingreso
            contratacion.fecha_vencimiento = fecha_vencimiento
            contratacion.direccionempresa = direccion_empresa
            contratacion.nombre_empresa_id = nombre_empresa_id
            contratacion.telefono_empresa = telefono_empresa
            contratacion.nombre_beneficiario = nombre_beneficiario
            contratacion.dni_beneficiario = identidad_beneficiario
            contratacion.parentesco_beneficiario = parentesco_beneficiario
            contratacion.porcentaje = porcentaje_beneficiario
            contratacion.ruta = ruta_imagen
            contratacion.nombreimagen = nombre_imagen

            # Guardar cambios en la base de datos
            contratacion.save()

            return JsonResponse({'success': True, 'message': 'Candidato actualizado exitosamente'}, status=200)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

class ExportExcelContratacion(View):
    def get(self, request, *args, **kwargs):
        template_path = os.path.join('reclutamiento/static/templates/FormatoContratacion.xlsx')
        
        # Cargar la plantilla de Excel
        try:
            wb = load_workbook(template_path) 
            sheet = wb.active  # Ajusta si la hoja tiene un nombre específico
        except Exception as e:
            return HttpResponse(f"Error al cargar la plantilla de Excel: {e}", status=500)

        # Obtener los datos de ContratacionEmpleados
        contrataciones = ContratacionEmpleados.objects.all().order_by('-id')
        
        # Definir la celda de inicio (C5)
        start_row = 5
        start_col = 3  # Columna C es la columna 3

        # Escribir los datos en la hoja de Excel
        for idx, contratacion in enumerate(contrataciones, start=start_row):
            try:
                sheet.cell(row=idx, column=start_col, value=contratacion.id)
                sheet.cell(row=idx, column=start_col + 1, value=contratacion.tipo_contratacion)
                sheet.cell(row=idx, column=start_col + 2, value=contratacion.nombre1)
                sheet.cell(row=idx, column=start_col + 3, value=contratacion.nombre2)
                sheet.cell(row=idx, column=start_col + 4, value=contratacion.apellido1)
                sheet.cell(row=idx, column=start_col + 5, value=contratacion.apellido2)
                sheet.cell(row=idx, column=start_col + 6, value=contratacion.fecha_nacimiento)
                
                # Departamento Honduras (asegúrate de que el campo en DepartamentoHonduras se llame 'nombre_departamento')
                sheet.cell(row=idx, column=start_col + 7, value=contratacion.departamento.nombre_departamento if contratacion.departamento else '')
                
                # Municipio Honduras (asegúrate de que el campo en MunicipioHonduras se llame 'nombre_municipio')
                sheet.cell(row=idx, column=start_col + 8, value=contratacion.municipio.nombre_municipio if contratacion.municipio else '')
                
                sheet.cell(row=idx, column=start_col + 9, value=contratacion.genero)
                sheet.cell(row=idx, column=start_col + 10, value=contratacion.dni)
                sheet.cell(row=idx, column=start_col + 11, value=contratacion.telefono)
                sheet.cell(row=idx, column=start_col + 12, value=contratacion.estado_civil)
                sheet.cell(row=idx, column=start_col + 13, value=contratacion.profecion_oficio)
                
                # Datos de contacto de emergencia
                sheet.cell(row=idx, column=start_col + 14, value=contratacion.nombre1_emergencia)
                sheet.cell(row=idx, column=start_col + 15, value=contratacion.parentesco1)
                sheet.cell(row=idx, column=start_col + 16, value=contratacion.telefonoemergencia1)
                sheet.cell(row=idx, column=start_col + 17, value=contratacion.nombre2_emergencia)
                sheet.cell(row=idx, column=start_col + 18, value=contratacion.parentesco2)
                sheet.cell(row=idx, column=start_col + 19, value=contratacion.telefonoemergencia2)
                
                # Dirección exacta, educación y salud
                sheet.cell(row=idx, column=start_col + 20, value=contratacion.direccionexacta)
                sheet.cell(row=idx, column=start_col + 21, value=contratacion.nivel_educativo)
                sheet.cell(row=idx, column=start_col + 22, value=contratacion.ultimo_grado_estudio)
                sheet.cell(row=idx, column=start_col + 23, value=contratacion.ultimogradodetalle)
                
                # Combinar padecimiento y detalle_enfermedad
                padecimiento_detalle = f"{'Si' if contratacion.padecimiento else 'No'} - {contratacion.detalle_enfermedad}"
                sheet.cell(row=idx, column=start_col + 24, value=padecimiento_detalle)
                
                # Puesto
                sheet.cell(row=idx, column=start_col + 25, value=contratacion.puestos.nombre_puestos if contratacion.puestos else '')
                sheet.cell(row=idx, column=start_col + 26, value=contratacion.salario)
                sheet.cell(row=idx, column=start_col + 27, value=contratacion.sucursal.nombre_sucursal if contratacion.sucursal else '')
                sheet.cell(row=idx, column=start_col + 28, value=contratacion.fecha_ingreso)
                
                # Beneficiario
                sheet.cell(row=idx, column=start_col + 29, value=contratacion.nombre_beneficiario)
                sheet.cell(row=idx, column=start_col + 30, value=contratacion.dni_beneficiario)
                sheet.cell(row=idx, column=start_col + 31, value=contratacion.parentesco_beneficiario)
                sheet.cell(row=idx, column=start_col + 32, value=contratacion.porcentaje)
                
                # Departamento empresa (Asegúrate de que el campo en Departamento sea correcto)
                sheet.cell(row=idx, column=start_col + 33, value=contratacion.departamento_empresa.nombre_departamento if contratacion.departamento_empresa else '')
                
                # Unidad de negocio
                sheet.cell(row=idx, column=start_col + 34, value=contratacion.unidad_de_negocio.nombre_unidad_de_negocio if contratacion.unidad_de_negocio else '')
                sheet.cell(row=idx, column=start_col + 35, value=contratacion.correo)
                
                # Tipo de contrato
                sheet.cell(row=idx, column=start_col + 36, value=contratacion.tipo_contrato.nombre_tipo_de_contrato if contratacion.tipo_contrato else '')
                
                # Información de la empresa
                sheet.cell(row=idx, column=start_col + 37, value=contratacion.telefono_empresa)
                sheet.cell(row=idx, column=start_col + 38, value=contratacion.direccionempresa)
                sheet.cell(row=idx, column=start_col + 39, value=contratacion.nombre_empresa.nombre_empresa if contratacion.nombre_empresa else '')
                
                # Hijos (Si o No)
                sheet.cell(row=idx, column=start_col + 40, value='Si' if contratacion.hijos else 'No')
                sheet.cell(row=idx, column=start_col + 41, value=contratacion.comision)
                sheet.cell(row=idx, column=start_col + 42, value=contratacion.bofa)
            except Exception as e:
                return HttpResponse(f"Error escribiendo datos en Excel en la fila {idx + 1}: {e}", status=500)

        # Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="FormatoContratacion.xlsx"'

        # Guardar el libro de trabajo en la respuesta
        wb.save(response)

        return response

#--------CESANTIAs VIEW--------#
@csrf_exempt
def cesantias_view(request, id=None):
    if request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_cesantias'):
            return redirect('sinacceso')
        
        all_estados_pago = Cesantias.objects.values_list('estado_pago', flat=True).distinct()
        all_estados_empleado = Cesantias.objects.values_list('estado_empleado', flat=True).distinct()
        all_años = Cesantias.objects.values_list('año', flat=True).distinct()
        all_autorizadores = Cesantias.objects.values_list('nombre_autoriza', flat=True).distinct()
        all_fechas_extencion = Cesantias.objects.values_list('fecha_extencion', flat=True).distinct()
        all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()
        all_empresas = Empresas.objects.filter(activo=True).order_by('nombre_empresa').distinct()
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
        all_empleados = Colaboradores.objects.order_by('nombrecolaborador').distinct()

        # Obtener parámetros de búsqueda y filtros desde el formulario
        search = request.GET.get('search', '')
        estados_pago = request.GET.getlist('estado_pago', [])  # Lista de estados de pago seleccionados
        estados_empleado = request.GET.getlist('estado_empleado', [])  # Lista de estados de empleado seleccionados
        años = request.GET.getlist('año', [])  # Lista de años seleccionados
        autorizadores = request.GET.getlist('nombre_autoriza', [])  # Lista de nombres de autorizadores seleccionados
        fechas_extencion = request.GET.getlist('fecha_extencion', [])  # Lista de fechas de extensión seleccionadas

        # Query inicial con todas las cesantías
        cesantias = Cesantias.objects.all()

        # Filtrar por búsqueda general (nombre autoriza, nombre empleado, correlativo, año)
        if search:
            try:
                search_date = datetime.strptime(search, '%d/%m/%Y').date()
                cesantias = cesantias.filter(Q(fecha_extencion=search_date))
            except ValueError:
                cesantias = cesantias.filter(
                    Q(nombre_autoriza__icontains=search) |
                    Q(nombre_empleado__icontains=search) |
                    Q(correlativo__icontains=search) |
                    Q(año__icontains=search)
                )

        # Filtrar por estado de pago si hay algún filtro aplicado
        if estados_pago:
            cesantias = cesantias.filter(estado_pago__in=estados_pago)

        # Filtrar por estado de empleado si hay algún filtro aplicado
        if estados_empleado:
            cesantias = cesantias.filter(estado_empleado__in=estados_empleado)

        # Filtrar por año si hay algún filtro aplicado
        if años:
            cesantias = cesantias.filter(año__in=años)

        # Filtrar por nombre del autorizador si hay algún filtro aplicado
        if autorizadores:
            cesantias = cesantias.filter(nombre_autoriza__in=autorizadores)

        # Filtrar por fechas de extensión si hay algún filtro aplicado
        if fechas_extencion:
            try:
                fechas_seleccionadas = [datetime.strptime(f, '%d/%m/%Y').date() for f in fechas_extencion]
                cesantias = cesantias.filter(fecha_extencion__in=fechas_seleccionadas)
            except ValueError:
                pass  # Si alguna fecha es inválida, ignorar el filtro

        # Paginación: mostrar 10 registros por página
        paginator = Paginator(cesantias, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calcular el rango de páginas a mostrar en la paginación
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        # Contexto para renderizar la plantilla
        context = {
            'cesantias': page_obj,
            'all_sucursales':all_sucursales,
            'all_empresas':all_empresas,
            'all_departamentos':all_departamentos,
            'search': search,
            'page_range': page_range,
            'all_estados_pago': all_estados_pago,
            'all_estados_empleado': all_estados_empleado,
            'all_años': all_años,
            'all_autorizadores': all_autorizadores,
            'all_fechas_extencion': all_fechas_extencion,  # Fechas de extensión disponibles
            'estados_pago': estados_pago,
            'estados_empleado': estados_empleado,
            'años': años,  # Años seleccionados
            'autorizadores': autorizadores,
            'fechas_extencion': fechas_extencion,
            'all_empleados': all_empleados,
        }

        return render(request, 'cesantias/cesantias.html', context)
    elif request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_cesantias'):
                return JsonResponse({'error': 'No tienes permisos para agregar cesantías'}, status=403)
            
            data = json.loads(request.body)
            
            cesantia = Cesantias(
                correlativo=data['correlativo'],
                nombre_autoriza=data['nombre_autoriza'],
                dni_autoriza=data['dni_autoriza'],
                cargo_autoriza=data['cargo_autoriza'],
                empresa=Empresas.objects.get(id=data['empresa']),
                sucursal=Sucursal.objects.get(id=data['sucursal']),
                departamento=Departamento.objects.get(id=data['departamento']),
                nombre_empleado=data['nombre_empleado'],
                dni_empleado=data['dni_empleado'],
                fecha_inicial=data['fecha_inicial'],
                fecha_final=data['fecha_final'],
                fecha_extencion=data['fecha_extencion'],
                sueldo_actual=data['sueldo_actual'],
                cesantia_actual=data['cesantia_actual'],
                cesantia_final=data['cesantia_final'],
                porcentaje=data['porcentaje'],
                año=data['año'],
                estado_empleado=data['estado_empleado'],
                estado_pago=data['estado_pago'],
                n_cheke=data.get('n_cheke', None),
                usuario_creo = request.user,
            )

            # Guardar la nueva cesantía
            cesantia.save()

            # Responder con éxito
            return JsonResponse({'success': True, 'message': 'Cesantía registrada correctamente.'})

        except Exception as e:
            # En caso de error, devolver una respuesta con el error
            return JsonResponse({'success': False, 'message': str(e)})
        
    elif request.method == 'PUT':
        try:
            if not request.user.has_perm('Reclutamiento.change_cesantias'):
                return JsonResponse({'error': 'No tienes permisos para editar cesantías'}, status=403)
            data = json.loads(request.body)

            # Buscar el registro a actualizar usando el id
            cesantia = get_object_or_404(Cesantias, id=id)

            # Actualizar los campos
            cesantia.nombre_autoriza = data['nombre_autoriza']
            cesantia.dni_autoriza = data['dni_autoriza']
            cesantia.cargo_autoriza = data['cargo_autoriza']
            cesantia.sucursal_id = data['sucursal']
            cesantia.empresa_id = data['empresa']
            cesantia.departamento_id = data['departamento']
            cesantia.nombre_empleado = data['nombre_empleado']
            cesantia.dni_empleado = data['dni_empleado']
            cesantia.fecha_inicial = data['fecha_inicial']
            cesantia.fecha_final = data['fecha_final']
            cesantia.fecha_extencion = data['fecha_extencion']
            cesantia.sueldo_actual = data['sueldo_actual']
            cesantia.cesantia_actual = data['cesantia_actual']
            cesantia.cesantia_final = data['cesantia_final']
            cesantia.porcentaje = data['porcentaje']
            cesantia.año = data['año']
            cesantia.estado_empleado = data['estado_empleado']
            cesantia.estado_pago = data['estado_pago']
            cesantia.n_cheke = data['n_cheke']
            cesantia.usuario_modifico = request.user

            # Guardar los cambios
            cesantia.save()

            return JsonResponse({'success': True, 'message': 'Cesantía actualizada correctamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Método no permitido.'})

def numero_a_letras_con_decimales(numero):
    parte_entera = int(numero)
    parte_decimal = int(round((numero - parte_entera) * 100))

    parte_entera_letras = num2words(parte_entera, lang='es').upper()

    if parte_decimal > 0:
        parte_decimal_letras = num2words(parte_decimal, lang='es').upper()
        return f"{parte_entera_letras} PUNTO {parte_decimal_letras}"
    else:
        return parte_entera_letras

def imprimircesantias_view(request, id):
    cesantia = get_object_or_404(Cesantias, id=id)

    # Convertir a palabras usando la nueva función para manejar decimales correctamente
    cesantia_letras = numero_a_letras_con_decimales(cesantia.cesantia_actual)
    cesantia_letras_final = numero_a_letras_con_decimales(cesantia.cesantia_final)
    
    # Formatear números con separador de miles y dos decimales
    cesantia_actual_formateado = "{:,.2f}".format(cesantia.cesantia_actual)
    cesantia_final_formateado = "{:,.2f}".format(cesantia.cesantia_final)
    sueldo_actual_formateado = "{:,.2f}".format(cesantia.sueldo_actual)

    context = {
        'cesantia': cesantia,
        'cesantia_letras': cesantia_letras,
        'cesantia_letras_final': cesantia_letras_final,
        'cesantia_actual_formateado': cesantia_actual_formateado,
        'cesantia_final_formateado': cesantia_final_formateado,
        'sueldo_actual_formateado': sueldo_actual_formateado
    }
    
    return render(request, 'cesantias/imprimircesantias.html', context)

def exportar_cesantias(request):
    datos = (
        Cesantias.objects
        .select_related('empresa', 'sucursal', 'departamento') 
        .values(
            'fecha_extencion', 
            'empresa__nombre_empresa', 
            'sucursal__nombre_sucursal',  
            'nombre_empleado', 
            'dni_empleado',
            'departamento__nombre_departamento',
            'año', 
            'cesantia_final'
        )
    )

    # Convertir los datos a un DataFrame de pandas
    df = pd.DataFrame(list(datos))
    
    # Crear un archivo Excel
    excel_file = "cesantias.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={excel_file}'

    # Escribir el DataFrame en el archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Cesantias')

    return response


#--------PERFILES DE PUESTOS VIEW--------#
def perfilpuesto_view(request):
    if not request.user.has_perm('Reclutamiento.view_perfilespuestos'):
       return redirect('sinacceso')
    
    all_tipos_perfil = PerfilesPuestos.objects.values_list('tipo_perfil', flat=True).distinct()
    all_empresas = PerfilesPuestos.objects.values_list('empresa', flat=True).distinct()
    all_departamentos = PerfilesPuestos.objects.values_list('departamento', flat=True).distinct()
    all_nombres_cargo = PerfilesPuestos.objects.values_list('nombre_cargo', flat=True).distinct()
    all_fechas_actualizacion = (
        PerfilesPuestos.objects.annotate(fecha_actualizacion_dia=TruncDate('modificado'))
        .values_list('fecha_actualizacion_dia', flat=True)
        .distinct()
    )

    # Obtener parámetros de búsqueda y filtros desde el formulario
    search = request.GET.get('search', '')  
    tipos_perfil = request.GET.getlist('tipo_perfil', [])  
    empresas = request.GET.getlist('empresa', [])  
    departamentos = request.GET.getlist('departamento', [])  
    nombres_cargo = request.GET.getlist('nombre_cargo', [])  
    fechas_actualizacion = request.GET.getlist('fecha_actualizacion', [])  

    # Query inicial con todos los perfiles de puesto
    perfiles_puesto = PerfilesPuestos.objects.all()

    # Filtrar por búsqueda general
    if search:
        perfiles_puesto = perfiles_puesto.filter(
            Q(nombre_cargo__icontains=search) |
            Q(tipo_perfil__icontains=search) |
            Q(empresa__icontains=search) |
            Q(departamento__icontains=search)
        )

    # Filtrar por tipo de perfil si hay algún filtro aplicado
    if tipos_perfil:
        perfiles_puesto = perfiles_puesto.filter(tipo_perfil__in=tipos_perfil)

    # Filtrar por empresa si hay algún filtro aplicado
    if empresas:
        perfiles_puesto = perfiles_puesto.filter(empresa__in=empresas)

    # Filtrar por departamento si hay algún filtro aplicado
    if departamentos:
        perfiles_puesto = perfiles_puesto.filter(departamento__in=departamentos)

    # Filtrar por nombre de cargo si hay algún filtro aplicado
    if nombres_cargo:
        perfiles_puesto = perfiles_puesto.filter(nombre_cargo__in=nombres_cargo)

    # Filtrar por fecha de actualización si hay algún filtro aplicado
    if fechas_actualizacion:
        try:
            fechas_seleccionadas = [datetime.strptime(f, '%d/%m/%Y') for f in fechas_actualizacion]
            perfiles_puesto = perfiles_puesto.filter(modificado__date__in=fechas_seleccionadas)
        except ValueError:
            pass  # Si alguna fecha es inválida, ignorar el filtro

    # Paginación: mostrar 10 registros por página
    paginator = Paginator(perfiles_puesto, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calcular el rango de páginas a mostrar en la paginación
    total_pages = paginator.num_pages
    current_page = page_obj.number
    start_page = max(1, current_page - 2)
    end_page = min(total_pages, current_page + 2)

    if current_page <= 3:
        end_page = min(5, total_pages)
    elif current_page >= total_pages - 2:
        start_page = max(1, total_pages - 4)

    page_range = range(start_page, end_page + 1)

    # Contexto para renderizar la plantilla
    context = {
        'perfiles_puesto': page_obj,
        'search': search,
        'page_range': page_range,
        'all_tipos_perfil': all_tipos_perfil,
        'all_empresas': all_empresas,
        'all_departamentos': all_departamentos,
        'all_nombres_cargo': all_nombres_cargo,
        'all_fechas_actualizacion': all_fechas_actualizacion,
        'tipos_perfil': tipos_perfil,
        'empresas': empresas,
        'departamentos': departamentos,
        'nombres_cargo': nombres_cargo,
        'fechas_actualizacion': fechas_actualizacion,
    }

    return render(request, 'perfilpuesto/perfilpuesto.html', context)

def obtener_pruebas_por_nivel(request):
    if request.method == 'GET':
        nivel = request.GET.get('nivel') 
        if nivel:
            pruebas = Spicosmart.objects.filter(nivel=nivel, activo=True)
            pruebas_data = list(pruebas.values('nombre_prueba'))
            return JsonResponse({'success': True, 'pruebas': pruebas_data}, status=200)
        else:
            return JsonResponse({'success': False, 'message': 'Nivel no proporcionado'}, status=400)
        
def perfilpuestoregister_view(request):
    if request.method == 'GET':
        all_empresas = Empresas.objects.filter(activo=True).order_by('nombre_empresa').distinct()
        all_puestos = Puestos.objects.filter(activo=True).order_by('nombre_puestos').distinct()
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()

        context = {
            'all_empresas': all_empresas,
            'all_puestos': all_puestos,
            'all_departamentos': all_departamentos
        }
        return render(request, 'perfilpuesto/perfilpuestoregistro.html', context)

    elif request.method == 'POST':
        if not request.user.has_perm('Reclutamiento.add_perfilespuestos'):
            return JsonResponse({'error': 'No tienes permisos para agregar perfiles de puesto'}, status=403)
        try:
            tipo_perfil = request.POST.get('tipo_perfil') or None
            empresa = request.POST.get('empresa') or None
            nombre_cargo = request.POST.get('nombre_cargo') or None
            departamento = request.POST.get('departamento') or None
            cargo_al_que_reporta = request.POST.get('cargo_al_que_reporta') or None
            cargo_que_le_reportan = request.POST.get('cargos_que_le_reportan') or None
            educacion_universitario = request.POST.get('educacion_universitario') or None
            postgrado_especializaciones = request.POST.get('postgrado_especializaciones') or None
            formacion_complementaria = request.POST.get('formacion_complementaria') or None
            idiomas = request.POST.get('idiomas') or None
            nivel_idioma = request.POST.get('nivel_idioma') or None
            anos_experiencia = request.POST.get('anos_experiencia') or None
            pensamiento_estrategico = request.POST.get('pensamiento_estrategico') or None
            enfoque_al_cliente = request.POST.get('enfoque_al_cliente') or None
            planificacion_y_organizacion = request.POST.get('planificacion_y_organizacion') or None
            comunicacion = request.POST.get('comunicacion') or None
            orientacion_al_logro = request.POST.get('orientacion_al_logro') or None
            mision_cargo = request.POST.get('mision_cargo') or None

            # Captura del campo de 'residir en el área', asignando null si está vacío
            residir_en_area = request.POST.get('residir_en_area') or None

            # Guardar imagen del organigrama si existe
            imagen = request.FILES.get('ruta_organigrama')
            nombre_imagen = None
            ruta_imagen = None
            if imagen:
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/organigrama/{nombre_imagen}"

                os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                # Guardar el archivo en la ruta correspondiente
                with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                    for chunk in imagen.chunks():
                        destino.write(chunk)

            # Capturar los campos dinámicos JSON (funciones, retos, materiales, etc.), con valores por defecto como lista vacía o null
            funciones_cargo = json.loads(request.POST.get('funciones_cargo', '[]') or '[]')
            retos = json.loads(request.POST.get('retos', '[]') or '[]')  
            materiales_equipos = json.loads(request.POST.get('materiales_equipos', '[]') or '[]')
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]') or '[]')
            beneficiospromaco = json.loads(request.POST.get('beneficiospromaco', '[]') or '[]')
            psicometricas = json.loads(request.POST.get('psicometricas', '[]') or '[]')
            principales_indicadores = json.loads(request.POST.get('principales_indicadores', '[]') or '[]')

            # Captura de otros campos, asignando null si están vacíos
            horario_turnos = request.POST.get('horario_turnos') or None
            otros = request.POST.get('otros') or None
            nivel_prueba = request.POST.get('nivel_prueba') or None
            montacargas = request.POST.get('montacargas') or None
            esquipo_pegado = request.POST.get('esquipo_pegado') or None

            # Crear nuevo perfil de puesto con todos los datos capturados, enviando null cuando corresponda
            perfil = PerfilesPuestos.objects.create(
                tipo_perfil=tipo_perfil,
                empresa=empresa,
                nombre_cargo=nombre_cargo,
                departamento=departamento,
                cargo_al_que_reporta=cargo_al_que_reporta,
                cargo_que_le_reportan=cargo_que_le_reportan,
                educacion_universitario=educacion_universitario,
                postgrado_especializaciones=postgrado_especializaciones,
                formacion_complementaria=formacion_complementaria,
                idiomas=idiomas,
                nivel_idioma=nivel_idioma,
                anos_experiencia=anos_experiencia,
                pensamiento_estrategico=pensamiento_estrategico,
                enfoque_al_cliente=enfoque_al_cliente,
                planificacion_y_organizacion=planificacion_y_organizacion,
                comunicacion=comunicacion,
                orientacion_al_logro=orientacion_al_logro,
                residir_en_area=residir_en_area,
                mision_cargo=mision_cargo,
                ruta_organigrama=ruta_imagen,
                nombre_organigrama=nombre_imagen,
                funciones_cargo=funciones_cargo,
                retos=retos,
                materiales_equipos=materiales_equipos,
                horario_turnos=horario_turnos,
                otros=otros,
                plan_de_compensacion=plan_de_compensacion,
                beneficiospromaco=beneficiospromaco,
                nivel_prueba=nivel_prueba,
                psicometricas=psicometricas,
                montacargas=montacargas,
                esquipo_pegado=esquipo_pegado,
                principales_indicadores=principales_indicadores,
                usuario_creo=request.user,
            )

            # Devolver una respuesta de éxito
            return JsonResponse({'success': True, 'message': 'Perfil de puesto registrado correctamente.'})

        except Exception as e:
            # Depuración de errores y POST data
            print(f"Datos POST: {request.POST}")
            print(f"Error: {str(e)}")  # Para depurar el error exacto
            return JsonResponse({'success': False, 'message': f'Error al procesar la solicitud: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

def imprimir_perfilpuesto_view(request, perfil_id):
    perfil_puesto = get_object_or_404(PerfilesPuestos, id=perfil_id)
    
    context = {
        'perfil_puesto': perfil_puesto,
    }

    # Renderiza la plantilla para la impresión
    return render(request, 'perfilpuesto/imprimirperfilpuesto.html', context)

def imprimir_perfilpuestocorto_view(request, perfil_id):
    perfil_puesto = get_object_or_404(PerfilesPuestos, id=perfil_id)
    
    context = {
        'perfil_puesto': perfil_puesto,
    }

    # Renderiza la plantilla para la impresión
    return render(request, 'perfilpuesto/imprimirperfilpuestocorto.html', context)

def update_perfilpuesto_view(request, id):
    if request.method == 'GET':
        perfil_puesto = get_object_or_404(PerfilesPuestos, id=id)
        all_empresas = Empresas.objects.filter(activo=True).order_by('nombre_empresa').distinct()
        all_puestos = Puestos.objects.filter(activo=True).order_by('nombre_puestos').distinct()
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()

        # Valores predefinidos para materiales
        predefined_materials = ["COMPUTADOR", "CELULAR", "UNIFORME", "ESCRITORIO", "LICENCIA SAP", "LICENCIA IVEND", "SILLA", "CORREO"]
        otros_materiales = [material for material in perfil_puesto.materiales_equipos if material not in predefined_materials]
        otros_materiales = otros_materiales[0] if otros_materiales else ''

        # Valores predefinidos para plan de compensación y beneficios
        predefined_compensacion = ["SALARIO BASE", "COMISION", "BONO", "DEPRECIACION Y COMBUSTIBLE", "HOSPEDAJE"]
        predefined_beneficios = ["SEGURO MEDICO PRIVADO", "SEGURO SOCIAL", "PERMISO DE ESTUDIO", "BONO SUPERMERCADO", 
                                "BONO CAFE PROMACO", "DESCUENTO EN SUPERMERCADO", "DESCUENTO EN CLINICA Y LAVORATORIOS", 
                                "SEGURO DE VEHICULO"]

        # Limpiar espacios extra en plan de compensación y filtrar "Otros"
        plan_de_compensacion_limpio = [comp.strip() for comp in perfil_puesto.plan_de_compensacion]
        otros_compensacion = [comp for comp in plan_de_compensacion_limpio if comp not in predefined_compensacion]
        otros_compensacion = otros_compensacion[0] if otros_compensacion else ''

        context = {
            'perfil_puesto': perfil_puesto,
            'all_empresas': all_empresas,
            'all_puestos': all_puestos,
            'all_departamentos': all_departamentos,
            'otros_materiales': otros_materiales,
            'otros_compensacion': otros_compensacion,  # Para el campo "Otros" en plan de compensación
            'predefined_materials': predefined_materials,  # Para comparar en el template
            'predefined_compensacion': predefined_compensacion,  # Para comparar en el template
            'predefined_beneficios': predefined_beneficios  # Para comparar en el template
        }

        return render(request, 'perfilpuesto/updateperfilpuesto.html', context)
    
    elif request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.change_perfilespuestos'):
                return JsonResponse({'error': 'No tienes permisos para editar perfiles de puesto'}, status=403)
            # Obtener el perfil de puesto existente
            perfil = get_object_or_404(PerfilesPuestos, id=id)
            
            # Obtener datos del formulario, asignando None si el valor es vacío
            tipo_perfil = request.POST.get('tipo_perfil') or None
            empresa = request.POST.get('empresa') or None
            nombre_cargo = request.POST.get('nombre_cargo') or None
            departamento = request.POST.get('departamento') or None
            cargo_al_que_reporta = request.POST.get('cargo_al_que_reporta') or None
            cargo_que_le_reportan = request.POST.get('cargos_que_le_reportan') or None
            educacion_universitario = request.POST.get('educacion_universitario') or None
            postgrado_especializaciones = request.POST.get('postgrado_especializaciones') or None
            formacion_complementaria = request.POST.get('formacion_complementaria') or None
            idiomas = request.POST.get('idiomas') or None
            nivel_idioma = request.POST.get('nivel_idioma') or None
            anos_experiencia = request.POST.get('anos_experiencia') or None
            pensamiento_estrategico = request.POST.get('pensamiento_estrategico') or None
            enfoque_al_cliente = request.POST.get('enfoque_al_cliente') or None
            planificacion_y_organizacion = request.POST.get('planificacion_y_organizacion') or None
            comunicacion = request.POST.get('comunicacion') or None
            orientacion_al_logro = request.POST.get('orientacion_al_logro') or None
            mision_cargo = request.POST.get('mision_cargo') or None

            # Captura del campo de 'residir en el área', asignando None si está vacío
            residir_en_area = request.POST.get('residir_en_area') or None

            # Variables para imagen del organigrama
            imagen = request.FILES.get('ruta_organigrama')
            nombre_imagen = perfil.nombre_organigrama  # Mantener el nombre existente si no hay imagen nueva
            ruta_imagen = perfil.ruta_organigrama  # Mantener la ruta existente si no hay imagen nueva

            if imagen:
                # Si ya existe una imagen, eliminar la imagen anterior
                if perfil.ruta_organigrama:
                    ruta_anterior = os.path.join(settings.BASE_DIR, perfil.ruta_organigrama)
                    if os.path.exists(ruta_anterior):
                        try:
                            os.remove(ruta_anterior)  # Elimina la imagen anterior
                        except Exception as e:
                            return JsonResponse({'success': False, 'message': f'Error al eliminar la imagen anterior: {str(e)}'}, status=500)

                # Generar un nombre único para la nueva imagen
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/organigrama/{nombre_imagen}"

                # Crear el directorio si no existe
                os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                # Guardar el archivo en la ruta correspondiente
                with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                    for chunk in imagen.chunks():
                        destino.write(chunk)

            # Capturar los campos dinámicos JSON (funciones, retos, materiales, etc.), con valores por defecto como lista vacía o None
            funciones_cargo = json.loads(request.POST.get('funciones_cargo', '[]') or '[]')
            retos = json.loads(request.POST.get('retos', '[]') or '[]')  
            materiales_equipos = json.loads(request.POST.get('materiales_equipos', '[]') or '[]')
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]') or '[]')
            beneficiospromaco = json.loads(request.POST.get('beneficiospromaco', '[]') or '[]')
            psicometricas = json.loads(request.POST.get('psicometricas', '[]') or '[]')
            principales_indicadores = json.loads(request.POST.get('principales_indicadores', '[]') or '[]')

            # Captura de otros campos, asignando None si están vacíos
            horario_turnos = request.POST.get('horario_turnos') or None
            otros = request.POST.get('otros') or None
            nivel_prueba = request.POST.get('nivel_prueba') or None
            montacargas = request.POST.get('montacargas') or None
            esquipo_pegado = request.POST.get('esquipo_pegado') or None

            # Asignar los valores al perfil existente
            perfil.tipo_perfil = tipo_perfil
            perfil.empresa = empresa
            perfil.nombre_cargo = nombre_cargo
            perfil.departamento = departamento
            perfil.cargo_al_que_reporta = cargo_al_que_reporta
            perfil.cargo_que_le_reportan = cargo_que_le_reportan
            perfil.educacion_universitario = educacion_universitario
            perfil.postgrado_especializaciones = postgrado_especializaciones
            perfil.formacion_complementaria = formacion_complementaria
            perfil.idiomas = idiomas
            perfil.nivel_idioma = nivel_idioma
            perfil.anos_experiencia = anos_experiencia
            perfil.pensamiento_estrategico = pensamiento_estrategico
            perfil.enfoque_al_cliente = enfoque_al_cliente
            perfil.planificacion_y_organizacion = planificacion_y_organizacion
            perfil.comunicacion = comunicacion
            perfil.orientacion_al_logro = orientacion_al_logro
            perfil.residir_en_area = residir_en_area
            perfil.mision_cargo = mision_cargo
            perfil.ruta_organigrama = ruta_imagen
            perfil.nombre_organigrama = nombre_imagen
            perfil.funciones_cargo = funciones_cargo
            perfil.retos = retos
            perfil.materiales_equipos = materiales_equipos
            perfil.horario_turnos = horario_turnos
            perfil.otros = otros
            perfil.plan_de_compensacion = plan_de_compensacion
            perfil.beneficiospromaco = beneficiospromaco
            perfil.nivel_prueba = nivel_prueba
            perfil.psicometricas = psicometricas
            perfil.montacargas = montacargas
            perfil.esquipo_pegado = esquipo_pegado
            perfil.principales_indicadores = principales_indicadores
            perfil.usuario_modifico = request.user

            # Guardar el perfil actualizado
            perfil.save()

            # Devolver una respuesta de éxito
            return JsonResponse({'success': True, 'message': 'Perfil de puesto actualizado correctamente.'})

        except Exception as e:
            # Depuración de errores y POST data
            print(f"Datos POST: {request.POST}")
            print(f"Error: {str(e)}")  # Para depurar el error exacto
            return JsonResponse({'success': False, 'message': f'Error al procesar la solicitud: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

def updatecompleto_perfilpuesto_view(request, id):
    if request.method == 'GET':
        perfil_puesto = get_object_or_404(PerfilesPuestos, id=id)
        all_empresas = Empresas.objects.filter(activo=True).order_by('nombre_empresa').distinct()
        all_puestos = Puestos.objects.filter(activo=True).order_by('nombre_puestos').distinct()
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()

        # Valores predefinidos para materiales
        predefined_materials = ["COMPUTADOR", "CELULAR", "UNIFORME", "ESCRITORIO", "LICENCIA SAP", "LICENCIA IVEND", "SILLA", "CORREO"]
        otros_materiales = [material for material in perfil_puesto.materiales_equipos if material not in predefined_materials]
        otros_materiales = otros_materiales[0] if otros_materiales else ''

        # Valores predefinidos para plan de compensación y beneficios
        predefined_compensacion = ["SALARIO BASE", "COMISION", "BONO", "DEPRECIACION Y COMBUSTIBLE", "HOSPEDAJE"]
        predefined_beneficios = ["SEGURO MEDICO PRIVADO", "SEGURO SOCIAL", "PERMISO DE ESTUDIO", "BONO SUPERMERCADO", 
                                "BONO CAFE PROMACO", "DESCUENTO EN SUPERMERCADO", "DESCUENTO EN CLINICA Y LAVORATORIOS", 
                                "SEGURO DE VEHICULO"]

        # Limpiar espacios extra en plan de compensación y filtrar "Otros"
        plan_de_compensacion_limpio = [comp.strip() for comp in perfil_puesto.plan_de_compensacion]
        otros_compensacion = [comp for comp in plan_de_compensacion_limpio if comp not in predefined_compensacion]
        otros_compensacion = otros_compensacion[0] if otros_compensacion else ''

        context = {
            'perfil_puesto': perfil_puesto,
            'all_empresas': all_empresas,
            'all_puestos': all_puestos,
            'all_departamentos': all_departamentos,
            'otros_materiales': otros_materiales,
            'otros_compensacion': otros_compensacion,  # Para el campo "Otros" en plan de compensación
            'predefined_materials': predefined_materials,  # Para comparar en el template
            'predefined_compensacion': predefined_compensacion,  # Para comparar en el template
            'predefined_beneficios': predefined_beneficios  # Para comparar en el template
        }

        return render(request, 'perfilpuesto/updatetotalperfilpuesto.html', context)


#--------REQUISAS DE PUESTOS VIEW--------#
def requisaperfil_view(request, id):
    if request.method == 'GET':
        perfil_puesto = get_object_or_404(PerfilesPuestos, id=id)
        all_empresas = Empresas.objects.filter(activo=True).order_by('nombre_empresa').distinct()
        all_puestos = Puestos.objects.filter(activo=True).order_by('nombre_puestos').distinct()
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
        all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(activo=True).order_by('nombre_tipo_de_contrato').distinct()

        # Extraer solo las funciones del JSONField
        funciones_list = [funcion['funcion'] for funcion in perfil_puesto.funciones_cargo]
        funciones_text = ', '.join(funciones_list)  # Unir las funciones con comas

        context = {
            'perfil_puesto': perfil_puesto,
            'all_empresas': all_empresas,
            'all_puestos': all_puestos,
            'all_departamentos': all_departamentos,
            'all_sucursales': all_sucursales,
            'funciones_text': funciones_text,
            'all_tipos_contrato':all_tipos_contrato
        }

        return render(request, 'requisas/requisaperfil.html', context)
    elif request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_requisa'):
                return JsonResponse({'error': 'No tienes permisos para agregar requisas'}, status=403)
            salario_base = request.POST.get('salario_base') or None
            
            # Decodificar el JSON enviado con los valores de plan_de_compensacion y habilidades
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]'))
            montos = json.loads(request.POST.get('montos', '{}'))

            puesto = request.POST.get('puesto') or None
            sucursal_id = request.POST.get('sucursal') or None
            departamento_id = request.POST.get('departamento') or None
            funciones_cargo = request.POST.get('funciones_cargo') or None

            hora_inicio = request.POST.get('hora_inicio') or None
            hora_fin = request.POST.get('hora_fin') or None

            centrocostos = request.POST.get('centrocostos') or None
            puestonuevo = bool(int(request.POST.get('puestonuevo', 0)))  
            incapacidad = bool(int(request.POST.get('incapacidad', 0)))  
            reemplazo = bool(int(request.POST.get('reemplazo', 0)))

            tiempoprimercontrato = request.POST.get('tiempoprimercontrato') or None
            tipo_contrato_id = request.POST.get('tipo_contrato') or None
            motivo = request.POST.get('motivo') or None
            nombrereemplazar = request.POST.get('nombrereemplazar') or None
            formacionacademica = request.POST.get('formacionacademica') or None

            # Decodificar los JSON de habilidades
            habilidadesferreteras = json.loads(request.POST.get('habilidadesferreteras', '[]'))
            habilidadesinformaticas = json.loads(request.POST.get('habilidadesinformaticas', '[]'))
            habilidadespersonales = json.loads(request.POST.get('habilidadespersonales', '[]'))
            habilidadesanaliticas = json.loads(request.POST.get('habilidadesanaliticas', '[]'))
            materialesequipo = json.loads(request.POST.get('materialesequipo', '[]'))

            # Obtener instancias relacionadas
            sucursal = Sucursal.objects.get(id=sucursal_id) if sucursal_id else None
            departamento = Departamento.objects.get(id=departamento_id) if departamento_id else None
            tipo_contrato = TipoContrato.objects.get(id=tipo_contrato_id) if tipo_contrato_id else None

            # Crear la nueva Requisa sin fechaRecepcion
            nueva_requisa = Requisa.objects.create(
                salario_base=salario_base,
                plan_de_compensacion=plan_de_compensacion,
                montos=montos,
                puesto=puesto,
                sucursal=sucursal,
                departamento=departamento,
                hora_inicio=hora_inicio,
                hora_fin=hora_fin,
                centrocostos=centrocostos,
                puestonuevo=puestonuevo,
                incapacidad=incapacidad,
                reemplazo=reemplazo,
                tiempoprimercontrato=tiempoprimercontrato,
                tipo_contrato=tipo_contrato,
                motivo=motivo,
                nombrereemplazar=nombrereemplazar,
                formacionacademica=formacionacademica,
                habilidadesferreteras=habilidadesferreteras,
                habilidadesinformaticas=habilidadesinformaticas,
                habilidadespersonales=habilidadespersonales,
                habilidadesanaliticas=habilidadesanaliticas,
                materialesequipo=materialesequipo,
                funciones_cargo=funciones_cargo,
                fechaRecepcion=None,
                usuario_creo=request.user,
                estado="EN PROCESO"
            )

            return JsonResponse({'success': True, 'message': 'Requisa registrada correctamente'})

        except Exception as e:
            # Si ocurre algún error, devolvemos una respuesta de error
            return JsonResponse({'success': False, 'message': str(e)})

def requisa_view(request):
    if not request.user.has_perm('Reclutamiento.view_requisa'):
        return redirect('sinacceso')
    all_sucursales = Requisa.objects.values_list('sucursal__nombre_sucursal', flat=True).distinct()
    all_departamentos = Requisa.objects.values_list('departamento__nombre_departamento', flat=True).distinct()
    all_estados = Requisa.objects.values_list('estado', flat=True).distinct()
    all_fechas_creacion = Requisa.objects.annotate(fecha_creacion_dia=TruncDate('creado')).values_list('fecha_creacion_dia', flat=True).distinct()
    all_fechas_recepcion = Requisa.objects.annotate(fecha_recepcion_dia=TruncDate('fechaRecepcion')).values_list('fecha_recepcion_dia', flat=True).distinct()
    all_puestos = Requisa.objects.values_list('puesto', flat=True).distinct()

    search = request.GET.get('search', '')
    sucursales = request.GET.getlist('sucursal', [])
    departamentos = request.GET.getlist('departamento', [])
    estados = request.GET.getlist('estado', [])
    fechas_creacion = request.GET.getlist('creado', [])
    fechas_recepcion = request.GET.getlist('fecharecepcion', [])
    puestos = request.GET.getlist('puesto', [])

    requisas = Requisa.objects.all()

    if search:
        requisas = requisas.filter(
            Q(puesto__icontains=search) |
            Q(sucursal__nombre_sucursal__icontains=search) |
            Q(departamento__nombre_departamento__icontains=search) |
            Q(estado__icontains=search)
        )

    if sucursales:
        requisas = requisas.filter(sucursal__nombre_sucursal__in=sucursales)

    if departamentos:
        requisas = requisas.filter(departamento__nombre_departamento__in=departamentos)

    if estados:
        requisas = requisas.filter(estado__in=estados)

    if puestos:
        requisas = requisas.filter(puesto__in=puestos)

    if fechas_creacion:
        try:
            fechas_seleccionadas = [datetime.strptime(f, '%d/%m/%Y').date() for f in fechas_creacion]
            requisas = requisas.filter(creado__date__in=fechas_seleccionadas)
        except ValueError:
            pass

    if fechas_recepcion:
        try:
            fechas_seleccionadas = [datetime.strptime(f, '%d/%m/%Y').date() for f in fechas_recepcion]
            requisas = requisas.filter(fechaRecepcion__date__in=fechas_seleccionadas)
        except ValueError:
            pass

    paginator = Paginator(requisas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages
    current_page = page_obj.number
    start_page = max(1, current_page - 2)
    end_page = min(total_pages, current_page + 2)

    if current_page <= 3:
        end_page = min(5, total_pages)
    elif current_page >= total_pages - 2:
        start_page = max(1, total_pages - 4)

    page_range = range(start_page, end_page + 1)

    context = {
        'requisas': page_obj,
        'search': search,
        'page_range': page_range,
        'all_sucursales': all_sucursales,
        'all_departamentos': all_departamentos,
        'all_estados': all_estados,
        'all_fechas_creacion': all_fechas_creacion,
        'all_fechas_recepcion': all_fechas_recepcion,
        'all_puestos': all_puestos,
        'sucursales': sucursales,
        'departamentos': departamentos,
        'estados': estados,
        'fechas_creacion': fechas_creacion,
        'fechas_recepcion': fechas_recepcion,
        'puestos': puestos,
    }

    return render(request, 'requisas/requisa.html', context)

def requisaregistrar_view(request):
    if request.method == 'GET':
        # Cargar el formulario
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
        all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(activo=True).order_by('nombre_tipo_de_contrato').distinct()

        context = {
            'all_departamentos': all_departamentos,
            'all_sucursales': all_sucursales,
            'all_tipos_contrato': all_tipos_contrato
        }

        return render(request, 'requisas/requisaregistrar.html', context)

    elif request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_requisa'):
                return JsonResponse({'error': 'No tienes permisos para agregar perfiles de puesto'}, status=403)
            # Recibir los datos enviados desde el formulario con FormData
            salario_base = request.POST.get('salario_base') or None
            
            # Decodificar el JSON enviado con los valores de plan_de_compensacion y habilidades
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]'))
            montos = json.loads(request.POST.get('montos', '{}'))

            puesto = request.POST.get('puesto') or None
            sucursal_id = request.POST.get('sucursal') or None
            departamento_id = request.POST.get('departamento') or None
            funciones_cargo = request.POST.get('funciones_cargo') or None

            hora_inicio = request.POST.get('hora_inicio') or None
            hora_fin = request.POST.get('hora_fin') or None

            centrocostos = request.POST.get('centrocostos') or None
            puestonuevo = bool(int(request.POST.get('puestonuevo', 0)))  
            incapacidad = bool(int(request.POST.get('incapacidad', 0)))  
            reemplazo = bool(int(request.POST.get('reemplazo', 0)))

            tiempoprimercontrato = request.POST.get('tiempoprimercontrato') or None
            tipo_contrato_id = request.POST.get('tipo_contrato') or None
            motivo = request.POST.get('motivo') or None
            nombrereemplazar = request.POST.get('nombrereemplazar') or None
            formacionacademica = request.POST.get('formacionacademica') or None

            # Decodificar los JSON de habilidades
            habilidadesferreteras = json.loads(request.POST.get('habilidadesferreteras', '[]'))
            habilidadesinformaticas = json.loads(request.POST.get('habilidadesinformaticas', '[]'))
            habilidadespersonales = json.loads(request.POST.get('habilidadespersonales', '[]'))
            habilidadesanaliticas = json.loads(request.POST.get('habilidadesanaliticas', '[]'))
            materialesequipo = json.loads(request.POST.get('materialesequipo', '[]'))

            # Obtener instancias relacionadas
            sucursal = Sucursal.objects.get(id=sucursal_id) if sucursal_id else None
            departamento = Departamento.objects.get(id=departamento_id) if departamento_id else None
            tipo_contrato = TipoContrato.objects.get(id=tipo_contrato_id) if tipo_contrato_id else None

            # Crear la nueva Requisa sin fechaRecepcion
            nueva_requisa = Requisa.objects.create(
                salario_base=salario_base,
                plan_de_compensacion=plan_de_compensacion,
                montos=montos,
                puesto=puesto,
                sucursal=sucursal,
                departamento=departamento,
                hora_inicio=hora_inicio,
                hora_fin=hora_fin,
                centrocostos=centrocostos,
                puestonuevo=puestonuevo,
                incapacidad=incapacidad,
                reemplazo=reemplazo,
                tiempoprimercontrato=tiempoprimercontrato,
                tipo_contrato=tipo_contrato,
                motivo=motivo,
                nombrereemplazar=nombrereemplazar,
                formacionacademica=formacionacademica,
                habilidadesferreteras=habilidadesferreteras,
                habilidadesinformaticas=habilidadesinformaticas,
                habilidadespersonales=habilidadespersonales,
                habilidadesanaliticas=habilidadesanaliticas,
                materialesequipo=materialesequipo,
                funciones_cargo=funciones_cargo,
                fechaRecepcion=None,
                usuario_creo=request.user,
                estado="EN PROCESO"
            )

            return JsonResponse({'success': True, 'message': 'Requisa registrada correctamente'})

        except Exception as e:
            # Si ocurre algún error, devolvemos una respuesta de error
            return JsonResponse({'success': False, 'message': str(e)})

def aprobar_requisa_view(request, id):
    if request.method == 'POST':
        try:
            requisa = get_object_or_404(Requisa, id=id)
            requisa.estado = 'APROBADO'
            requisa.save()
            
            return JsonResponse({'success': True, 'message': 'Requisa aprobada correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

def cancelar_requisa_view(request, id):
    if request.method == 'POST':
        try:
            requisa = get_object_or_404(Requisa, id=id)
            requisa.estado = 'CANCELADA'
            requisa.save()
            
            return JsonResponse({'success': True, 'message': 'Requisa cancelada correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

def duplicar_requisa_view(request, id):
    if request.method == 'POST':
        try:
            # Obtener la requisición original
            requisa_original = get_object_or_404(Requisa, id=id)
            
            # Crear una nueva requisición duplicando la original (omitimos el campo id)
            nueva_requisa = Requisa.objects.create(
                salario_base=requisa_original.salario_base,
                plan_de_compensacion=requisa_original.plan_de_compensacion,
                montos=requisa_original.montos,
                puesto=requisa_original.puesto,
                sucursal=requisa_original.sucursal,
                departamento=requisa_original.departamento,
                hora_inicio=requisa_original.hora_inicio,
                hora_fin=requisa_original.hora_fin,
                centrocostos=requisa_original.centrocostos,
                puestonuevo=requisa_original.puestonuevo,
                incapacidad=requisa_original.incapacidad,
                reemplazo=requisa_original.reemplazo,
                tiempoprimercontrato=requisa_original.tiempoprimercontrato,
                tipo_contrato=requisa_original.tipo_contrato,
                motivo=requisa_original.motivo,
                nombrereemplazar=requisa_original.nombrereemplazar,
                formacionacademica=requisa_original.formacionacademica,
                habilidadesferreteras=requisa_original.habilidadesferreteras,
                habilidadesinformaticas=requisa_original.habilidadesinformaticas,
                habilidadespersonales=requisa_original.habilidadespersonales,
                habilidadesanaliticas=requisa_original.habilidadesanaliticas,
                materialesequipo=requisa_original.materialesequipo,
                funciones_cargo=requisa_original.funciones_cargo,
                estado="EN PROCESO",
                usuario_creo=request.user,
            )
            
            return JsonResponse({'success': True, 'message': 'Requisa duplicada correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
        
def requisaupdate_view(request, id):
    if request.method == 'GET':
        requisa = get_object_or_404(Requisa, id=id)
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
        all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(activo=True).order_by('nombre_tipo_de_contrato').distinct()

        context = {
            'all_departamentos': all_departamentos,
            'all_sucursales': all_sucursales,
            'all_tipos_contrato': all_tipos_contrato,
            'requisa': requisa,
        }
        return render(request, 'requisas/updaterequisa.html', context)

    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.change_requisa'):
                return JsonResponse({'error': 'No tienes permisos para editar requisas'}, status=403)
            requisa = get_object_or_404(Requisa, id=id)

            # Recibir los datos enviados desde el formulario con FormData
            salario_base = request.POST.get('salario_base')
            fecharecepcion = request.POST.get('fecharecepcion')

            # Decodificar el JSON enviado con los valores de plan_de_compensacion y montos
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]'))
            montos = json.loads(request.POST.get('montos', '{}'))

            puesto = request.POST.get('puesto')
            sucursal_id = request.POST.get('sucursal')
            departamento_id = request.POST.get('departamento')
            funciones_cargo = request.POST.get('funciones_cargo')

            hora_inicio = request.POST.get('hora_inicio')
            hora_fin = request.POST.get('hora_fin')

            centrocostos = request.POST.get('centrocostos')
            puestonuevo = bool(int(request.POST.get('puestonuevo', 0)))
            incapacidad = bool(int(request.POST.get('incapacidad', 0)))
            reemplazo = bool(int(request.POST.get('reemplazo', 0)))

            tiempoprimercontrato = request.POST.get('tiempoprimercontrato')
            tipo_contrato_id = request.POST.get('tipo_contrato')
            motivo = request.POST.get('motivo')
            nombrereemplazar = request.POST.get('nombrereemplazar')
            formacionacademica = request.POST.get('formacionacademica')

            # Decodificar los JSON de habilidades
            habilidadesferreteras = json.loads(request.POST.get('habilidadesferreteras', '[]'))
            habilidadesinformaticas = json.loads(request.POST.get('habilidadesinformaticas', '[]'))
            habilidadespersonales = json.loads(request.POST.get('habilidadespersonales', '[]'))
            habilidadesanaliticas = json.loads(request.POST.get('habilidadesanaliticas', '[]'))
            materialesequipo = json.loads(request.POST.get('materialesequipo', '[]'))

            # Actualizar los datos de la Requisa
            requisa.salario_base = salario_base
            requisa.fechaRecepcion = fecharecepcion if fecharecepcion else None
            requisa.plan_de_compensacion = plan_de_compensacion
            requisa.montos = montos
            requisa.puesto = puesto
            requisa.sucursal_id = sucursal_id if sucursal_id else None
            requisa.departamento_id = departamento_id if departamento_id else None
            requisa.hora_inicio = hora_inicio
            requisa.hora_fin = hora_fin
            requisa.centrocostos = centrocostos
            requisa.puestonuevo = puestonuevo
            requisa.incapacidad = incapacidad
            requisa.reemplazo = reemplazo
            requisa.tiempoprimercontrato = tiempoprimercontrato
            requisa.tipo_contrato_id = tipo_contrato_id if tipo_contrato_id else None
            requisa.motivo = motivo
            requisa.nombrereemplazar = nombrereemplazar
            requisa.formacionacademica = formacionacademica
            requisa.habilidadesferreteras = habilidadesferreteras
            requisa.habilidadesinformaticas = habilidadesinformaticas
            requisa.habilidadespersonales = habilidadespersonales
            requisa.habilidadesanaliticas = habilidadesanaliticas
            requisa.materialesequipo = materialesequipo
            requisa.funciones_cargo = funciones_cargo
            requisa.usuario_modifico = request.user
            # Guardar la requisa actualizada
            requisa.save()

            return JsonResponse({'success': True, 'message': 'Requisa actualizada correctamente'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

def imprimirrequisa_view(request, id):
    if request.method == 'GET':
        requisa = get_object_or_404(Requisa, id=id)
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
        all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(activo=True).order_by('nombre_tipo_de_contrato').distinct()

        context = {
            'all_departamentos': all_departamentos,
            'all_sucursales': all_sucursales,
            'all_tipos_contrato': all_tipos_contrato,
            'requisa': requisa,
        }
        return render(request, 'requisas/imprimirrequisa.html', context)


#--------CONTRATACIONES MULTI--------#  
def contratacionesform_multi_view(request):
    if request.method == 'GET':
        all_sucursales = Sucursal.objects.filter(activo=True).order_by('nombre_sucursal').distinct()


        all_municipios = MunicipioPais.objects.all().order_by('nombre_municipio').distinct()
        all_departamentos_hn = DepartamentoPais.objects.all().order_by('nombre_departamento').distinct()
        all_paises = Pais.objects.all().order_by('nombre_pais').distinct()

        context = {
            'all_sucursales': all_sucursales,
            'all_municipios': all_municipios, 
            'all_paises': all_paises,
            'all_departamentos_hn': all_departamentos_hn,
        }
        return render(request, 'contrataciones_multi/registrarcontrataciones_multi.html', context)
    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_contratacionempleadosmulti'):
                return JsonResponse({'error': 'No tienes permisos para registrar registros'}, status=403)
            tipo_contratacion = request.POST.get('tipoIngreso') or None
            primer_nombre = request.POST.get('primerNombre') or None
            segundo_nombre = request.POST.get('segundoNombre') or None
            primer_apellido = request.POST.get('primerApellido') or None
            segundo_apellido = request.POST.get('segundoApellido') or None
            fecha_nacimiento = request.POST.get('fechaNacimiento') or None
            municipio = request.POST.get('municipio') or None
            genero = request.POST.get('genero') or None
            direccion = request.POST.get('direccion') or None
            dni = request.POST.get('cedula') or None
            estado_civil = request.POST.get('estadoCivil') or None
            profesion = request.POST.get('profesion') or None
            departamento = request.POST.get('departamento') or None
            telefono = request.POST.get('telefono') or None
            ultimo_grado_estudio = request.POST.get('ultimoGrado') or None
            pais = request.POST.get('pais') or None

            # Datos de emergencia
            emergencia1 = request.POST.get('emergencia1') or None
            parentesco1 = request.POST.get('parentesco1') or None
            telefono_emergencia1 = request.POST.get('telefonoEmergencia1') or None
            enfermedad = request.POST.get('enfermedad') or None

            # Datos laborales
            puesto = request.POST.get('puesto') or None
            sucursal = request.POST.get('sucursal') or None
            salario = request.POST.get('salario') or None
            fecha_ingreso = request.POST.get('fechaIngreso') or None

            # Beneficiario
            nombre_beneficiario = request.POST.get('nombreBeneficiario') or None
            identidad_beneficiario = request.POST.get('identidadBeneficiario') or None
            parentesco_beneficiario = request.POST.get('parentescoBeneficiario') or None
            porcentaje_beneficiario = request.POST.get('porcentajeBeneficiario') or None

            # Imagen
            imagen = request.FILES.get('imagen')
            nombre_imagen = None
            ruta_imagen = None

            if imagen:
                # Si hay imagen, generar un nombre único para el archivo
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/contrataciones/contrataciones_multi/{nombre_imagen}"

                # Crear directorio si no existe
                try:
                    os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                    # Guardar la imagen en la ruta especificada
                    with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                        for chunk in imagen.chunks():
                            destino.write(chunk)
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Error al guardar la imagen: {str(e)}'}, status=500)

            # Crear nueva entrada en ContratacionEmpleadosmulti
            contratacion = ContratacionEmpleadosmulti.objects.create(
                tipo_contratacion=tipo_contratacion,
                nombre1=primer_nombre,
                nombre2=segundo_nombre,
                apellido1=primer_apellido,
                apellido2=segundo_apellido,
                fecha_nacimiento=fecha_nacimiento,
                municipio_id=municipio,
                genero=genero,
                direccionexacta=direccion,
                dni=dni,
                estado_civil=estado_civil,
                profecion_oficio=profesion,
                departamento_id=departamento,
                telefono=telefono,
                nombre_emergencia=emergencia1,
                parentesco=parentesco1,
                telefonoemergencia=telefono_emergencia1,
                enfermedad=enfermedad,
                puestos=puesto,
                sucursal_id=sucursal,
                salario=salario,
                fecha_ingreso=fecha_ingreso,
                ultimo_grado_estudio=ultimo_grado_estudio,
                nombre_beneficiario=nombre_beneficiario,
                dni_beneficiario=identidad_beneficiario,
                parentesco_beneficiario=parentesco_beneficiario,
                porcentaje=porcentaje_beneficiario,
                ruta=ruta_imagen,
                nombreimagen=nombre_imagen,
                pais_id=pais,
                usuario_creo=request.user,
            )

            return JsonResponse({'success': True, 'message': 'Candidato registrado exitosamente'}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
                          
def contratacionesmulti_view(request, id=None):
    if request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_contratacionempleadosmulti'):
            return redirect('sinacceso')
        search = request.GET.get('search', '')

        # Consulta los datos del nuevo modelo ContratacionEmpleadosmulti
        contrataciones_multi = ContratacionEmpleadosmulti.objects.select_related(
            'sucursal', 
            'municipio',  # Relaciona con el modelo Municipio para el filtro de ciudad
        )

        # Obtener valores únicos para los filtros
        municipios_options = set(contrataciones_multi.values_list('municipio__nombre_municipio', flat=True))
        puestos_options = set(contrataciones_multi.values_list('puestos', flat=True))
        sucursales_options = set(contrataciones_multi.values_list('sucursal__nombre_sucursal', flat=True))
        fechas_ingreso_options = set(contrataciones_multi.values_list('fecha_ingreso', flat=True))

        # Filtrar por los filtros seleccionados (si se aplican)
        municipio_filter = request.GET.getlist('municipio')
        puesto_filter = request.GET.getlist('puesto')
        sucursal_filter = request.GET.getlist('sucursal')
        fecha_ingreso_filter = request.GET.getlist('fecha_ingreso')

        # Aplicar filtros a la consulta solo si existen
        if municipio_filter:
            contrataciones_multi = contrataciones_multi.filter(municipio__nombre_municipio__in=municipio_filter)
        
        if puesto_filter:
            contrataciones_multi = contrataciones_multi.filter(puestos__in=puesto_filter)
        
        if sucursal_filter:
            contrataciones_multi = contrataciones_multi.filter(sucursal__nombre_sucursal__in=sucursal_filter)

        if fecha_ingreso_filter:
            contrataciones_multi = contrataciones_multi.filter(fecha_ingreso__in=fecha_ingreso_filter)

        # Filtro de búsqueda general
        if search:
            contrataciones_multi = contrataciones_multi.filter(
                models.Q(nombre1__icontains=search) |
                models.Q(nombre2__icontains=search) |
                models.Q(apellido1__icontains=search) |
                models.Q(apellido2__icontains=search) |
                models.Q(telefono__icontains=search) |
                models.Q(dni__icontains=search) |
                models.Q(correo__icontains=search)
            )

        # Paginación
        paginator = Paginator(contrataciones_multi, 10)  # 10 registros por página
        page_number = request.GET.get('page')  # Capturar el número de página actual
        page_obj = paginator.get_page(page_number)  # Obtener la página solicitada

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear un rango de páginas a mostrar

        # Contexto para la plantilla
        context = {
            'contrataciones': page_obj,  # Pasar el objeto paginado
            'search': search,  # Pasar el valor de búsqueda si aplica
            'page_range': page_range,  # Pasar el rango de páginas
            'municipios_options': list(municipios_options),
            'puestos_options': list(puestos_options),
            'sucursales_options': list(sucursales_options),
            'fechas_ingreso_options': list(fechas_ingreso_options),
            'municipio_filter': municipio_filter,
            'puesto_filter': puesto_filter,
            'sucursal_filter': sucursal_filter,
            'fecha_ingreso_filter': fecha_ingreso_filter,
        }

        return render(request, 'contrataciones_multi/contrataciones_multi.html', context)

def imprimir_contratacion_multi_view(request, contratacion_id):
    contratacion_multi = get_object_or_404(ContratacionEmpleadosmulti, id=contratacion_id)
    
    context = {
        'contratacion_multi': contratacion_multi,
    }
    return render(request, 'contrataciones_multi/imprimir_contratacion_multi.html', context)

def updatecontrataciones_multi_view(request, contratacion_id):
    contratacion_multi = get_object_or_404(ContratacionEmpleadosmulti, id=contratacion_id)
    if request.method == 'GET':
        context = {
            'contratacion_multi': contratacion_multi,
            'all_municipios': MunicipioPais.objects.all(),
            'all_departamentos': DepartamentoPais.objects.all(),
            'all_sucursales': Sucursal.objects.all(),
            'all_paises': Pais.objects.all(),
        }
        return render(request, 'contrataciones_multi/updatecontrataciones_multi.html', context)

    if request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.change_contratacionempleadosmulti'):
                return JsonResponse({'error': 'No tienes permisos para editar registros'}, status=403)
            primer_nombre = request.POST.get('nombre1') or None
            segundo_nombre = request.POST.get('nombre2') or None
            primer_apellido = request.POST.get('apellido1') or None
            segundo_apellido = request.POST.get('apellido2') or None
            fecha_nacimiento = request.POST.get('fechaNacimiento') or None

            # Claves foráneas deben ser números o None
            municipio = request.POST.get('municipio')
            municipio_id = int(municipio) if municipio and municipio.isdigit() else None

            genero = request.POST.get('genero') or None
            direccion = request.POST.get('direccionexacta') or None
            dni = request.POST.get('dni') or None
            estado_civil = request.POST.get('estadoCivil') or None
            profesion = request.POST.get('profecion_oficio') or None
            telefono = request.POST.get('telefono') or None

            # Datos laborales
            puesto = request.POST.get('puestos')
            sucursal = request.POST.get('sucursal')
            sucursal_id = int(sucursal) if sucursal and sucursal.isdigit() else None
            salario = request.POST.get('salario') or None
            fecha_ingreso = request.POST.get('fechaIngreso') or None

            # Datos de emergencia
            nombre_emergencia = request.POST.get('emergencia1') or None
            parentesco_emergencia = request.POST.get('parentesco1') or None
            telefono_emergencia = request.POST.get('telefonoEmergencia1') or None

            # Beneficiario
            nombre_beneficiario = request.POST.get('nombre_beneficiario') or None
            identidad_beneficiario = request.POST.get('dni_beneficiario') or None
            parentesco_beneficiario = request.POST.get('parentesco_beneficiario') or None
            porcentaje_beneficiario = request.POST.get('porcentajeBeneficiario') or None
            pais = request.POST.get('pais') or None

            # Imagen
            imagen = request.FILES.get('imagen')
            nombre_imagen = None
            ruta_imagen = None

            if imagen:
                # Verificar si existe una imagen previa en el campo `ruta`
                if contratacion_multi.ruta:
                    ruta_anterior = os.path.join(settings.BASE_DIR, contratacion_multi.ruta)
                    
                    # Comprobar si la ruta anterior realmente existe en el sistema de archivos
                    if os.path.exists(ruta_anterior):
                        try:
                            os.remove(ruta_anterior)  # Eliminar la imagen anterior
                            print(f"Imagen anterior eliminada: {ruta_anterior}")  # Log para depuración
                        except Exception as e:
                            print(f"Error al eliminar la imagen anterior: {str(e)}")
                            return JsonResponse({'success': False, 'message': f'Error al eliminar la imagen anterior: {str(e)}'}, status=500)
                    else:
                        print(f"La ruta de la imagen anterior no existe: {ruta_anterior}")  # Log para depuración

                # Guardar la nueva imagen
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/contrataciones/contrataciones_multi/{nombre_imagen}"
                os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)
                
                # Guardar el archivo en la nueva ruta
                with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                    for chunk in imagen.chunks():
                        destino.write(chunk)

                # Actualizar el modelo con la ruta y el nombre de la nueva imagen
                contratacion_multi.ruta = ruta_imagen
                contratacion_multi.nombreimagen = nombre_imagen

            # Actualizar el modelo con otros campos
            contratacion_multi.nombre1 = primer_nombre
            contratacion_multi.nombre2 = segundo_nombre
            contratacion_multi.apellido1 = primer_apellido
            contratacion_multi.apellido2 = segundo_apellido
            contratacion_multi.fecha_nacimiento = fecha_nacimiento
            contratacion_multi.municipio_id = municipio_id
            contratacion_multi.genero = genero
            contratacion_multi.direccionexacta = direccion
            contratacion_multi.dni = dni
            contratacion_multi.estado_civil = estado_civil
            contratacion_multi.profecion_oficio = profesion
            contratacion_multi.telefono = telefono
            contratacion_multi.puestos = puesto
            contratacion_multi.sucursal_id = sucursal_id
            contratacion_multi.fecha_ingreso = fecha_ingreso
            contratacion_multi.salario = salario
            contratacion_multi.pais_id = pais
            contratacion_multi.usuario_modifico = request.user

            # Datos de emergencia
            contratacion_multi.nombre_emergencia = nombre_emergencia
            contratacion_multi.parentesco = parentesco_emergencia
            contratacion_multi.telefonoemergencia = telefono_emergencia

            # Beneficiario
            contratacion_multi.nombre_beneficiario = nombre_beneficiario
            contratacion_multi.dni_beneficiario = identidad_beneficiario
            contratacion_multi.parentesco_beneficiario = parentesco_beneficiario
            contratacion_multi.porcentaje = porcentaje_beneficiario
            contratacion_multi.usuario_modifico = request.user 

            # Guardar los cambios
            contratacion_multi.save()

            return JsonResponse({'success': True, 'message': 'Candidato actualizado exitosamente'}, status=200)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

class ExportExcelContratacionMulti(View):
    def get(self, request, *args, **kwargs):
        template_path = os.path.join('reclutamiento/static/templates/FormatoContratacion_multi.xlsx')
        
        # Cargar la plantilla de Excel
        try:
            wb = load_workbook(template_path)
            sheet = wb.active  # Ajusta si la hoja tiene un nombre específico
        except Exception as e:
            return HttpResponse(f"Error al cargar la plantilla de Excel: {e}", status=500)

        # Obtener los datos de ContratacionEmpleadosmulti
        contrataciones = ContratacionEmpleadosmulti.objects.all().order_by('-id')
        
        # Definir la celda de inicio (C5)
        start_row = 5
        start_col = 3  # Columna C es la columna 3

        # Escribir los datos en la hoja de Excel
        for idx, contratacion in enumerate(contrataciones, start=start_row):
            try:
                sheet.cell(row=idx, column=start_col, value=contratacion.id)
                sheet.cell(row=idx, column=start_col + 1, value=contratacion.tipo_contratacion)
                sheet.cell(row=idx, column=start_col + 2, value=contratacion.nombre1)
                sheet.cell(row=idx, column=start_col + 3, value=contratacion.nombre2)
                sheet.cell(row=idx, column=start_col + 4, value=contratacion.apellido1)
                sheet.cell(row=idx, column=start_col + 5, value=contratacion.apellido2)
                sheet.cell(row=idx, column=start_col + 6, value=contratacion.fecha_nacimiento)
                
                # Departamento Honduras
                sheet.cell(row=idx, column=start_col + 7, value=contratacion.pais.nombre_pais if contratacion.pais else '')
                sheet.cell(row=idx, column=start_col + 8, value=getattr(contratacion.departamento, 'nombre_departamento', '') if contratacion.departamento else '')
                # Municipio Honduras
                sheet.cell(row=idx, column=start_col + 9, value=contratacion.municipio.nombre_municipio if contratacion.municipio else '')
                
                sheet.cell(row=idx, column=start_col + 10, value=contratacion.genero)
                sheet.cell(row=idx, column=start_col + 11, value=contratacion.dni)
                sheet.cell(row=idx, column=start_col + 12, value=contratacion.telefono)
                sheet.cell(row=idx, column=start_col + 13, value=contratacion.profecion_oficio)
                sheet.cell(row=idx, column=start_col + 14, value=contratacion.estado_civil)

                
                # Datos de contacto de emergencia
                sheet.cell(row=idx, column=start_col + 15, value=contratacion.nombre_emergencia)
                sheet.cell(row=idx, column=start_col + 16, value=contratacion.parentesco)
                sheet.cell(row=idx, column=start_col + 17, value=contratacion.telefonoemergencia)
                
                # Enfermedad
                sheet.cell(row=idx, column=start_col + 18, value=contratacion.enfermedad)
                
                # Dirección exacta y último grado de estudio
                sheet.cell(row=idx, column=start_col + 19, value=contratacion.direccionexacta)
                sheet.cell(row=idx, column=start_col + 20, value=contratacion.ultimo_grado_estudio)
                
                # Puestos y salario
                sheet.cell(row=idx, column=start_col + 21, value=contratacion.puestos)
                sheet.cell(row=idx, column=start_col + 22, value=contratacion.salario)
                
                # Sucursal y fecha de ingreso
                sheet.cell(row=idx, column=start_col + 23, value=contratacion.sucursal.nombre_sucursal if contratacion.sucursal else '')
                sheet.cell(row=idx, column=start_col + 24, value=contratacion.fecha_ingreso)
                
                # Beneficiario
                sheet.cell(row=idx, column=start_col + 25, value=contratacion.nombre_beneficiario)
                sheet.cell(row=idx, column=start_col + 26, value=contratacion.dni_beneficiario)
                sheet.cell(row=idx, column=start_col + 27, value=contratacion.parentesco_beneficiario)
                sheet.cell(row=idx, column=start_col + 28, value=contratacion.porcentaje)
            except Exception as e:
                return HttpResponse(f"Error escribiendo datos en Excel en la fila {idx + 1}: {e}", status=500)

        # Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="FormatoContratacion_multi.xlsx"'

        # Guardar el libro de trabajo en la respuesta
        wb.save(response)

        return response

#--------TELEFONIA--------#
def telefonosinventario_view(request, id=None):
    if request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_inventariotelefonos'):
            return redirect('sinacceso')
        all_nombres_telefono = Inventariotelefonos.objects.values_list('nombretelefono', flat=True).distinct()

        search = request.GET.get('search', '')
        nombres_telefono = request.GET.getlist('nombretelefono', [])

        # Obtener todos los inventarios
        inventarios = Inventariotelefonos.objects.all()

        # Filtrar por búsqueda
        if search:
            inventarios = inventarios.filter(
                Q(nombretelefono__icontains=search) |
                Q(correlativo__icontains=search) 
            )

        # Filtrar por nombres de teléfono seleccionados
        if nombres_telefono:
            inventarios = inventarios.filter(nombretelefono__in=nombres_telefono)

        # Formatear el valor total con comas para miles y punto para decimales
        for inventario in inventarios:
            inventario.valortotal_formateado = "{:,.2f}".format(inventario.valortotal)

        # Paginación
        paginator = Paginator(inventarios, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'inventarios': page_obj,
            'search': search,
            'page_range': page_range,
            'all_nombres_telefono': all_nombres_telefono,
            'nombres_telefono': nombres_telefono,
        }

        # Crear una lista de inventarios para el script JSON
        context['inventarios_json'] = [
            {
                'id': inventario.id,
                'correlativo': inventario.correlativo,
                'nombretelefono': inventario.nombretelefono,
                'valortotal': inventario.valortotal,
                'estadotelefono': inventario.estadotelefono,
                'estado': inventario.estado,
            }
            for inventario in inventarios
        ]

        return render(request, 'inventario/inventariotelefono.html', context)


    elif request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_inventariotelefonos'):
                return JsonResponse({'error': 'No tienes permisos para crear registros'}, status=403)
            # Parsear los datos JSON recibidos desde el fetch
            data = json.loads(request.body)
            nombre_telefono = data.get('nombretelefono')
            valor_telefono = float(data.get('valortelefono', 0))
            cantidad = int(data.get('cantidad', 1))  # Cantidad de registros a crear
            estado = data.get('estado', 'REGISTRADO')
            estadotelefono = data.get('estadotelefono')

            # Validación básica
            if not nombre_telefono or valor_telefono <= 0 or cantidad <= 0:
                return JsonResponse({'success': False, 'message': 'Datos no válidos.'}, status=400)
            
            # Crear los registros según la cantidad especificada
            for _ in range(cantidad):
                Inventariotelefonos.objects.create(
                    nombretelefono=nombre_telefono,
                    estado=estado,
                    valortotal=valor_telefono,
                    estadotelefono = estadotelefono,
                    usuario_creo=request.user,
                )
            
            return JsonResponse({'success': True, 'message': f'{cantidad} teléfonos registrados exitosamente.'}, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Error al procesar la solicitud.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        
    elif request.method == 'DELETE':
        if id:
            telefono = get_object_or_404(Inventariotelefonos, id=id)
            telefono.delete()
            return JsonResponse({'success': True, 'message': 'Teléfono eliminado con éxito.'})

        return JsonResponse({'success': False, 'message': 'ID de teléfono no encontrado.'})
        
    elif request.method == 'PUT':
        telefono_id = id  # Asegúrate de que el ID se pase correctamente a la vista
        telefono = get_object_or_404(Inventariotelefonos, id=telefono_id)

        # Obtener los datos del cuerpo de la solicitud
        try:
            if not request.user.has_perm('Reclutamiento.change_inventariotelefonos'):
                return JsonResponse({'error': 'No tienes permisos para editar registros'}, status=403)
            data = json.loads(request.body)
            valortelefono = data.get('valortelefono')
            estadotelefono = data.get('estadotelefono')
            estado = data.get('estado')

            # Actualizar los campos del objeto
            if valortelefono is not None:
                telefono.valortotal = float(valortelefono)  # Asegúrate de convertir a float

            if estadotelefono is not None:
                telefono.estadotelefono = estadotelefono

            if estado is not None:
                # Si el nuevo estado es "REGISTRADO" y el anterior era "ASIGNADO" o "REASIGNADO", actualizar correlativo a null
                if estado == "REGISTRADO" and telefono.estado in ["ASIGNADO", "REASIGNADO"]:
                    telefono.correlativo = None  # Establecer correlativo a null
                telefono.estado = estado

            telefono.usuario_modifico = request.user

            # Guardar los cambios en la base de datos
            telefono.save()

            return JsonResponse({'success': True, 'message': 'Teléfono actualizado con éxito.'})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Error al decodificar los datos JSON.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Método no permitido.'})

def obtener_id_telefono(request):
    nombre_telefono = request.GET.get('nombretelefono', None)
    estadotelefono = request.GET.get('estadotelefono', None)  # Nuevo parámetro de estado de teléfono

    if nombre_telefono and estadotelefono:
        # Filtrar por nombre, estado 'REGISTRADO' y estado del teléfono ('NUEVO' o 'USADO')
        telefono = (Inventariotelefonos.objects
                    .filter(nombretelefono=nombre_telefono, estado='REGISTRADO', estadotelefono=estadotelefono)
                    .order_by('id')
                    .first())

        if telefono:
            return JsonResponse({'success': True, 'id': telefono.id})
        else:
            return JsonResponse({'success': False, 'message': 'No se encontró el teléfono en el inventario.'}, status=404)

    return JsonResponse({'success': False, 'message': 'Parámetros nombretelefono o estadotelefono no proporcionados.'}, status=400)

def telefonos_view(request):
    if request.method == 'GET':
        if not request.user.has_perm('Reclutamiento.view_telefonia'):
            return redirect('sinacceso')
        all_unidades_negocio = Unidad_Negocio.objects.filter(activo=True).order_by('nombre_unidad_de_negocio').distinct()
        all_departamentos = Departamento.objects.filter(activo=True).order_by('nombre_departamento').distinct()
        all_telefonosinventario = (Inventariotelefonos.objects
                                .filter(estado='REGISTRADO')
                                .values('nombretelefono', 'estadotelefono')
                                .distinct()
                                .order_by('nombretelefono', 'estadotelefono'))

        # Search and filter functionality
        search = request.GET.get('search', '')
        telefonia = Telefonia.objects.all()

        if search:
            try:
                # Intentar parsear la fecha si se busca en el formato día/mes/año
                search_date = datetime.strptime(search, '%d/%m/%Y').date()
                telefonia = telefonia.filter(fecha=search_date)
            except ValueError:
                # Si no es una fecha válida, buscar en los demás campos
                telefonia = telefonia.filter(
                    Q(correlativo__icontains=search) |
                    Q(nombre__icontains=search) |
                    Q(dni__icontains=search) |
                    Q(unidad_de_negocio__nombre_unidad_de_negocio__icontains=search) |
                    Q(departamento__nombre_departamento__icontains=search) |
                    Q(nombretelefono__icontains=search)
                )

        # Paginación
        paginator = Paginator(telefonia, 10)  # 10 registros por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Rango de paginación
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'all_unidades_negocio': all_unidades_negocio,
            'all_departamentos': all_departamentos,
            'all_telefonosinventario': all_telefonosinventario,
            'telefonia': page_obj,  # Pasamos los datos paginados
            'search': search,
            'page_range': page_range
        }
        return render(request, 'inventario/telefonos.html', context)

    elif request.method == 'POST':
        try:
            if not request.user.has_perm('Reclutamiento.add_telefonia'):
                return JsonResponse({'error': 'No tienes permisos para crear registros'}, status=403)
            data = json.loads(request.body)
            correlativo = data.get('correlativo') 
            fecha = data.get('fecha')
            nombre = data.get('nombre')
            dni = data.get('dni')
            MEI = data.get('MEI')
            unidad_negocio_id = data.get('unidadnegocio')
            departamento_id = data.get('departamento')
            telefono_id = data.get('telefono') 
            lineatelefonica = data.get('lineatelefono')
            caracter = data.get('caracter')
            tiempopago = data.get('tiempopago')
            quincenas = data.get('quincenas')
            observacion = data.get('observacion')

            # Obtén el objeto `Inventariotelefonos` para el `telefono_id`
            inventario_telefono1 = Inventariotelefonos.objects.filter(id=telefono_id, estado='REGISTRADO').first()
            
            # Verifica que el teléfono existe en el inventario y tiene el estado `REGISTRADO`
            if not inventario_telefono1:
                return JsonResponse({'success': False, 'message': 'No se encontró el teléfono en el inventario.'}, status=404)
            
            valor_total = inventario_telefono1.valortotal
            estado_telefono = inventario_telefono1.estadotelefono  # Obtiene el `estadotelefono` desde `Inventariotelefonos`

            # Validación de campos obligatorios
            if not nombre or not dni or not unidad_negocio_id or not departamento_id or not telefono_id or not caracter:
                return JsonResponse({'success': False, 'message': 'Faltan campos obligatorios.'}, status=400)
            
            # Crear el registro en `Telefonia` con `estadotelefono`
            telefonia = Telefonia.objects.create(
                correlativo=correlativo, 
                fecha=fecha,
                nombre=nombre,
                dni=dni,
                unidad_de_negocio_id=unidad_negocio_id,
                departamento_id=departamento_id,
                nombretelefono=telefono_id,
                MEI=MEI,
                lineatelefonica=lineatelefonica,
                caracter=caracter,
                valor=valor_total,
                tiempopago=tiempopago,
                quinsena=quincenas,
                estado='ASIGNADO',
                observacion=observacion,
                estadotelefono=estado_telefono,
                usuario_creo=request.user,
            )

            # Actualizar el estado en `Inventariotelefonos`
            inventario_telefono1.estado = 'ASIGNADO'
            inventario_telefono1.correlativo = correlativo
            inventario_telefono1.save()

            return JsonResponse({'success': True, 'message': 'Registro de telefonía y actualización de inventario exitosos.'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Error en los datos enviados.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

def imprimir_telefono_view(request, id):
    telefonia = get_object_or_404(Telefonia, id=id)
    inventario_telefono = get_object_or_404(Inventariotelefonos, id=telefonia.nombretelefono)

    telefonia.valor_formateado = "{:,.2f}".format(telefonia.valor) if telefonia.valor else None
    if telefonia.quinsena:
        # Verificar si el número es entero o tiene decimales
        if telefonia.quinsena.is_integer():
            telefonia.quincena_formateada = "{:,.0f}".format(telefonia.quinsena)  # Sin decimales
        else:
            telefonia.quincena_formateada = "{:,.2f}".format(telefonia.quinsena)  # Con dos decimales
    else:
        telefonia.quincena_formateada = None

    context = {
        'telefonia': telefonia,
        'nombre_telefono': inventario_telefono.nombretelefono,
    }
    return render(request, 'inventario/impimirtelefonos.html', context)

@csrf_exempt
def telefonosreasignar_view(request, correlativo):
    if request.method == 'PUT':
        try:
            data = json.loads(request.body.decode('utf-8'))

            telefonia = get_object_or_404(Telefonia, correlativo=correlativo)

            # Actualizar los campos en Telefonia
            telefonia.asignacionnueva = data.get('asignacionnueva')
            telefonia.estado = 'REASIGNADO'
            telefonia.save()

            # Buscar el registro en Inventariotelefonos por correlativo
            inventario_telefono = Inventariotelefonos.objects.filter(correlativo=correlativo, estado='ASIGNADO').first()

            if inventario_telefono:
                inventario_telefono.estado = 'REASIGNADO'
                inventario_telefono.save()
            else:
                return JsonResponse({'success': False, 'message': 'No se encontró el teléfono en el inventario.'}, status=404)

            return JsonResponse({'success': True, 'message': 'Registro reasignado correctamente'})

        except Telefonia.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Correlativo no encontrado'}, status=404)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

def deduccion_telefono(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        correlativo = data.get('correlativo')
        fecha_extravio = data.get('fecha_extravio')

        if not correlativo or not fecha_extravio:
            return JsonResponse({'success': False, 'message': 'Datos incompletos.'})

        telefono_inventario = get_object_or_404(Inventariotelefonos, correlativo=correlativo)
        telefono_asignacion = get_object_or_404(Telefonia, correlativo=correlativo)

        valor_total = telefono_inventario.valortotal
        fecha_asignacion = telefono_asignacion.fecha
        fecha_extravio_dt = datetime.strptime(fecha_extravio, '%Y-%m-%d').date()

        # Calcular los meses de uso hasta la fecha de extravío
        meses_uso = (fecha_extravio_dt.year - fecha_asignacion.year) * 12 + (fecha_extravio_dt.month - fecha_asignacion.month) + 1

        # Definir tasa de depreciación y calcular el valor de depreciación mensual
        tasa_depreciacion_mensual = 5.2632 / 100
        depreciacion_mensual = round(valor_total * tasa_depreciacion_mensual)

        # Calcular el tiempo de pago restante y el valor mensual a pagar
        vida_util_meses = 18
        tiempo_pago_restante = vida_util_meses - meses_uso
        valor_mensual_a_pagar = depreciacion_mensual

        # Actualizar el estado en inventario y telefonía, y valores de deducción
        telefono_inventario.estado = 'DEDUCIDO'
        telefono_inventario.save()

        telefono_asignacion.estado = 'DEDUCIDO'
        telefono_asignacion.fechaextravio = fecha_extravio_dt
        telefono_asignacion.tiempopago = tiempo_pago_restante
        telefono_asignacion.valorquincena = valor_mensual_a_pagar
        telefono_asignacion.save()

        return JsonResponse({
            'success': True,
            'message': 'Deducción calculada y guardada correctamente',
            'valor_mensual_a_pagar': valor_mensual_a_pagar,
            'tiempo_pago_restante': tiempo_pago_restante
        })
    return JsonResponse({'success': False, 'message': 'Método no permitido'})

def exportar_inventario_telefonos(request):
    # Obtener todos los registros de Inventariotelefonos
    datos = (
        Inventariotelefonos.objects
        .values(
            'id', 
            'nombretelefono', 
            'correlativo', 
            'estado', 
            'estadotelefono', 
            'valortotal', 
            'creado'  # Incluye la fecha de creación
        )
    )

    # Convertir los datos a un DataFrame de pandas
    df = pd.DataFrame(list(datos))

    # Formatear la fecha de creación para que solo muestre la fecha
    df['creado'] = pd.to_datetime(df['creado']).dt.date  # Convertir a solo fecha

    # Crear un archivo Excel
    excel_file = "inventario_telefonos.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={excel_file}'

    # Escribir el DataFrame en el archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='InventarioTelefonos')

    return response
